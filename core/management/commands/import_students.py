from django.core.management.base import BaseCommand
from core.models import Student, School
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook


class Command(BaseCommand):
    help = "Import students from Excel"

    def add_arguments(self, parser):
        parser.add_argument("file", type=str, help="Path to Excel file")

    def handle(self, *args, **options):
        file_path = options["file"]

        wb = load_workbook(file_path)
        ws = wb.active

        headers = []
        created, skipped = 0, 0

        with transaction.atomic():
            for row_index, row in enumerate(ws.iter_rows(values_only=True)):
                # header row
                if row_index == 0:
                    headers = [str(h).strip() if h else "" for h in row]
                    continue

                data = dict(zip(headers, row))

                full_name = data.get("full_name")
                parent_phone = data.get("parent_phone")

                if not full_name or not parent_phone:
                    skipped += 1
                    continue

                # --- school ---
                school = None
                school_name = data.get("school_name")
                if school_name:
                    school, _ = School.objects.get_or_create(
                        name=str(school_name).strip()
                    )

                # --- enroll date ---
                enroll_date = data.get("enroll_date") or timezone.localdate()

                Student.objects.create(
                    full_name=str(full_name).strip(),
                    nickname=str(data.get("nickname") or "").strip(),
                    grade_level=str(data.get("grade_level") or "").strip(),
                    academic_year=str(data.get("academic_year") or "").strip(),
                    school=school,
                    parent_phone=str(parent_phone).strip(),
                    contact_channel=data.get("contact_channel") or "line",
                    enroll_date=enroll_date,
                    referral_source=data.get("referral_source") or "referral",
                    note=str(data.get("note") or "").strip(),
                    is_active=bool(data.get("is_active", True)),
                )
                created += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Import complete: {created} created, {skipped} skipped"
            )
        )

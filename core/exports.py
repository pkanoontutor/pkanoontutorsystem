"""Full-system Excel export.

One workbook, one sheet per data area, formatted so it is readable without any
post-processing: frozen + styled header row, auto-sized columns, autofilter.

Used by both the on-demand download view and the scheduled email job, so the
file the mailbox receives is byte-for-byte the same report as the manual one.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO

from django.db.models import Count, Q, Sum
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="0F172A")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="0F172A")

MAX_COL_WIDTH = 52


def _fmt_dt(value) -> str:
    if not value:
        return ""
    if timezone.is_aware(value):
        value = timezone.localtime(value)
    return value.strftime("%Y-%m-%d %H:%M")


def _fmt_d(value) -> str:
    return value.strftime("%Y-%m-%d") if value else ""


def _money(value) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _add_sheet(wb: Workbook, title: str, headers: list, rows) -> None:
    """Append one formatted sheet. `rows` may be any iterable of sequences."""
    # Excel sheet titles: max 31 chars, and []:*?/\ are illegal.
    safe_title = title[:31]
    for bad in "[]:*?/\\":
        safe_title = safe_title.replace(bad, "-")
    ws = wb.create_sheet(safe_title)

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 26

    count = 0
    for row in rows:
        ws.append(list(row))
        count += 1

    # Freeze the header and let the user filter/sort straight away.
    ws.freeze_panes = "A2"
    if count:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{count + 1}"

    for idx in range(1, len(headers) + 1):
        letter = get_column_letter(idx)
        widest = len(str(headers[idx - 1]))
        for cell in ws[letter][1:]:
            value = cell.value
            if value is not None:
                widest = max(widest, len(str(value)))
        ws.column_dimensions[letter].width = min(widest + 3, MAX_COL_WIDTH)


def _build_summary(wb: Workbook, generated_at: datetime, sheet_specs: list) -> None:
    """Cover sheet: what's in this file and how big each area is."""
    ws = wb.create_sheet("สรุปภาพรวม", 0)
    ws["A1"] = "Pkanoon Tutor — Export ข้อมูลทั้งระบบ"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"สร้างเมื่อ: {generated_at.strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(size=11, color="475569")

    ws["A4"] = "Sheet"
    ws["B4"] = "เนื้อหา"
    ws["C4"] = "จำนวนแถว"
    for cell in (ws["A4"], ws["B4"], ws["C4"]):
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for offset, (name, desc, count) in enumerate(sheet_specs, start=5):
        ws.cell(row=offset, column=1, value=name)
        ws.cell(row=offset, column=2, value=desc)
        ws.cell(row=offset, column=3, value=count)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 62
    ws.column_dimensions["C"].width = 12
    ws.freeze_panes = "A5"


def build_full_workbook() -> tuple[BytesIO, str]:
    """Build the whole-system workbook. Returns (buffer, filename)."""
    from .models import (
        Attendance,
        AdmissionInquiry,
        CoursePayment,
        CourseRenewalNotice,
        Enrollment,
        SchoolExpense,
        Sheet,
        SheetAllocation,
        SheetInventory,
        SheetInventoryMovement,
        SheetPrintOrder,
        Student,
        TeachingProgressUpdate,
        TutoringClass,
        Tutor,
        TutorPayrollEntry,
        WeeklyTest,
        WeeklyTestScore,
    )

    generated_at = timezone.localtime()
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet
    specs: list[tuple[str, str, int]] = []

    def add(title, desc, headers, rows):
        rows = list(rows)
        _add_sheet(wb, title, headers, rows)
        specs.append((title, desc, len(rows)))

    # ---------- Students & enrollments ----------
    # remaining_sessions is a Python property (sessions_total - used_sessions()),
    # not a column, so it can't be aggregated in SQL — prefetch and sum in Python.
    students = (
        Student.objects
        .order_by("-is_active", "grade_level", "student_code")
        .annotate(active_enrollments=Count("enrollments", filter=Q(enrollments__is_active=True)))
        .prefetch_related("enrollments")
    )

    def _student_rows():
        for s in students:
            remaining = sum(
                int(e.remaining_sessions or 0)
                for e in s.enrollments.all()
                if e.is_active
            )
            yield [
                s.id, s.student_code or "", s.nickname or "", s.full_name or "",
                s.grade_level or "", getattr(s.school, "name", "") or "",
                s.parent_phone or "", bool(s.is_active),
                int(s.active_enrollments or 0), remaining,
            ]

    add(
        "นักเรียน", "ข้อมูลนักเรียนทั้งหมด พร้อมจำนวนคอร์สที่ยัง active",
        ["ID", "รหัสนักเรียน", "ชื่อเล่น", "ชื่อ-นามสกุล", "ระดับชั้น", "โรงเรียน",
         "เบอร์ผู้ปกครอง", "ใช้งาน", "คอร์สที่ active", "ครั้งคงเหลือรวม"],
        _student_rows(),
    )

    enrollments = (
        Enrollment.objects.select_related("student", "tutoring_class")
        .order_by("-is_active", "tutoring_class__name", "student__nickname")
    )
    add(
        "คอร์สเรียน", "Enrollment ทุกรายการ พร้อมราคาและครั้งคงเหลือ",
        ["ID", "Sale Run No", "รหัสนักเรียน", "ชื่อเล่น", "Class", "ประเภท",
         "ครั้งทั้งหมด", "ครั้งคงเหลือ", "ราคาคอร์ส", "ส่วนลด", "ราคาสุทธิ",
         "ใช้งาน", "วันที่สร้าง"],
        (
            [e.id, e.sale_run_no or "", getattr(e.student, "student_code", "") or "",
             getattr(e.student, "nickname", "") or "",
             getattr(e.tutoring_class, "name", "") or "",
             e.get_enrollment_type_display() if hasattr(e, "get_enrollment_type_display") else "",
             int(e.sessions_total or 0),
             int(e.remaining_sessions or 0) if e.remaining_sessions is not None else 0,
             _money(e.course_price), _money(e.discount_amount), _money(e.net_price),
             bool(e.is_active), _fmt_dt(e.created_at)]
            for e in enrollments
        ),
    )

    add(
        "ห้องเรียน", "Class ทั้งหมดและจำนวนที่นั่ง",
        ["ID", "ชื่อ Class", "รอบเวลา", "ชม./ครั้ง", "ราคาคอร์ส", "ที่นั่งทั้งหมด", "ใช้งาน"],
        (
            [c.id, c.name or "", c.time_slot or "", _money(c.hours_per_session),
             _money(c.course_price), int(c.total_seats or 0), bool(c.is_active)]
            for c in TutoringClass.objects.order_by("name")
        ),
    )

    # ---------- Money ----------
    add(
        "ใบเสร็จรับเงิน", "CoursePayment ทุกใบ (รวมที่ยกเลิก)",
        ["เลขที่ใบเสร็จ", "วันที่", "นักเรียน", "Class", "ประเภท", "ช่องทาง",
         "จำนวนครั้ง", "ราคาคอร์ส", "ส่วนลด", "ยอดสุทธิ", "ยอดที่ชำระ", "สถานะ", "หมายเหตุ"],
        (
            [p.receipt_no or "", _fmt_d(p.payment_date),
             p.student.display_name if p.student_id else "",
             p.tutoring_class.name if p.tutoring_class_id else "",
             p.get_payment_type_display() if hasattr(p, "get_payment_type_display") else p.payment_type,
             p.get_payment_method_display() if hasattr(p, "get_payment_method_display") else p.payment_method,
             int(p.sessions_granted or 0), _money(p.course_price), _money(p.discount_amount),
             _money(p.net_amount), _money(p.amount_paid),
             p.get_status_display() if hasattr(p, "get_status_display") else p.status,
             p.note or ""]
            for p in CoursePayment.objects.select_related("student", "tutoring_class")
            .order_by("-payment_date", "-created_at")
        ),
    )

    add(
        "รายจ่ายโรงเรียน", "SchoolExpense ทุกรายการ",
        ["วันที่", "หมวด", "ผู้ขาย/ผู้รับเงิน", "รายละเอียด", "จำนวนเงิน", "ช่องทาง", "หมายเหตุ"],
        (
            [_fmt_d(x.expense_date), getattr(x.category, "name", "") or "",
             x.vendor or "", x.description or "", _money(x.amount),
             x.get_payment_method_display() if hasattr(x, "get_payment_method_display") else x.payment_method,
             x.note or ""]
            for x in SchoolExpense.objects.select_related("category").order_by("-expense_date", "-id")
        ),
    )

    add(
        "ค่าสอนติวเตอร์", "TutorPayrollEntry รายวัน พร้อมเรทที่ใช้จริง",
        ["วันที่สอน", "ติวเตอร์", "ชม. onsite", "เรท/ชม.", "ค่าสอน onsite", "เรทพิเศษ 325",
         "ชม.ออนไลน์", "ค่าสอนออนไลน์", "ค่าเดินทาง", "ค่านั่งว่าง", "ยอดรวม", "หมายเหตุ"],
        (
            [_fmt_d(p.work_date), p.tutor.name if p.tutor_id else "",
             _money(p.teaching_hours), _money(p.hourly_rate), _money(p.teaching_fee),
             "ใช่" if p.special_rate_325 else "ไม่", _money(p.online_teaching_hours),
             _money(p.online_teaching_fee), _money(p.travel_fee), _money(p.idle_fee),
             _money(p.total_amount), p.note or ""]
            for p in TutorPayrollEntry.objects.select_related("tutor").order_by("-work_date", "tutor__name")
        ),
    )

    add(
        "ติวเตอร์", "รายชื่อติวเตอร์และค่าเริ่มต้นเรทพิเศษ",
        ["ชื่อ", "เบอร์ติดต่อ", "ค่าเริ่มต้นเรท 325", "ใช้งาน", "หมายเหตุ"],
        (
            [t.name, t.phone or "", "ใช่" if t.default_special_rate_325 else "ไม่",
             bool(t.is_active), t.note or ""]
            for t in Tutor.objects.order_by("name")
        ),
    )

    # ---------- Sheets / stock ----------
    add(
        "คลังชีท", "ยอดคงเหลือชีทแต่ละเล่ม",
        ["รหัสชีท", "ชื่อชีท", "วิชา", "ระดับชั้น", "คงเหลือ", "ขั้นต่ำ", "เป้าหมาย", "ลิงก์ไฟล์"],
        (
            [getattr(i.sheet, "code", "") or "", getattr(i.sheet, "title", "") or "",
             getattr(getattr(i.sheet, "subject", None), "name", "") or "",
             getattr(i.sheet, "grade_level", "") or "",
             int(i.quantity or 0), int(getattr(i, "minimum_stock", 0) or 0),
             int(getattr(i, "target_stock", 0) or 0), i.onedrive_url or ""]
            for i in SheetInventory.objects.select_related("sheet", "sheet__subject").order_by("sheet__code")
        ),
    )

    add(
        "Movement ชีท", "ประวัติเพิ่ม/ลด/ปรับยอดชีท (ล่าสุด 5000 รายการ)",
        ["เวลา", "รหัสชีท", "ประเภท", "จำนวน", "คงเหลือหลังทำรายการ", "ผู้ทำรายการ", "หมายเหตุ"],
        (
            [_fmt_dt(m.created_at), getattr(m.sheet, "code", "") or "",
             m.get_movement_type_display() if hasattr(m, "get_movement_type_display") else m.movement_type,
             int(m.quantity or 0), int(getattr(m, "balance_after", 0) or 0),
             getattr(m.created_by, "username", "") or "", m.note or ""]
            for m in SheetInventoryMovement.objects.select_related("sheet", "created_by")
            .order_by("-created_at")[:5000]
        ),
    )

    add(
        "รายการสั่งปรินท์", "SheetPrintOrder ทุกสถานะ",
        ["ID", "รหัส/ชื่อเอกสาร", "จำนวน", "ปรินท์เสร็จแล้ว", "สถานะ", "การเย็บ",
         "ครบกำหนด", "สั่งเมื่อ", "ร้านแจ้งเสร็จ", "ตรวจรับเมื่อ", "หมายเหตุ"],
        (
            [o.id, f"{o.display_code} · {o.display_title}", int(o.quantity or 0),
             int(o.printed_quantity or 0),
             o.get_status_display() if hasattr(o, "get_status_display") else o.status,
             getattr(o, "binding_label", "") or "", _fmt_d(o.due_date), _fmt_dt(o.created_at),
             _fmt_dt(o.completed_at), _fmt_dt(o.received_at), o.note or ""]
            for o in SheetPrintOrder.objects.select_related("sheet", "sheet__subject")
            .order_by("-created_at")
        ),
    )

    add(
        "การแจกชีท", "SheetAllocation รายเด็ก (ล่าสุด 5000 รายการ)",
        ["วันที่แจก", "ผู้รับ", "ประเภทผู้รับ", "Class", "รหัสชีท", "ชื่อชีท",
         "จำนวน", "ผู้แจก", "หมายเหตุ"],
        (
            [_fmt_d(a.allocation_date),
             (getattr(a.student, "display_name", "") if a.student_id else "")
             or a.manual_nickname or "",
             a.get_recipient_type_display() if hasattr(a, "get_recipient_type_display") else (a.recipient_type or ""),
             getattr(a.tutoring_class, "name", "") or "",
             getattr(a.sheet, "code", "") or "", getattr(a.sheet, "title", "") or "",
             int(a.quantity or 0),
             getattr(a.created_by, "username", "") or "", a.note or ""]
            for a in SheetAllocation.objects.select_related(
                "student", "sheet", "tutoring_class", "created_by"
            ).order_by("-allocation_date", "-id")[:5000]
        ),
    )

    add(
        "ชีททั้งหมด", "Master data ของชีท",
        ["รหัสชีท", "ชื่อชีท", "วิชา", "ระดับชั้น"],
        (
            [s.code or "", s.title or "", getattr(s.subject, "name", "") or "",
             s.grade_level or ""]
            for s in Sheet.objects.select_related("subject").order_by("code")
        ),
    )

    # ---------- Teaching & tests ----------
    add(
        "บันทึกการสอน", "TeachingProgressUpdate — สอนถึงชีท/หน้า/ข้อไหน (ล่าสุด 5000 รายการ)",
        ["วันที่สอน", "ชีท", "ถึงหน้า", "ถึงข้อ", "ไม่มีการสอน", "ชีทใกล้หมด",
         "ผู้บันทึก", "บันทึกเมื่อ"],
        (
            [_fmt_d(t.teaching_date), t.sheet_name or "",
             t.page_to if t.page_to is not None else "",
             t.question_to if t.question_to is not None else "",
             "ใช่" if t.no_teaching else "", "ใช่" if t.sheet_near_end else "",
             t.updated_by_name or "", _fmt_dt(t.created_at)]
            for t in TeachingProgressUpdate.objects.select_related("assignment")
            .order_by("-teaching_date", "-id")[:5000]
        ),
    )

    add(
        "Test ย่อยรายสัปดาห์", "WeeklyTest ที่สร้างไว้",
        ["ID", "สัปดาห์เริ่ม", "ระดับชั้น", "วันที่สอบ", "วิชา", "หัวข้อ", "ความยาก", "หมายเหตุ"],
        (
            [w.id, _fmt_d(w.week_start), w.grade_level or "", _fmt_d(w.test_date),
             w.subject_name or getattr(w.subject, "name", "") or "", w.topic or "",
             w.get_difficulty_display() if hasattr(w, "get_difficulty_display") else (w.difficulty or ""),
             w.note or ""]
            for w in WeeklyTest.objects.select_related("subject").order_by("-week_start", "-id")
        ),
    )

    add(
        "คะแนน Test ย่อย", "WeeklyTestScore รายคน",
        ["สัปดาห์", "ระดับชั้น", "วิชา", "Class", "นักเรียน", "วันที่เช็คชื่อ",
         "สถานะเข้าเรียน", "ผลสอบ", "หมายเหตุ"],
        (
            [_fmt_d(getattr(s.weekly_test, "week_start", None)),
             getattr(s.weekly_test, "grade_level", "") or "",
             getattr(s.weekly_test, "subject_name", "") or "",
             getattr(s.tutoring_class, "name", "") or "",
             getattr(s.student, "display_name", "") or "",
             _fmt_d(s.attendance_date),
             s.get_attendance_status_display() if hasattr(s, "get_attendance_status_display") else (s.attendance_status or ""),
             s.result or "", s.note or ""]
            for s in WeeklyTestScore.objects.select_related(
                "weekly_test", "student", "tutoring_class"
            ).order_by("-id")[:10000]
        ),
    )

    # ---------- Ops ----------
    add(
        "การเช็คชื่อ", "Attendance (ล่าสุด 10000 รายการ)",
        ["วันที่", "นักเรียน", "Class", "สถานะ", "ตัดครั้งแล้ว", "เช็คเมื่อ"],
        (
            [_fmt_d(a.attendance_date),
             getattr(a.student, "display_name", "") if a.student_id else "",
             getattr(getattr(a.enrollment, "tutoring_class", None), "name", "") or "",
             a.get_status_display() if hasattr(a, "get_status_display") else (a.status or ""),
             "ใช่" if a.deducted else "", _fmt_dt(a.checked_at)]
            for a in Attendance.objects.select_related(
                "student", "enrollment", "enrollment__tutoring_class"
            ).order_by("-attendance_date", "-id")[:10000]
        ),
    )

    add(
        "ใบสมัคร/ทดลองเรียน", "AdmissionInquiry ทุกรายการ",
        ["วันที่", "ชื่อเล่น", "ชื่อ-นามสกุล", "โรงเรียนเดิม", "ระดับชั้น", "ประเภทคำขอ",
         "เบอร์ติดต่อ", "รอบที่สนใจ", "Class เป้าหมาย", "เตรียมชีทแล้ว", "มาทดลองเรียน",
         "ผลทดลองเรียน", "ปิดงานแล้ว", "หมายเหตุภายใน"],
        (
            [_fmt_dt(a.created_at), a.nickname or "",
             f"{a.first_name or ''} {a.last_name or ''}".strip(),
             a.school_name or "", a.grade_level or "",
             a.get_request_type_display() if hasattr(a, "get_request_type_display") else (a.request_type or ""),
             a.contact_phone or "", a.preferred_time_slot or "",
             getattr(a.target_class, "name", "") or "",
             "ใช่" if a.sheet_prepared else "", "ใช่" if a.trial_attended else "",
             a.trial_result or "", "ใช่" if a.is_completed else "", a.internal_note or ""]
            for a in AdmissionInquiry.objects.select_related("target_class").order_by("-created_at")
        ),
    )

    add(
        "ใบแจ้งต่อคอร์ส", "CourseRenewalNotice ที่ออกไปแล้ว",
        ["สร้างเมื่อ", "นักเรียน", "Class", "ประเภท", "คอร์สจบประมาณ",
         "แพ็ก 10 (สุทธิ)", "แพ็ก 20 (สุทธิ)", "แพ็ก 30 (สุทธิ)",
         "ผ่อน-ยอดคงเหลือ", "ส่งให้ผู้ปกครองแล้ว"],
        (
            [_fmt_dt(n.created_at),
             getattr(n.student, "display_name", "") if n.student_id else "",
             getattr(n.tutoring_class, "name", "") or "",
             n.get_notice_type_display() if hasattr(n, "get_notice_type_display") else (n.notice_type or ""),
             _fmt_d(n.expected_course_end_date),
             _money(n.package_10_net_price), _money(n.package_20_net_price),
             _money(n.package_30_net_price), _money(n.installment_remaining_amount),
             "ใช่" if n.is_sent_to_parent else ""]
            for n in CourseRenewalNotice.objects.select_related("student", "tutoring_class")
            .order_by("-created_at")
        ),
    )

    _build_summary(wb, generated_at, specs)

    buff = BytesIO()
    wb.save(buff)
    buff.seek(0)
    filename = f"pkanoon_full_export_{generated_at.strftime('%Y%m%d_%H%M')}.xlsx"
    return buff, filename

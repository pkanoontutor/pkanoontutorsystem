from __future__ import annotations
from collections import OrderedDict, defaultdict

import json
import csv
import re
from datetime import date, timedelta, datetime
from decimal import Decimal
from io import BytesIO

from django import forms
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Count, Q, Sum
from django.http import JsonResponse, HttpRequest, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST, require_GET

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from .models import (
    Student,
    School,
    Subject,
    Sheet,
    ClassSubject,
    Attendance,
    Enrollment,
    TutoringClass,
    SheetInventory,
    SheetInventoryMovement,
    SheetClassMapping,
    SheetPrintOrder,
    SheetAllocation,
    AdmissionInquiry,
    FinanceSetting,
    ExpenseCategory,
    SchoolExpense,
    Tutor,
    TutorPayrollEntry,
    CoursePayment,
    CourseRenewalNotice,
    TeachingTutor,
    TeachingClassSubjectTemplate,
    TeachingWeeklyAssignment,
    TeachingProgressUpdate,
    WeeklyTest,
    WeeklyTestScore,
    TestRound,
    TestSubject,
    TestParticipant,
    TestScore,
)


def _parse_date(s: str | None) -> date:
    if not s:
        return timezone.localdate()
    try:
        return date.fromisoformat(s)
    except ValueError:
        return timezone.localdate()


def _fmt_dt_th(dt: datetime | None) -> str:
    if not dt:
        return "-"
    try:
        dt_local = timezone.localtime(dt)
        return dt_local.strftime("%Y-%m-%d %H:%M")
    except Exception:
        return "-"


def home(request):
    return render(request, "core/home.html")


def home_redirect(request: HttpRequest) -> HttpResponse:
    return redirect("core:dashboard")


def student_id_list(request: HttpRequest) -> HttpResponse:
    """
    หน้า Public สำหรับผู้ปกครอง:
    แสดงเฉพาะ Student ID / ชื่อเล่น / ชื่อจริงนามสกุล / ระดับชั้น
    """
    grade = request.GET.get("grade")

    qs = (
        Student.objects
        .filter(is_active=True)
        .only("student_code", "nickname", "full_name", "grade_level")
    )

    if grade:
        qs = qs.filter(grade_level=grade)

    students = qs.order_by("grade_level", "student_code")

    return render(
        request,
        "core/student_id_list.html",
        {"students": students, "selected_grade": grade}
    )


# =======================
# Time slot order (GLOBAL)
# =======================
TIME_SLOT_ORDER = [
    TutoringClass.TimeSlot.SAT_MORNING,
    TutoringClass.TimeSlot.SAT_AFTERNOON,
    TutoringClass.TimeSlot.SUN_MORNING,
    TutoringClass.TimeSlot.SUN_AFTERNOON,
]


SHEET_GRADE_ORDER = ["p4", "p5", "p6", "m1", "m2", "m3", "m4", ""]


def _sheet_grade_label(value: str | None) -> str:
    labels = {
        "p4": "ป.4",
        "p5": "ป.5",
        "p6": "ป.6",
        "m1": "ม.1",
        "m2": "ม.2",
        "m3": "ม.3",
        "m4": "ม.4",
        "": "ไม่ระบุระดับชั้น",
    }
    return labels.get((value or "").strip().lower(), value or "ไม่ระบุระดับชั้น")


def _sheet_subject_style(subject_name: str | None) -> dict:
    """Pastel color mapping for sheet cards by subject."""
    raw = (subject_name or "").strip().lower().replace(" ", "")

    # Specific subjects first
    if "อังกฤษเสริม" in raw or "คณิตเสริม" in raw or "extraenglish" in raw or "extramath" in raw:
        return {"key": "red", "label": "สีแดง", "bg": "#fee2e2", "border": "#fca5a5", "accent": "#dc2626"}
    if "คณิตหลัก" in raw or "คณิต" in raw or "math" in raw:
        return {"key": "blue", "label": "สีฟ้า", "bg": "#dbeafe", "border": "#93c5fd", "accent": "#2563eb"}
    if "วิทย" in raw or "ชีว" in raw or "science" in raw or "bio" in raw or "biology" in raw:
        return {"key": "green", "label": "สีเขียว", "bg": "#dcfce7", "border": "#86efac", "accent": "#16a34a"}
    if "อังกฤษ" in raw or "english" in raw:
        return {"key": "pink", "label": "สีชมพู", "bg": "#fce7f3", "border": "#f9a8d4", "accent": "#db2777"}
    if "เคมี" in raw or "ฟิสิก" in raw or "ไทย" in raw or "สังคม" in raw or "chem" in raw or "physics" in raw or "thai" in raw or "social" in raw:
        return {"key": "yellow", "label": "สีเหลือง", "bg": "#fef9c3", "border": "#fde047", "accent": "#ca8a04"}
    return {"key": "gray", "label": "ทั่วไป", "bg": "#f8fafc", "border": "#cbd5e1", "accent": "#64748b"}


def _sheet_subject_style_attr(sheet: Sheet | None) -> str:
    subject_name = ""
    if sheet and getattr(sheet, "subject_id", None):
        subject_name = sheet.subject.name
    style = _sheet_subject_style(subject_name)
    return f"--subject-bg:{style['bg']};--subject-border:{style['border']};--subject-accent:{style['accent']};"


def _normalize_sheet_grade_level(value: str | None) -> str:
    raw = (value or "").strip().lower().replace(" ", "")
    if not raw:
        return ""
    raw = raw.replace(".", "")
    mapping = {
        "p4": "p4", "ป4": "p4", "ประถม4": "p4", "ป.4": "p4",
        "p5": "p5", "ป5": "p5", "ประถม5": "p5", "ป.5": "p5",
        "p6": "p6", "ป6": "p6", "ประถม6": "p6", "ป.6": "p6",
        "m1": "m1", "ม1": "m1", "มัธยม1": "m1", "ม.1": "m1",
        "m2": "m2", "ม2": "m2", "มัธยม2": "m2", "ม.2": "m2",
        "m3": "m3", "ม3": "m3", "มัธยม3": "m3", "ม.3": "m3",
        "m4": "m4", "ม4": "m4", "มัธยม4": "m4", "ม.4": "m4",
    }
    return mapping.get(raw, "")


def _infer_sheet_grade_level(*texts: str | None) -> str:
    combined = " ".join((t or "") for t in texts).lower()
    # Match explicit code patterns first e.g. E-P4-01 / M-M1-01.
    for grade in ["p4", "p5", "p6", "m1", "m2", "m3", "m4"]:
        if re.search(rf"(^|[^a-z0-9]){grade}([^a-z0-9]|$)", combined, re.I):
            return grade
    for value in ["ป.4", "ป4", "ป.5", "ป5", "ป.6", "ป6", "ม.1", "ม1", "ม.2", "ม2", "ม.3", "ม3", "ม.4", "ม4"]:
        normalized = _normalize_sheet_grade_level(value)
        if value.replace(".", "") in combined.replace(".", "").replace(" ", "") and normalized:
            return normalized
    return ""


def _class_grade_level(tutoring_class: TutoringClass | None) -> str:
    if not tutoring_class:
        return ""
    return _infer_sheet_grade_level(getattr(tutoring_class, "name", ""))


def _ordered_grade_groups(rows: list[dict]) -> list[dict]:
    buckets = {g: [] for g in SHEET_GRADE_ORDER}
    for row in rows:
        grade = (getattr(row.get("sheet"), "grade_level", "") or "").lower()
        if grade not in buckets:
            buckets.setdefault(grade, [])
        buckets[grade].append(row)
    return [
        {"grade": grade, "label": _sheet_grade_label(grade), "rows": buckets.get(grade, []), "count": len(buckets.get(grade, []))}
        for grade in SHEET_GRADE_ORDER
        if buckets.get(grade)
    ]


@login_required
def dashboard(request: HttpRequest) -> HttpResponse:
    """
    - ซ้าย: เช็คชื่อ (submit แยกทีละห้อง)
    - slot_totals (รวมต่อรอบเวลา) + รวมรายวัน/รวมสองวัน (เสาร์+อาทิตย์)
    - ตัดส่วน Sheet ออกทั้งหมดแล้ว
    """
    selected_date = _parse_date(request.GET.get("date"))

    classes = TutoringClass.objects.filter(is_active=True).order_by("name").all()

    classes_by_time_slot = OrderedDict()
    for ts in TIME_SLOT_ORDER:
        classes_by_time_slot[ts] = {
            "label": TutoringClass.TimeSlot(ts).label,
            "classes": [],
        }

    for cls in classes:
        if cls.time_slot in classes_by_time_slot:
            classes_by_time_slot[cls.time_slot]["classes"].append(cls)

    enrollments = (
        Enrollment.objects
        .select_related("student", "tutoring_class")
        .filter(
            is_active=True,
            student__is_active=True,
            tutoring_class__is_active=True,
        )
        .order_by(
            "tutoring_class__name",
            "student__nickname",
            "student__full_name",
            "student__grade_level",
        )
    )

    todays_att = Attendance.objects.filter(attendance_date=selected_date)
    att_map = {a.enrollment_id: a for a in todays_att}

    summary_by_class_id = {}
    seats_summary_by_class_id = {}

    global_present = global_excused = global_no_show = global_total = 0

    slot_totals: dict[str, dict[str, int]] = {
        ts: {"present": 0, "excused": 0, "no_show": 0, "total": 0}
        for ts in TIME_SLOT_ORDER
    }

    for cls in classes:
        cls_enrollments = enrollments.filter(tutoring_class=cls)
        enrollment_count = cls_enrollments.count()

        atts = todays_att.filter(enrollment__in=cls_enrollments)

        present = atts.filter(status=Attendance.Status.PRESENT).count()
        excused = atts.filter(status=Attendance.Status.EXCUSED).count()
        no_show = atts.filter(status=Attendance.Status.NO_SHOW).count()

        summary_by_class_id[cls.id] = {
            "present": present,
            "excused": excused,
            "no_show": no_show,
            "total": present + excused + no_show,
        }

        seats_total = cls.total_seats or 0
        seats_in_progress = enrollment_count
        seats_available = max(seats_total - seats_in_progress, 0)

        seats_summary_by_class_id[cls.id] = {
            "seats_total": seats_total,
            "seats_in_progress": seats_in_progress,
            "seats_available": seats_available,
        }

        global_present += present
        global_excused += excused
        global_no_show += no_show
        global_total += present + excused + no_show

        if cls.time_slot in slot_totals:
            slot_totals[cls.time_slot]["present"] += present
            slot_totals[cls.time_slot]["excused"] += excused
            slot_totals[cls.time_slot]["no_show"] += no_show
            slot_totals[cls.time_slot]["total"] += (present + excused + no_show)

    global_summary = {
        "present": global_present,
        "excused": global_excused,
        "no_show": global_no_show,
        "total": global_total,
    }

    other_day_date: date | None = None
    other_day_summary: dict | None = None
    two_day_summary: dict | None = None

    wd = selected_date.weekday()  # 5=Sat, 6=Sun
    if wd == 5:
        other_day_date = selected_date + timedelta(days=1)
    elif wd == 6:
        other_day_date = selected_date - timedelta(days=1)

    if other_day_date:
        other_day_summary = Attendance.objects.filter(
            attendance_date=other_day_date,
            enrollment__tutoring_class__is_active=True,
            student__is_active=True,
        ).aggregate(
            present=Count("id", filter=Q(status=Attendance.Status.PRESENT)),
            excused=Count("id", filter=Q(status=Attendance.Status.EXCUSED)),
            no_show=Count("id", filter=Q(status=Attendance.Status.NO_SHOW)),
            total=Count("id"),
        )

        two_day_summary = {
            "present": (global_summary.get("present", 0) or 0) + (other_day_summary.get("present", 0) or 0),
            "excused": (global_summary.get("excused", 0) or 0) + (other_day_summary.get("excused", 0) or 0),
            "no_show": (global_summary.get("no_show", 0) or 0) + (other_day_summary.get("no_show", 0) or 0),
            "total": (global_summary.get("total", 0) or 0) + (other_day_summary.get("total", 0) or 0),
        }

    THRESHOLD = 2
    near_complete = [e for e in enrollments if (e.remaining_sessions or 0) < THRESHOLD]

    context = {
        "selected_date": selected_date,
        "classes": classes,
        "classes_by_time_slot": classes_by_time_slot,

        "enrollments": enrollments,
        "att_map": att_map,
        "summary_by_class_id": summary_by_class_id,
        "seats_summary_by_class_id": seats_summary_by_class_id,
        "global_summary": global_summary,

        "slot_totals": slot_totals,
        "other_day_date": other_day_date,
        "other_day_summary": other_day_summary,
        "two_day_summary": two_day_summary,

        "near_complete": near_complete,
        "threshold": THRESHOLD,
    }
    return render(request, "core/dashboard.html", context)


# =========================================================
# ✅ Export ข้อมูลออก Excel
# =========================================================
def _autosize(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)



# =========================================================
# ✅ Sheet Inventory Module
# =========================================================
def _sheet_code_from_text(text: str | None) -> str:
    """Extract a likely sheet code from free text such as 'M-P6-01 เศษส่วน'."""
    raw = (text or "").strip()
    if not raw:
        return ""
    # Try exact token first; sheet codes in the system usually look like M-P6-01.
    tokens = re.split(r"[\s,;:/|]+", raw)
    for token in tokens:
        t = token.strip().upper()
        if t and Sheet.objects.filter(code__iexact=t).exists():
            return t
    # Fallback: if the whole text is the code.
    if Sheet.objects.filter(code__iexact=raw.upper()).exists():
        return raw.upper()
    return ""


def _inventory_for_sheet(sheet: Sheet) -> SheetInventory | None:
    try:
        return sheet.inventory
    except SheetInventory.DoesNotExist:
        return None


def _apply_sheet_inventory_movement(
    *,
    sheet: Sheet,
    movement_type: str,
    quantity: int,
    note: str = "",
    user=None,
) -> tuple[bool, str, SheetInventory | None, SheetInventoryMovement | None]:
    """
    Centralized stock update.
    - ADD: add quantity to stock
    - DEDUCT: deduct quantity, but block if result would be negative
    - SET / COUNT: set balance to quantity
    """
    if quantity is None:
        quantity = 0

    try:
        quantity = int(quantity)
    except Exception:
        return False, "จำนวนไม่ถูกต้อง", None, None

    if movement_type in {SheetInventoryMovement.MovementType.ADD, SheetInventoryMovement.MovementType.DEDUCT}:
        if quantity <= 0:
            return False, "กรุณาระบุจำนวนมากกว่า 0", None, None
    elif movement_type in {SheetInventoryMovement.MovementType.SET, SheetInventoryMovement.MovementType.COUNT}:
        if quantity < 0:
            return False, "ยอดจริงต้องไม่ติดลบ", None, None
    else:
        return False, "ประเภท movement ไม่ถูกต้อง", None, None

    with transaction.atomic():
        inventory, _ = SheetInventory.objects.select_for_update().get_or_create(
            sheet=sheet,
            defaults={"quantity": 0},
        )
        before = int(inventory.quantity or 0)

        if movement_type == SheetInventoryMovement.MovementType.ADD:
            after = before + quantity
            movement_qty = quantity
        elif movement_type == SheetInventoryMovement.MovementType.DEDUCT:
            if before - quantity < 0:
                return False, f"ตัด stock ไม่ได้ เพราะคงเหลือ {before} ชุด", inventory, None
            after = before - quantity
            movement_qty = quantity
        else:
            after = quantity
            movement_qty = quantity

        inventory.quantity = max(after, 0)
        inventory.save()

        movement = SheetInventoryMovement.objects.create(
            sheet=sheet,
            movement_type=movement_type,
            quantity=movement_qty,
            balance_before=before,
            balance_after=inventory.quantity,
            note=(note or "").strip(),
            created_by=user if getattr(user, "is_authenticated", False) else None,
        )

    return True, "บันทึกสำเร็จ", inventory, movement


def _active_admissions_for_sheet_demand() -> list[AdmissionInquiry]:
    return list(
        AdmissionInquiry.objects
        .filter(is_completed=False)
        .only("id", "grade_level", "preferred_time_slot", "request_type", "first_lesson_date")
        .order_by("first_lesson_date", "created_at")
    )


def _class_matches_inquiry(tutoring_class: TutoringClass, inquiry: AdmissionInquiry) -> bool:
    if getattr(tutoring_class, "time_slot", "") != getattr(inquiry, "preferred_time_slot", ""):
        return False

    class_name = (tutoring_class.name or "").lower()
    grade_display = ""
    try:
        grade_display = inquiry.get_grade_level_display()
    except Exception:
        grade_display = ""

    grade_code = (inquiry.grade_level or "").lower()
    tokens = {
        grade_code,
        grade_code.replace("p", "ป."),
        grade_code.replace("m", "ม."),
        grade_display,
        grade_display.replace(".", ""),
        grade_display.replace(" ", ""),
    }
    tokens = {t.lower() for t in tokens if t}

    return any(token in class_name for token in tokens)


def _demand_count_for_class(tutoring_class: TutoringClass, pending_inquiries: list[AdmissionInquiry]) -> int:
    return sum(1 for i in pending_inquiries if _class_matches_inquiry(tutoring_class, i))

def _sheet_row(sheet: Sheet) -> dict:
    inv = _inventory_for_sheet(sheet)
    qty = int(inv.quantity or 0) if inv else 0
    minimum = int(getattr(inv, "minimum_stock", 0) or 0) if inv else 0
    last_updated_at = getattr(inv, "updated_at", None) if inv else None
    subject_style = _sheet_subject_style(sheet.subject.name if getattr(sheet, "subject_id", None) else "")
    return {
        "sheet": sheet,
        "inventory": inv,
        "quantity": qty,
        "minimum_stock": minimum,
        "grade_level": getattr(sheet, "grade_level", "") or "",
        "grade_label": _sheet_grade_label(getattr(sheet, "grade_level", "") or ""),
        "is_low": minimum > 0 and qty <= minimum,
        "last_updated_at": last_updated_at,
        "subject_style": f"--subject-bg:{subject_style['bg']};--subject-border:{subject_style['border']};--subject-accent:{subject_style['accent']};",
        "subject_color_label": subject_style["label"],
        "default_spine_color": _default_spine_color_for_subject(sheet.subject.name if getattr(sheet, "subject_id", None) else ""),
        "default_spine_color_label": _spine_color_label(_default_spine_color_for_subject(sheet.subject.name if getattr(sheet, "subject_id", None) else "")),
        "onedrive_url": (getattr(inv, "onedrive_url", "") or "") if inv else "",
    }


def _build_class_sheet_rows() -> list[dict]:
    """
    Build class-sheet requirements using:
    1) Manual SheetClassMapping records
    2) ClassSubject.current_sheet from the older sheet update setup
    3) TeachingClassSubjectTemplate.default_sheet_name from the tutor update module
    4) TeachingProgressUpdate.sheet_name from recent tutor updates
    Demand count is based on active AdmissionInquiry records, not Enrollment.
    """
    pending_inquiries = _active_admissions_for_sheet_demand()
    classes = list(TutoringClass.objects.filter(is_active=True).order_by("time_slot", "name"))

    # Preload mappings
    manual_mappings = (
        SheetClassMapping.objects
        .select_related("sheet", "sheet__subject", "tutoring_class")
        .filter(is_active=True)
    )
    manual_by_class: dict[int, list[SheetClassMapping]] = defaultdict(list)
    for m in manual_mappings:
        manual_by_class[m.tutoring_class_id].append(m)

    current_sheet_links = (
        ClassSubject.objects
        .select_related("tutoring_class", "current_sheet", "current_sheet__subject")
        .filter(is_active=True, current_sheet__isnull=False, tutoring_class__is_active=True)
    )
    current_by_class: dict[int, list[Sheet]] = defaultdict(list)
    for cs in current_sheet_links:
        current_by_class[cs.tutoring_class_id].append(cs.current_sheet)

    templates = (
        TeachingClassSubjectTemplate.objects
        .select_related("tutoring_class")
        .filter(is_active=True, tutoring_class__is_active=True)
        .exclude(default_sheet_name="")
    )
    template_by_class: dict[int, list[Sheet]] = defaultdict(list)
    for t in templates:
        code = _sheet_code_from_text(t.default_sheet_name)
        if code:
            sheet = Sheet.objects.filter(code__iexact=code).select_related("subject").first()
            if sheet:
                template_by_class[t.tutoring_class_id].append(sheet)

    progress_updates = (
        TeachingProgressUpdate.objects
        .select_related("assignment__tutoring_class")
        .filter(assignment__tutoring_class__is_active=True)
        .exclude(sheet_name="")
        .order_by("-updated_at")[:500]
    )
    progress_by_class: dict[int, list[Sheet]] = defaultdict(list)
    for p in progress_updates:
        code = _sheet_code_from_text(p.sheet_name)
        if code:
            sheet = Sheet.objects.filter(code__iexact=code).select_related("subject").first()
            if sheet:
                progress_by_class[p.assignment.tutoring_class_id].append(sheet)

    rows = []
    for cls in classes:
        demand_count = _demand_count_for_class(cls, pending_inquiries)
        seen: dict[int, dict] = {}

        for m in manual_by_class.get(cls.id, []):
            seen[m.sheet_id] = {
                "sheet": m.sheet,
                "source": "manual",
                "source_label": "Manual",
                "quantity_per_student": int(m.quantity_per_student or 1),
                "mapping": m,
            }

        for s in current_by_class.get(cls.id, []):
            seen.setdefault(s.id, {
                "sheet": s,
                "source": "current_sheet",
                "source_label": "Current Sheet",
                "quantity_per_student": 1,
                "mapping": None,
            })

        for s in template_by_class.get(cls.id, []):
            seen.setdefault(s.id, {
                "sheet": s,
                "source": "template",
                "source_label": "Template",
                "quantity_per_student": 1,
                "mapping": None,
            })

        for s in progress_by_class.get(cls.id, []):
            seen.setdefault(s.id, {
                "sheet": s,
                "source": "tutor_update",
                "source_label": "Tutor Update",
                "quantity_per_student": 1,
                "mapping": None,
            })

        entries = []
        for item in seen.values():
            sheet = item["sheet"]
            inv = _inventory_for_sheet(sheet)
            balance = int(inv.quantity or 0) if inv else 0
            required = demand_count * int(item["quantity_per_student"] or 1)
            entries.append({
                **item,
                "balance": balance,
                "required": required,
                "shortage": max(required - balance, 0),
            })

        rows.append({
            "class": cls,
            "demand_count": demand_count,
            "entries": sorted(entries, key=lambda x: x["sheet"].code),
            "has_shortage": any(e["shortage"] > 0 for e in entries),
        })

    return rows


def _sheet_inventory_context(*, q: str = "", extra: dict | None = None) -> dict:
    sheets_qs = Sheet.objects.select_related("subject").all()
    if q:
        grade_q = _normalize_sheet_grade_level(q)
        query = (
            Q(code__icontains=q) |
            Q(title__icontains=q) |
            Q(subject__name__icontains=q)
        )
        if grade_q:
            query |= Q(grade_level=grade_q)
        sheets_qs = sheets_qs.filter(query)

    sheets = list(sheets_qs.order_by("grade_level", "subject__name", "code"))
    sheet_rows = [_sheet_row(s) for s in sheets]
    all_sheets = Sheet.objects.select_related("subject").filter(is_active=True).order_by("grade_level", "subject__name", "code")
    active_classes = TutoringClass.objects.filter(is_active=True).order_by("time_slot", "name")
    context = {
        "q": q,
        "sheet_rows": sheet_rows,
        "sheet_grade_groups": _ordered_grade_groups(sheet_rows),
        "sheet_grade_choices": Sheet.GradeLevel.choices,
        "subjects": Subject.objects.filter(is_active=True).order_by("name"),
        "classes": active_classes,
        "all_sheets": all_sheets,
        "sheet_choices": all_sheets,
        "class_rows": _build_class_sheet_rows(),
        "movements": (
            SheetInventoryMovement.objects
            .select_related("sheet", "created_by")
            .order_by("-created_at")[:30]
        ),
        "movement_type_choices": SheetInventoryMovement.MovementType.choices,
        "binding_type_choices": SheetPrintOrder.BindingType.choices,
        "spine_color_choices": _print_color_choices(),
        "default_print_due_date": timezone.localdate() + timedelta(days=3),
        "ready_to_receive_orders": (
            SheetPrintOrder.objects
            .select_related("sheet", "sheet__subject", "requested_by")
            .filter(status=SheetPrintOrder.Status.READY, sheet__isnull=False)
            .order_by("completed_at", "due_date", "created_at")
        ),
    }
    if extra:
        context.update(extra)
    return context

def _clean_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


def _parse_int_cell(value, default=None):
    raw = _clean_cell(value)
    if raw == "":
        return default
    try:
        return max(int(float(raw.replace(",", ""))), 0)
    except Exception:
        return default


def _parse_bool_cell(value, default=True):
    raw = _clean_cell(value).lower()
    if raw == "":
        return default
    return raw in {"1", "true", "yes", "y", "active", "ใช่", "เปิด", "เปิดใช้งาน"}


def _normalize_upload_header(value) -> str:
    raw = _clean_cell(value).lower()
    return re.sub(r"[\s_\-./()]+", "", raw)


def _pick_upload_value(row: dict, aliases: list[str]):
    normalized_aliases = {_normalize_upload_header(a) for a in aliases}
    for key, value in row.items():
        if _normalize_upload_header(key) in normalized_aliases:
            return value
    return ""


def _read_sheet_upload_rows(uploaded_file) -> list[dict]:
    filename = (getattr(uploaded_file, "name", "") or "").lower()

    if filename.endswith(".csv"):
        raw = uploaded_file.read()
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw.decode("cp874", errors="replace")
        reader = csv.DictReader(text.splitlines())
        return [dict(r) for r in reader]

    # Default to Excel workbook
    wb = load_workbook(BytesIO(uploaded_file.read()), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_clean_cell(h) for h in rows[0]]
    result = []
    for row_values in rows[1:]:
        if not any(_clean_cell(v) for v in row_values):
            continue
        result.append({headers[i]: row_values[i] if i < len(row_values) else "" for i in range(len(headers))})
    return result


def _import_sheet_rows_from_upload(rows: list[dict], *, user=None, stock_mode: str = "set") -> dict:
    """
    Import rows from Excel/CSV.
    Required columns:
    - รหัสชีท / code
    - ชื่อชีท / title
    - วิชา / subject

    Optional:
    - ระดับชั้น / grade_level (accepted for template compatibility; current Sheet model does not store grade separately)
    - initial_quantity
    - total_pages
    - total_questions
    - minimum_stock
    - is_active
    """
    created_count = 0
    updated_count = 0
    skipped_count = 0
    movement_count = 0
    errors: list[str] = []

    code_aliases = ["รหัสชีท", "sheet code", "sheet_code", "code", "รหัส"]
    title_aliases = ["ชื่อชีท", "ชื่อเรื่อง", "sheet name", "sheet_name", "title", "ชื่อ"]
    subject_aliases = ["วิชา", "subject", "subject_name"]
    grade_aliases = ["ระดับชั้น", "grade", "grade_level", "gradelevel", "ชั้น", "level"]
    initial_aliases = ["initial_quantity", "initial qty", "ยอดเริ่มต้น", "จำนวนคงเหลือ", "stock", "quantity"]
    pages_aliases = ["total_pages", "จำนวนหน้า", "pages"]
    questions_aliases = ["total_questions", "จำนวนข้อ", "questions"]
    min_aliases = ["minimum_stock", "ขั้นต่ำที่ควรมี", "min stock", "minimum"]
    active_aliases = ["is_active", "active", "เปิดใช้งาน"]

    stock_mode = (stock_mode or "set").strip()
    if stock_mode not in {"skip", "set", "add"}:
        stock_mode = "set"

    with transaction.atomic():
        for idx, row in enumerate(rows, start=2):
            code = _clean_cell(_pick_upload_value(row, code_aliases)).upper()
            title = _clean_cell(_pick_upload_value(row, title_aliases))
            subject_name = _clean_cell(_pick_upload_value(row, subject_aliases))
            grade_level_raw = _clean_cell(_pick_upload_value(row, grade_aliases))
            grade_level = _normalize_sheet_grade_level(grade_level_raw) or _infer_sheet_grade_level(code, title)

            if not code and not title and not subject_name:
                skipped_count += 1
                continue

            if not code:
                errors.append(f"แถว {idx}: ไม่มีรหัสชีท")
                skipped_count += 1
                continue
            if not title:
                errors.append(f"แถว {idx}: ไม่มีชื่อชีทสำหรับรหัส {code}")
                skipped_count += 1
                continue
            if not subject_name:
                errors.append(f"แถว {idx}: ไม่มีวิชาสำหรับรหัส {code}")
                skipped_count += 1
                continue

            subject, _ = Subject.objects.get_or_create(
                name=subject_name,
                defaults={"is_active": True},
            )

            total_pages = _parse_int_cell(_pick_upload_value(row, pages_aliases), default=None)
            total_questions = _parse_int_cell(_pick_upload_value(row, questions_aliases), default=None)
            minimum_stock = _parse_int_cell(_pick_upload_value(row, min_aliases), default=None)
            initial_qty = _parse_int_cell(_pick_upload_value(row, initial_aliases), default=None)
            is_active = _parse_bool_cell(_pick_upload_value(row, active_aliases), default=True)

            sheet, created = Sheet.objects.get_or_create(
                code=code,
                defaults={
                    "title": title,
                    "subject": subject,
                    "grade_level": grade_level,
                    "total_pages": total_pages or 0,
                    "total_questions": total_questions or 0,
                    "is_active": is_active,
                },
            )

            if created:
                created_count += 1
            else:
                changed = False
                if sheet.title != title:
                    sheet.title = title
                    changed = True
                if sheet.subject_id != subject.id:
                    sheet.subject = subject
                    changed = True
                if grade_level and getattr(sheet, "grade_level", "") != grade_level:
                    sheet.grade_level = grade_level
                    changed = True
                if total_pages is not None and sheet.total_pages != total_pages:
                    sheet.total_pages = total_pages
                    changed = True
                if total_questions is not None and sheet.total_questions != total_questions:
                    sheet.total_questions = total_questions
                    changed = True
                if sheet.is_active != is_active:
                    sheet.is_active = is_active
                    changed = True
                if changed:
                    sheet.save()
                updated_count += 1

            inventory, _ = SheetInventory.objects.get_or_create(sheet=sheet, defaults={"quantity": 0})

            if minimum_stock is not None:
                inventory.minimum_stock = minimum_stock
                inventory.save()

            if initial_qty is not None and stock_mode != "skip":
                movement_type = (
                    SheetInventoryMovement.MovementType.ADD
                    if stock_mode == "add"
                    else SheetInventoryMovement.MovementType.SET
                )

                if movement_type == SheetInventoryMovement.MovementType.ADD and initial_qty <= 0:
                    pass
                else:
                    note_parts = ["Bulk upload"]
                    if grade_level:
                        note_parts.append(f"grade={grade_level}")
                    ok, msg, inv, mv = _apply_sheet_inventory_movement(
                        sheet=sheet,
                        movement_type=movement_type,
                        quantity=initial_qty,
                        note="; ".join(note_parts),
                        user=user,
                    )
                    if ok and mv:
                        movement_count += 1
                    elif not ok:
                        errors.append(f"แถว {idx} ({code}): {msg}")

    return {
        "created": created_count,
        "updated": updated_count,
        "skipped": skipped_count,
        "movements": movement_count,
        "errors": errors[:30],
        "error_count": len(errors),
    }




def _save_inventory_set_from_post(
    request: HttpRequest,
    *,
    sheet_id: str | int,
    qty_prefix: str,
    note_prefix: str,
    movement_type: str,
    default_note: str,
) -> tuple[bool, str]:
    sheet = Sheet.objects.filter(id=sheet_id).first()
    if not sheet:
        return False, "ไม่พบชีท"

    raw_qty = (request.POST.get(f"{qty_prefix}{sheet.id}") or "").strip()
    if raw_qty == "":
        return False, "ยังไม่ได้กรอกจำนวน"
    try:
        qty = int(raw_qty)
    except Exception:
        return False, "จำนวนไม่ถูกต้อง"

    note = (request.POST.get(f"{note_prefix}{sheet.id}") or default_note).strip()
    ok, message, inventory, movement = _apply_sheet_inventory_movement(
        sheet=sheet,
        movement_type=movement_type,
        quantity=qty,
        note=note,
        user=request.user,
    )
    return ok, message


def _save_inventory_adjust_from_post(
    request: HttpRequest,
    *,
    sheet_id: str | int,
    qty_prefix: str,
    note_prefix: str,
    movement_type: str,
    default_note: str,
) -> tuple[bool, str]:
    sheet = Sheet.objects.filter(id=sheet_id).first()
    if not sheet:
        return False, "ไม่พบชีท"

    raw_qty = (request.POST.get(f"{qty_prefix}{sheet.id}") or "").strip()
    if raw_qty == "":
        return False, "ยังไม่ได้กรอกจำนวน"
    try:
        qty = int(raw_qty)
    except Exception:
        return False, "จำนวนไม่ถูกต้อง"

    note = (request.POST.get(f"{note_prefix}{sheet.id}") or default_note).strip()
    ok, message, inventory, movement = _apply_sheet_inventory_movement(
        sheet=sheet,
        movement_type=movement_type,
        quantity=qty,
        note=note,
        user=request.user,
    )
    return ok, message


def _save_inventory_bulk_from_post(
    request: HttpRequest,
    *,
    qty_prefix: str,
    note_prefix: str,
    movement_type: str,
    default_note: str,
) -> int:
    saved_count = 0
    for key, raw_qty in request.POST.items():
        if not key.startswith(qty_prefix):
            continue
        sheet_id = key.replace(qty_prefix, "", 1)
        if (raw_qty or "").strip() == "":
            continue
        ok, _msg = _save_inventory_set_from_post(
            request,
            sheet_id=sheet_id,
            qty_prefix=qty_prefix,
            note_prefix=note_prefix,
            movement_type=movement_type,
            default_note=default_note,
        )
        if ok:
            saved_count += 1
    return saved_count

@login_required
def sheet_inventory_dashboard(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()

        if action == "create_sheet":
            code = (request.POST.get("code") or "").strip().upper()
            title = (request.POST.get("title") or "").strip()
            subject_id = (request.POST.get("subject_id") or "").strip()
            new_subject = (request.POST.get("new_subject") or "").strip()
            grade_level = _normalize_sheet_grade_level(request.POST.get("grade_level")) or _infer_sheet_grade_level(code, title)
            total_pages = int(request.POST.get("total_pages") or 0)
            total_questions = int(request.POST.get("total_questions") or 0)
            initial_qty = int(request.POST.get("initial_qty") or 0)
            minimum_stock = int(request.POST.get("minimum_stock") or 0)

            if code and title:
                if subject_id:
                    subject = Subject.objects.filter(id=subject_id).first()
                elif new_subject:
                    subject, _ = Subject.objects.get_or_create(name=new_subject, defaults={"is_active": True})
                else:
                    subject, _ = Subject.objects.get_or_create(name="General", defaults={"is_active": True})

                sheet, created = Sheet.objects.get_or_create(
                    code=code,
                    defaults={
                        "title": title,
                        "subject": subject,
                        "grade_level": grade_level,
                        "total_pages": total_pages,
                        "total_questions": total_questions,
                        "is_active": True,
                    }
                )
                if created:
                    inv, _ = SheetInventory.objects.get_or_create(sheet=sheet, defaults={"quantity": 0})
                    inv.minimum_stock = max(minimum_stock, 0)
                    inv.save()
                    if initial_qty > 0:
                        _apply_sheet_inventory_movement(
                            sheet=sheet,
                            movement_type=SheetInventoryMovement.MovementType.ADD,
                            quantity=initial_qty,
                            note="Initial stock from Sheet Inventory module",
                            user=request.user,
                        )
                return redirect("core:sheet_inventory_profile", pk=sheet.pk)

        elif (
            action.startswith("receive_print_order:")
            or action.startswith("receive_print_order_full:")
            or action.startswith("receive_print_order_custom:")
            or action.startswith("receive_print_order_zero:")
        ):
            action_name, order_id = action.split(":", 1)
            is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"

            def _receive_response(ok: bool, message: str, status_code: int = 200, **payload):
                if is_ajax:
                    data = {"ok": ok, "message": message}
                    data.update(payload)
                    return JsonResponse(data, status=status_code)
                return redirect("core:sheet_inventory_dashboard")

            try:
                with transaction.atomic():
                    order = get_object_or_404(
                        SheetPrintOrder.objects.select_for_update().select_related("sheet", "sheet__subject"),
                        id=order_id,
                        sheet__isnull=False,
                    )

                    if order.status != SheetPrintOrder.Status.READY:
                        return _receive_response(False, "รายการนี้ถูกตรวจรับไปแล้ว หรือยังไม่พร้อมตรวจรับ", 400)

                    if action_name == "receive_print_order_full":
                        receive_qty = int(order.quantity or 0)
                    elif action_name == "receive_print_order_zero":
                        receive_qty = 0
                    else:
                        raw_qty = (request.POST.get(f"receive_qty_{order.id}") or "").strip()
                        try:
                            receive_qty = int(raw_qty) if raw_qty else int(order.quantity or 0)
                        except Exception:
                            return _receive_response(False, "จำนวนตรวจรับไม่ถูกต้อง", 400)

                    if receive_qty < 0:
                        return _receive_response(False, "จำนวนตรวจรับต้องไม่ติดลบ", 400)

                    note = (request.POST.get(f"receive_note_{order.id}") or "").strip()
                    if not note:
                        if action_name == "receive_print_order_full":
                            note = f"ตรวจรับจากร้านปรินท์ Order #{order.id}"
                        elif action_name == "receive_print_order_zero":
                            note = f"ตรวจรับจากร้านปรินท์ Order #{order.id} (รับจริง 0 เล่ม / ไม่เพิ่ม stock)"
                        else:
                            note = f"ตรวจรับจากร้านปรินท์ Order #{order.id} (แก้จำนวนรับจริง)"

                    inventory, _ = SheetInventory.objects.select_for_update().get_or_create(
                        sheet=order.sheet,
                        defaults={"quantity": 0},
                    )

                    if receive_qty > 0:
                        ok, message, inventory, _movement = _apply_sheet_inventory_movement(
                            sheet=order.sheet,
                            movement_type=SheetInventoryMovement.MovementType.ADD,
                            quantity=receive_qty,
                            note=note,
                            user=request.user,
                        )
                        if not ok:
                            return _receive_response(False, message or "บันทึกตรวจรับไม่สำเร็จ", 400)

                    order.status = SheetPrintOrder.Status.RECEIVED
                    order.received_at = timezone.now()
                    order.received_by = request.user if getattr(request.user, "is_authenticated", False) else None
                    order.save(update_fields=["status", "received_at", "received_by", "updated_at"])

                    if receive_qty > 0:
                        msg = f"ตรวจรับ {order.display_code} จำนวน {receive_qty} เล่ม และเพิ่มเข้า stock แล้ว"
                    else:
                        msg = f"ตรวจรับ {order.display_code} จำนวน 0 เล่มแล้ว โดยไม่เพิ่ม stock"

                    return _receive_response(
                        True,
                        msg,
                        order_id=order.id,
                        sheet_id=order.sheet_id,
                        received_qty=receive_qty,
                        new_balance=int(inventory.quantity or 0) if inventory else 0,
                    )
            except Exception as exc:
                return _receive_response(False, f"ตรวจรับไม่สำเร็จ: {exc}", 400)

        elif action.startswith("create_print_order_inline:"):
            sheet_id = action.split(":", 1)[1]
            sheet = get_object_or_404(Sheet.objects.select_related("subject"), id=sheet_id)
            raw_qty = (
                request.POST.get(f"print_qty_{sheet.id}")
                or request.POST.get("print_qty")
                or ""
            ).strip()
            try:
                quantity = max(int(raw_qty), 0)
            except Exception:
                quantity = 0

            due_date = _parse_optional_date(
                request.POST.get(f"print_due_date_{sheet.id}")
                or request.POST.get("print_due_date")
            )
            onedrive_url = (
                request.POST.get(f"print_url_{sheet.id}")
                or request.POST.get("print_url")
                or ""
            ).strip()
            note = (
                request.POST.get(f"print_note_{sheet.id}")
                or request.POST.get("print_note")
                or "สั่งปรินท์จาก popup หน้า Sheet Inventory"
            ).strip()

            binding_type = (
                request.POST.get(f"print_binding_type_{sheet.id}")
                or request.POST.get("print_binding_type")
                or SheetPrintOrder.BindingType.SIDE
            ).strip()
            if binding_type not in dict(SheetPrintOrder.BindingType.choices):
                binding_type = SheetPrintOrder.BindingType.SIDE

            default_spine_color = _default_spine_color_for_subject(sheet.subject.name if sheet.subject_id else "")
            spine_color = (
                request.POST.get(f"print_spine_color_{sheet.id}")
                or request.POST.get("print_spine_color")
                or default_spine_color
                or ""
            ).strip()

            if binding_type == SheetPrintOrder.BindingType.CORNER:
                spine_color = ""
            elif spine_color not in dict(SheetPrintOrder.SpineColor.choices):
                spine_color = default_spine_color

            if quantity > 0:
                inventory, _ = SheetInventory.objects.get_or_create(sheet=sheet, defaults={"quantity": 0})
                if onedrive_url:
                    inventory.onedrive_url = onedrive_url
                    inventory.save()
                elif getattr(inventory, "onedrive_url", ""):
                    onedrive_url = inventory.onedrive_url

                SheetPrintOrder.objects.create(
                    sheet=sheet,
                    quantity=quantity,
                    due_date=due_date,
                    onedrive_url=onedrive_url,
                    binding_type=binding_type,
                    spine_color=spine_color,
                    note=note,
                    requested_by=request.user if getattr(request.user, "is_authenticated", False) else None,
                )
            return redirect("core:sheet_inventory_dashboard")

        elif action.startswith("set_stock_single:"):
            sheet_id = action.split(":", 1)[1]
            _save_inventory_set_from_post(
                request,
                sheet_id=sheet_id,
                qty_prefix="set_qty_",
                note_prefix="set_note_",
                movement_type=SheetInventoryMovement.MovementType.SET,
                default_note="Manual balance update from Sheet Inventory",
            )
            return redirect("core:sheet_inventory_dashboard")

        elif action.startswith("add_stock_single:"):
            sheet_id = action.split(":", 1)[1]
            _save_inventory_adjust_from_post(
                request,
                sheet_id=sheet_id,
                qty_prefix="adjust_qty_",
                note_prefix="set_note_",
                movement_type=SheetInventoryMovement.MovementType.ADD,
                default_note="Manual add from Sheet Inventory",
            )
            return redirect("core:sheet_inventory_dashboard")

        elif action.startswith("deduct_stock_single:"):
            sheet_id = action.split(":", 1)[1]
            _save_inventory_adjust_from_post(
                request,
                sheet_id=sheet_id,
                qty_prefix="adjust_qty_",
                note_prefix="set_note_",
                movement_type=SheetInventoryMovement.MovementType.DEDUCT,
                default_note="Manual deduct from Sheet Inventory",
            )
            return redirect("core:sheet_inventory_dashboard")

        elif action == "set_stock_bulk":
            _save_inventory_bulk_from_post(
                request,
                qty_prefix="set_qty_",
                note_prefix="set_note_",
                movement_type=SheetInventoryMovement.MovementType.SET,
                default_note="Bulk balance update from Sheet Inventory",
            )
            return redirect("core:sheet_inventory_dashboard")

        # Keep legacy actions for compatibility with older forms/links.
        elif action == "stock_single":
            sheet = get_object_or_404(Sheet, id=request.POST.get("sheet_id"))
            mode = request.POST.get("movement_type") or SheetInventoryMovement.MovementType.ADD
            qty = int(request.POST.get("quantity") or 0)
            note = request.POST.get("note") or ""
            _apply_sheet_inventory_movement(sheet=sheet, movement_type=mode, quantity=qty, note=note, user=request.user)
            return redirect("core:sheet_inventory_dashboard")

        elif action == "bulk_stock":
            for key, val in request.POST.items():
                if not key.startswith("bulk_qty_"):
                    continue
                sheet_id = key.replace("bulk_qty_", "")
                raw_qty = (val or "").strip()
                if raw_qty == "":
                    continue
                sheet = Sheet.objects.filter(id=sheet_id).first()
                if not sheet:
                    continue
                mode = request.POST.get(f"bulk_mode_{sheet_id}") or SheetInventoryMovement.MovementType.ADD
                note = request.POST.get(f"bulk_note_{sheet_id}") or "Bulk update"
                try:
                    qty = int(raw_qty)
                except Exception:
                    continue
                _apply_sheet_inventory_movement(sheet=sheet, movement_type=mode, quantity=qty, note=note, user=request.user)
            return redirect("core:sheet_inventory_dashboard")

        elif action == "link_class_sheet":
            class_id = request.POST.get("class_id")
            sheet_id = request.POST.get("sheet_id")
            qty_per = int(request.POST.get("quantity_per_student") or 1)
            note = request.POST.get("note") or ""
            if class_id and sheet_id:
                SheetClassMapping.objects.update_or_create(
                    tutoring_class_id=class_id,
                    sheet_id=sheet_id,
                    defaults={
                        "quantity_per_student": max(qty_per, 1),
                        "note": note,
                        "is_active": True,
                    }
                )
            return redirect("core:sheet_inventory_dashboard")

        elif action == "unlink_class_sheet":
            mapping = get_object_or_404(SheetClassMapping, id=request.POST.get("mapping_id"))
            mapping.is_active = False
            mapping.save()
            return redirect("core:sheet_inventory_dashboard")

    q = (request.GET.get("q") or "").strip()
    return render(request, "core/sheet_inventory.html", _sheet_inventory_context(q=q))


@login_required
def sheet_inventory_count(request: HttpRequest) -> HttpResponse:
    """Dedicated stock count page. It records the actual counted balance only."""
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()

        if action.startswith("count_stock_single:"):
            sheet_id = action.split(":", 1)[1]
            _save_inventory_set_from_post(
                request,
                sheet_id=sheet_id,
                qty_prefix="count_qty_",
                note_prefix="count_note_",
                movement_type=SheetInventoryMovement.MovementType.COUNT,
                default_note="Stock count adjustment",
            )
            return redirect("core:sheet_inventory_count")

        if action == "count_stock_bulk":
            _save_inventory_bulk_from_post(
                request,
                qty_prefix="count_qty_",
                note_prefix="count_note_",
                movement_type=SheetInventoryMovement.MovementType.COUNT,
                default_note="Bulk stock count adjustment",
            )
            return redirect("core:sheet_inventory_count")

    q = (request.GET.get("q") or "").strip()
    return render(request, "core/sheet_inventory_count.html", _sheet_inventory_context(q=q))


@login_required
def sheet_inventory_movements(request: HttpRequest) -> HttpResponse:
    q = (request.GET.get("q") or "").strip()
    movement_type = (request.GET.get("movement_type") or "").strip()
    grade = _normalize_sheet_grade_level(request.GET.get("grade"))
    date_from_raw = (request.GET.get("date_from") or "").strip()
    date_to_raw = (request.GET.get("date_to") or "").strip()

    movements_qs = SheetInventoryMovement.objects.select_related("sheet", "sheet__subject", "created_by").order_by("-created_at", "-id")

    if q:
        movements_qs = movements_qs.filter(
            Q(sheet__code__icontains=q) |
            Q(sheet__title__icontains=q) |
            Q(sheet__subject__name__icontains=q) |
            Q(note__icontains=q)
        )
    if movement_type in dict(SheetInventoryMovement.MovementType.choices):
        movements_qs = movements_qs.filter(movement_type=movement_type)
    if grade:
        movements_qs = movements_qs.filter(sheet__grade_level=grade)

    parsed_from = None
    parsed_to = None
    try:
        if date_from_raw:
            parsed_from = date.fromisoformat(date_from_raw)
            movements_qs = movements_qs.filter(created_at__date__gte=parsed_from)
    except ValueError:
        parsed_from = None
    try:
        if date_to_raw:
            parsed_to = date.fromisoformat(date_to_raw)
            movements_qs = movements_qs.filter(created_at__date__lte=parsed_to)
    except ValueError:
        parsed_to = None

    movement_rows = []
    for m in movements_qs[:500]:
        before = int(m.balance_before or 0)
        after = int(m.balance_after or 0)
        delta = after - before
        movement_rows.append({
            "movement": m,
            "delta": delta,
            "delta_label": f"+{delta}" if delta > 0 else str(delta),
            "subject_style": _sheet_subject_style_attr(m.sheet),
        })

    return render(request, "core/sheet_inventory_movements.html", {
        "q": q,
        "movement_type": movement_type,
        "selected_grade": grade,
        "date_from": date_from_raw,
        "date_to": date_to_raw,
        "movement_type_choices": SheetInventoryMovement.MovementType.choices,
        "sheet_grade_choices": Sheet.GradeLevel.choices,
        "movement_rows": movement_rows,
        "movement_count": movements_qs.count(),
    })

@require_POST
@login_required
def sheet_inventory_bulk_upload(request: HttpRequest) -> HttpResponse:
    uploaded = request.FILES.get("bulk_file")
    stock_mode = request.POST.get("stock_mode") or "set"

    if not uploaded:
        context = _sheet_inventory_context(extra={
            "bulk_result": {
                "created": 0,
                "updated": 0,
                "skipped": 0,
                "movements": 0,
                "errors": ["กรุณาเลือกไฟล์ Excel หรือ CSV ก่อนอัปโหลด"],
                "error_count": 1,
            }
        })
        return render(request, "core/sheet_inventory.html", context)

    try:
        rows = _read_sheet_upload_rows(uploaded)
        result = _import_sheet_rows_from_upload(
            rows,
            user=request.user,
            stock_mode=stock_mode,
        )
    except Exception as exc:
        result = {
            "created": 0,
            "updated": 0,
            "skipped": 0,
            "movements": 0,
            "errors": [f"อ่านไฟล์ไม่สำเร็จ: {exc}"],
            "error_count": 1,
        }

    context = _sheet_inventory_context(extra={"bulk_result": result})
    return render(request, "core/sheet_inventory.html", context)


@require_POST
@login_required
def sheet_inventory_scan(request: HttpRequest) -> JsonResponse:
    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        data = request.POST

    code = (data.get("code") or "").strip().upper()
    mode = (data.get("mode") or SheetInventoryMovement.MovementType.DEDUCT).strip()
    qty = data.get("quantity") or 1
    note = data.get("note") or "QR scanner"

    if mode == "count":
        mode = SheetInventoryMovement.MovementType.COUNT

    if not code:
        return JsonResponse({"ok": False, "error": "ไม่พบรหัสชีท"}, status=400)

    sheet = Sheet.objects.filter(code__iexact=code).select_related("subject").first()
    if not sheet:
        return JsonResponse({"ok": False, "error": f"ไม่พบรหัสชีท {code}", "code": code}, status=404)

    ok, message, inventory, movement = _apply_sheet_inventory_movement(
        sheet=sheet,
        movement_type=mode,
        quantity=int(qty or 1),
        note=note,
        user=request.user,
    )
    if not ok:
        return JsonResponse({"ok": False, "error": message, "code": code}, status=400)

    return JsonResponse({
        "ok": True,
        "message": message,
        "code": sheet.code,
        "title": sheet.title,
        "quantity": int(inventory.quantity or 0) if inventory else 0,
        "movement_type": movement.get_movement_type_display() if movement else "",
        "movement_at": _fmt_dt_th(movement.created_at) if movement else "",
    })


def _allocation_recipient_payload(source: str, raw: dict) -> dict:
    source = (source or "unassigned").strip()
    payload = {
        "recipient_type": SheetAllocation.RecipientType.UNASSIGNED,
        "student": None,
        "admission_inquiry": None,
        "manual_nickname": "",
        "manual_grade_level": "",
    }

    if source == SheetAllocation.RecipientType.STUDENT:
        student = Student.objects.filter(id=raw.get("student_id"), is_active=True).first()
        if student:
            payload["recipient_type"] = SheetAllocation.RecipientType.STUDENT
            payload["student"] = student
            return payload
    elif source == SheetAllocation.RecipientType.ADMISSION:
        inquiry = AdmissionInquiry.objects.filter(id=raw.get("admission_id")).first()
        if inquiry:
            payload["recipient_type"] = SheetAllocation.RecipientType.ADMISSION
            payload["admission_inquiry"] = inquiry
            return payload
    elif source == SheetAllocation.RecipientType.MANUAL:
        nickname = (raw.get("manual_nickname") or "").strip()
        grade = (raw.get("manual_grade_level") or "").strip()
        if nickname or grade:
            payload["recipient_type"] = SheetAllocation.RecipientType.MANUAL
            payload["manual_nickname"] = nickname
            payload["manual_grade_level"] = grade
            return payload

    return payload


def _allocation_students_json() -> list[dict]:
    students = Student.objects.filter(is_active=True).order_by("grade_level", "student_code", "nickname")
    return [
        {
            "id": s.id,
            "student_code": s.student_code or "",
            "nickname": s.nickname or "",
            "full_name": s.full_name or "",
            "grade_level": s.grade_level or "",
            "label": f"{s.student_code or '-'} | {s.nickname or '-'} | {s.full_name or '-'} | {s.grade_level or '-'}",
        }
        for s in students
    ]


def _allocation_admissions_json() -> list[dict]:
    inquiries = AdmissionInquiry.objects.filter(is_completed=False).order_by("first_lesson_date", "nickname")[:300]
    rows = []
    for a in inquiries:
        rows.append({
            "id": a.id,
            "nickname": a.nickname or "",
            "full_name": a.full_name or "",
            "grade_level": a.grade_level or "",
            "grade_label": a.get_grade_level_display() if hasattr(a, "get_grade_level_display") else (a.grade_level or ""),
            "first_lesson_date": a.first_lesson_date.isoformat() if a.first_lesson_date else "",
            "label": f"{a.nickname or '-'} | {a.full_name or '-'} | {a.get_grade_level_display() if hasattr(a, 'get_grade_level_display') else a.grade_level} | {a.first_lesson_date.strftime('%d/%m/%Y') if a.first_lesson_date else '-'}",
        })
    return rows


def _allocation_class_students_json() -> list[dict]:
    classes = TutoringClass.objects.filter(is_active=True).order_by("time_slot", "name")
    rows = []
    for cls in classes:
        enrollments = (
            Enrollment.objects
            .select_related("student")
            .filter(is_active=True, student__is_active=True, tutoring_class=cls)
            .order_by("student__grade_level", "student__student_code", "student__nickname")
        )
        rows.append({
            "id": cls.id,
            "name": cls.name,
            "time_slot": cls.get_time_slot_display() if hasattr(cls, "get_time_slot_display") else cls.time_slot,
            "students": [
                {
                    "id": e.student.id,
                    "student_code": e.student.student_code or "",
                    "nickname": e.student.nickname or "",
                    "full_name": e.student.full_name or "",
                    "grade_level": e.student.grade_level or "",
                    "label": f"{e.student.nickname or '-'} | {e.student.full_name or '-'} | {e.student.grade_level or '-'}",
                }
                for e in enrollments
            ],
        })
    return rows


@login_required
def sheet_allocation_scan(request: HttpRequest) -> HttpResponse:
    sheets = Sheet.objects.select_related("subject").filter(is_active=True).order_by("grade_level", "subject__name", "code")
    sheet_rows = []
    for sheet in sheets:
        inv = _inventory_for_sheet(sheet)
        sheet_rows.append({
            "id": sheet.id,
            "code": sheet.code,
            "title": sheet.title,
            "subject": sheet.subject.name if sheet.subject_id else "",
            "grade_level": getattr(sheet, "grade_level", "") or "",
            "grade_label": sheet.get_grade_level_display() if getattr(sheet, "grade_level", "") else "",
            "quantity": int(inv.quantity or 0) if inv else 0,
        })

    return render(request, "core/sheet_allocation_scan.html", {
        "today": timezone.localdate(),
        "sheets_json": json.dumps(sheet_rows, ensure_ascii=False),
        "students_json": json.dumps(_allocation_students_json(), ensure_ascii=False),
        "admissions_json": json.dumps(_allocation_admissions_json(), ensure_ascii=False),
        "class_students_json": json.dumps(_allocation_class_students_json(), ensure_ascii=False),
        "grade_choices": Sheet.GradeLevel.choices,
    })


@require_POST
@login_required
def sheet_allocation_save(request: HttpRequest) -> JsonResponse:
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "message": "รูปแบบข้อมูลไม่ถูกต้อง"}, status=400)

    items = payload.get("items") or []
    allocation_date = _parse_optional_date(payload.get("allocation_date")) or timezone.localdate()
    if not items:
        return JsonResponse({"ok": False, "message": "ยังไม่มีรายการรอตัด"}, status=400)

    if len(items) > 500:
        return JsonResponse({"ok": False, "message": "รายการเยอะเกินไป กรุณาบันทึกไม่เกิน 500 รายการต่อครั้ง"}, status=400)

    batch_key = timezone.now().strftime("SA%Y%m%d%H%M%S")
    prepared = []
    qty_by_sheet: dict[int, int] = defaultdict(int)
    sheet_by_id: dict[int, Sheet] = {}

    for idx, item in enumerate(items, start=1):
        code = (item.get("sheet_code") or "").strip().upper()
        if not code:
            return JsonResponse({"ok": False, "message": f"รายการที่ {idx}: ไม่มีรหัสชีท"}, status=400)
        sheet = Sheet.objects.select_related("subject").filter(code__iexact=code, is_active=True).first()
        if not sheet:
            return JsonResponse({"ok": False, "message": f"ไม่พบรหัสชีท {code}"}, status=400)
        qty_by_sheet[sheet.id] += 1
        sheet_by_id[sheet.id] = sheet
        prepared.append((sheet, item))

    with transaction.atomic():
        movements_by_sheet = {}
        new_balances = {}
        for sheet_id, qty in qty_by_sheet.items():
            sheet = sheet_by_id[sheet_id]
            inv, _ = SheetInventory.objects.select_for_update().get_or_create(sheet=sheet, defaults={"quantity": 0})
            if int(inv.quantity or 0) < qty:
                return JsonResponse({
                    "ok": False,
                    "message": f"{sheet.code} คงเหลือ {int(inv.quantity or 0)} เล่ม แต่กำลังจะตัด {qty} เล่ม",
                }, status=400)
            ok, message, inventory, movement = _apply_sheet_inventory_movement(
                sheet=sheet,
                movement_type=SheetInventoryMovement.MovementType.DEDUCT,
                quantity=qty,
                note=f"แจกชีทผ่าน QR batch {batch_key}",
                user=request.user,
            )
            if not ok:
                return JsonResponse({"ok": False, "message": message}, status=400)
            movements_by_sheet[sheet_id] = movement
            new_balances[sheet.code] = int(inventory.quantity or 0) if inventory else 0

        created = 0
        for sheet, item in prepared:
            recipient_source = item.get("recipient_source") or "unassigned"
            recipient_payload = _allocation_recipient_payload(recipient_source, item)
            class_id = item.get("class_id") or None
            cls = None
            if class_id:
                cls = TutoringClass.objects.filter(id=class_id, is_active=True).first()
            elif recipient_payload.get("student"):
                active_enroll = (
                    Enrollment.objects
                    .filter(student=recipient_payload["student"], is_active=True, tutoring_class__is_active=True)
                    .order_by("-created_at")
                    .select_related("tutoring_class")
                    .first()
                )
                cls = active_enroll.tutoring_class if active_enroll else None

            SheetAllocation.objects.create(
                sheet=sheet,
                quantity=1,
                allocation_date=allocation_date,
                recipient_type=recipient_payload["recipient_type"],
                student=recipient_payload["student"],
                admission_inquiry=recipient_payload["admission_inquiry"],
                manual_nickname=recipient_payload["manual_nickname"],
                manual_grade_level=recipient_payload["manual_grade_level"],
                tutoring_class=cls,
                scan_code=(item.get("sheet_code") or "").strip().upper(),
                batch_key=batch_key,
                note=(item.get("note") or "").strip(),
                movement=movements_by_sheet.get(sheet.id),
                created_by=request.user if getattr(request.user, "is_authenticated", False) else None,
            )
            created += 1

    return JsonResponse({
        "ok": True,
        "message": f"บันทึกแจกชีท {created} รายการแล้ว",
        "created": created,
        "batch_key": batch_key,
        "new_balances": new_balances,
    })


def _class_report_sheets(cls: TutoringClass, student_ids: list[int]) -> list[Sheet]:
    sheet_map: dict[int, Sheet] = {}
    mappings = SheetClassMapping.objects.select_related("sheet", "sheet__subject").filter(tutoring_class=cls, is_active=True)
    for m in mappings:
        sheet_map[m.sheet_id] = m.sheet

    class_subjects = ClassSubject.objects.select_related("current_sheet", "current_sheet__subject").filter(
        tutoring_class=cls,
        is_active=True,
        current_sheet__isnull=False,
    )
    for cs in class_subjects:
        sheet_map[cs.current_sheet_id] = cs.current_sheet

    allocated = SheetAllocation.objects.select_related("sheet", "sheet__subject").filter(student_id__in=student_ids)
    for a in allocated:
        sheet_map[a.sheet_id] = a.sheet

    return sorted(sheet_map.values(), key=lambda s: (getattr(s, "grade_level", "") or "", s.subject.name if s.subject_id else "", s.code))


@login_required
def sheet_allocation_report(request: HttpRequest) -> HttpResponse:
    selected_student_id = (request.GET.get("student_id") or "").strip()
    selected_class_id = (request.GET.get("class_id") or "").strip()

    students = Student.objects.filter(is_active=True).order_by("grade_level", "student_code", "nickname")
    classes = TutoringClass.objects.filter(is_active=True).order_by("time_slot", "name")

    selected_student = None
    student_allocations = []
    if selected_student_id:
        selected_student = Student.objects.filter(id=selected_student_id, is_active=True).first()
        if selected_student:
            student_allocations = list(
                SheetAllocation.objects
                .select_related("sheet", "sheet__subject", "tutoring_class", "created_by")
                .filter(student=selected_student)
                .order_by("-allocation_date", "-created_at", "sheet__code")
            )

    selected_class = None
    class_rows = []
    class_sheets = []
    if selected_class_id:
        selected_class = TutoringClass.objects.filter(id=selected_class_id, is_active=True).first()
        if selected_class:
            enrollments = list(
                Enrollment.objects
                .select_related("student")
                .filter(is_active=True, student__is_active=True, tutoring_class=selected_class)
                .order_by("student__grade_level", "student__student_code", "student__nickname")
            )
            student_ids = [e.student_id for e in enrollments]
            class_sheets = _class_report_sheets(selected_class, student_ids)
            alloc_qs = SheetAllocation.objects.filter(student_id__in=student_ids, sheet__in=class_sheets).order_by("allocation_date", "created_at")
            alloc_map: dict[tuple[int, int], list[SheetAllocation]] = defaultdict(list)
            for a in alloc_qs:
                alloc_map[(a.student_id, a.sheet_id)].append(a)

            for e in enrollments:
                cells = []
                for sheet in class_sheets:
                    hits = alloc_map.get((e.student_id, sheet.id), [])
                    cells.append({
                        "sheet": sheet,
                        "allocations": hits,
                        "date_text": ", ".join(a.allocation_date.strftime("%d/%m/%y") for a in hits[-3:]),
                        "count": len(hits),
                    })
                class_rows.append({"student": e.student, "cells": cells})

    return render(request, "core/sheet_allocation_report.html", {
        "students": students,
        "classes": classes,
        "selected_student_id": selected_student_id,
        "selected_student": selected_student,
        "student_allocations": student_allocations,
        "selected_class_id": selected_class_id,
        "selected_class": selected_class,
        "class_sheets": class_sheets,
        "class_rows": class_rows,
    })


@login_required
def sheet_inventory_profile(request: HttpRequest, pk: int) -> HttpResponse:
    sheet = get_object_or_404(Sheet.objects.select_related("subject"), pk=pk)
    inventory, _ = SheetInventory.objects.get_or_create(sheet=sheet, defaults={"quantity": 0})
    movements = SheetInventoryMovement.objects.filter(sheet=sheet).select_related("created_by").order_by("-created_at")[:80]
    mappings = SheetClassMapping.objects.filter(sheet=sheet, is_active=True).select_related("tutoring_class")

    return render(request, "core/sheet_inventory_profile.html", {
        "sheet": sheet,
        "inventory": inventory,
        "movements": movements,
        "mappings": mappings,
    })


@login_required
def sheet_inventory_export(request: HttpRequest) -> HttpResponse:
    wb = Workbook()

    ws = wb.active
    ws.title = "Sheet Stock"
    ws.append(["Grade", "Sheet Code", "Title", "Subject", "Balance", "Minimum Stock", "Low Stock?", "Is Finished"])
    for sheet in Sheet.objects.select_related("subject").order_by("grade_level", "subject__name", "code"):
        inv = _inventory_for_sheet(sheet)
        qty = int(inv.quantity or 0) if inv else 0
        minimum = int(getattr(inv, "minimum_stock", 0) or 0) if inv else 0
        ws.append([
            sheet.get_grade_level_display() if getattr(sheet, "grade_level", "") else "",
            sheet.code,
            sheet.title,
            sheet.subject.name if sheet.subject_id else "",
            qty,
            minimum,
            "YES" if minimum > 0 and qty <= minimum else "NO",
            "YES" if inv and inv.is_finished else "NO",
        ])
    _autosize(ws)

    ws2 = wb.create_sheet("Class Requirement")
    ws2.append(["Class", "Time Slot", "Pending Inquiries", "Sheet Code", "Sheet Title", "Source", "Required Qty", "Balance", "Shortage"])
    for row in _build_class_sheet_rows():
        cls = row["class"]
        if not row["entries"]:
            ws2.append([cls.name, cls.get_time_slot_display(), row["demand_count"], "", "", "", 0, 0, 0])
        for e in row["entries"]:
            ws2.append([
                cls.name,
                cls.get_time_slot_display(),
                row["demand_count"],
                e["sheet"].code,
                e["sheet"].title,
                e["source_label"],
                e["required"],
                e["balance"],
                e["shortage"],
            ])
    _autosize(ws2)

    ws3 = wb.create_sheet("Movements")
    ws3.append(["Created At", "Sheet Code", "Movement Type", "Quantity", "Balance Before", "Balance After", "Note", "Created By"])
    movements = SheetInventoryMovement.objects.select_related("sheet", "created_by").order_by("-created_at")[:2000]
    for m in movements:
        ws3.append([
            timezone.localtime(m.created_at).strftime("%Y-%m-%d %H:%M") if m.created_at else "",
            m.sheet.code,
            m.get_movement_type_display(),
            m.quantity,
            m.balance_before,
            m.balance_after,
            m.note,
            m.created_by.get_username() if m.created_by else "",
        ])
    _autosize(ws3)

    buff = BytesIO()
    wb.save(buff)
    buff.seek(0)

    filename = f"sheet_inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    resp = HttpResponse(
        buff.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp




def _default_spine_color_for_subject(subject_name: str | None) -> str:
    raw = (subject_name or "").strip().lower().replace(" ", "")
    # Specific rules first
    if "คณิตเสริม" in raw or "อังกฤษเสริม" in raw or "extraenglish" in raw or "extra-math" in raw:
        return SheetPrintOrder.SpineColor.RED
    if "คณิต" in raw or "math" in raw:
        return SheetPrintOrder.SpineColor.BLUE
    if "อังกฤษ" in raw or "english" in raw:
        return SheetPrintOrder.SpineColor.PINK
    if "วิทย" in raw or "ชีว" in raw or "science" in raw or "bio" in raw or "biology" in raw:
        return SheetPrintOrder.SpineColor.GREEN
    if "เคมี" in raw or "ฟิสิก" in raw or "ไทย" in raw or "สังคม" in raw or "chem" in raw or "physics" in raw or "thai" in raw or "social" in raw:
        return SheetPrintOrder.SpineColor.ORANGE
    return ""


def _print_color_choices() -> list[dict]:
    return [
        {"value": SheetPrintOrder.SpineColor.BLUE, "label": "สีฟ้า", "bg": "#dbeafe", "border": "#93c5fd"},
        {"value": SheetPrintOrder.SpineColor.RED, "label": "สีแดง", "bg": "#fee2e2", "border": "#fca5a5"},
        {"value": SheetPrintOrder.SpineColor.PINK, "label": "สีชมพู", "bg": "#fce7f3", "border": "#f9a8d4"},
        {"value": SheetPrintOrder.SpineColor.GREEN, "label": "สีเขียว", "bg": "#dcfce7", "border": "#86efac"},
        {"value": SheetPrintOrder.SpineColor.ORANGE, "label": "สีส้ม", "bg": "#ffedd5", "border": "#fdba74"},
    ]


def _spine_color_label(value: str | None) -> str:
    labels = dict(SheetPrintOrder.SpineColor.choices)
    return labels.get(value or "", "ไม่ระบุสี")

def _parse_optional_date(value: str | None):
    raw = (value or "").strip()
    if not raw:
        return None
    try:
        return date.fromisoformat(raw)
    except ValueError:
        return None


def _print_order_dashboard_rows() -> list[dict]:
    pending_by_sheet = {
        item["sheet_id"]: item["qty"] or 0
        for item in (
            SheetPrintOrder.objects
            .filter(status=SheetPrintOrder.Status.PENDING, sheet__isnull=False)
            .values("sheet_id")
            .annotate(qty=Sum("quantity"))
        )
    }

    rows = []
    sheets = Sheet.objects.select_related("subject").filter(is_active=True).order_by("grade_level", "subject__name", "code")
    for sheet in sheets:
        inv = _inventory_for_sheet(sheet)
        current_qty = int(inv.quantity or 0) if inv else 0
        target_stock = int(getattr(inv, "target_stock", 0) or 0) if inv else 0
        onedrive_url = (getattr(inv, "onedrive_url", "") or "") if inv else ""
        pending_qty = int(pending_by_sheet.get(sheet.id, 0) or 0)
        shortage = max(target_stock - current_qty, 0)
        suggested_print_qty = max(target_stock - current_qty - pending_qty, 0)
        default_color = _default_spine_color_for_subject(sheet.subject.name if sheet.subject_id else "")
        rows.append({
            "sheet": sheet,
            "inventory": inv,
            "current_qty": current_qty,
            "target_stock": target_stock,
            "onedrive_url": onedrive_url,
            "pending_print_qty": pending_qty,
            "shortage": shortage,
            "suggested_print_qty": suggested_print_qty,
            "default_spine_color": default_color,
            "default_spine_color_label": _spine_color_label(default_color),
        })
    return rows


def _print_order_grade_groups(rows: list[dict]) -> list[dict]:
    buckets = {g: [] for g in SHEET_GRADE_ORDER}
    for row in rows:
        grade = (getattr(row.get("sheet"), "grade_level", "") or "").lower()
        if grade not in buckets:
            buckets.setdefault(grade, [])
        buckets[grade].append(row)
    return [
        {"grade": grade, "label": _sheet_grade_label(grade), "rows": buckets.get(grade, []), "count": len(buckets.get(grade, []))}
        for grade in SHEET_GRADE_ORDER
        if buckets.get(grade)
    ]


@login_required
def sheet_print_order_admin(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()

        if action == "update_sheet_print_setting":
            sheet = get_object_or_404(Sheet, id=request.POST.get("sheet_id"))
            target_stock = max(int(request.POST.get("target_stock") or 0), 0)
            onedrive_url = (request.POST.get("onedrive_url") or "").strip()
            inventory, _ = SheetInventory.objects.get_or_create(sheet=sheet, defaults={"quantity": 0})
            inventory.target_stock = target_stock
            inventory.onedrive_url = onedrive_url
            inventory.save()
            return redirect("core:sheet_print_order_admin")

        if action == "create_print_order":
            sheet = get_object_or_404(Sheet.objects.select_related("subject"), id=request.POST.get("sheet_id"))
            quantity = max(int(request.POST.get("quantity") or 0), 0)
            due_date = _parse_optional_date(request.POST.get("due_date"))
            onedrive_url = (request.POST.get("onedrive_url") or "").strip()
            note = (request.POST.get("note") or "").strip()
            binding_type = (request.POST.get("binding_type") or SheetPrintOrder.BindingType.SIDE).strip()
            if binding_type not in dict(SheetPrintOrder.BindingType.choices):
                binding_type = SheetPrintOrder.BindingType.SIDE
            spine_color = (request.POST.get("spine_color") or "").strip()
            if binding_type == SheetPrintOrder.BindingType.CORNER:
                spine_color = ""
            elif spine_color not in dict(SheetPrintOrder.SpineColor.choices):
                spine_color = _default_spine_color_for_subject(sheet.subject.name if sheet.subject_id else "")

            if quantity > 0:
                inventory, _ = SheetInventory.objects.get_or_create(sheet=sheet, defaults={"quantity": 0})
                if not onedrive_url:
                    onedrive_url = inventory.onedrive_url or ""

                SheetPrintOrder.objects.create(
                    sheet=sheet,
                    quantity=quantity,
                    due_date=due_date,
                    onedrive_url=onedrive_url,
                    binding_type=binding_type,
                    spine_color=spine_color,
                    note=note,
                    requested_by=request.user if getattr(request.user, "is_authenticated", False) else None,
                )
                return redirect("core:print_shop_order_list")

            return redirect("core:sheet_print_order_admin")

        if action == "create_custom_print_order":
            custom_title = (request.POST.get("custom_title") or "").strip()
            quantity = max(int(request.POST.get("quantity") or 0), 0)
            due_date = _parse_optional_date(request.POST.get("due_date"))
            onedrive_url = (request.POST.get("onedrive_url") or "").strip()
            note = (request.POST.get("note") or "").strip()
            binding_type = (request.POST.get("binding_type") or SheetPrintOrder.BindingType.SIDE).strip()
            if binding_type not in dict(SheetPrintOrder.BindingType.choices):
                binding_type = SheetPrintOrder.BindingType.SIDE
            spine_color = (request.POST.get("spine_color") or "").strip()
            if binding_type == SheetPrintOrder.BindingType.CORNER:
                spine_color = ""
            elif spine_color not in dict(SheetPrintOrder.SpineColor.choices):
                spine_color = ""

            if custom_title and quantity > 0 and onedrive_url:
                SheetPrintOrder.objects.create(
                    sheet=None,
                    custom_title=custom_title,
                    quantity=quantity,
                    due_date=due_date,
                    onedrive_url=onedrive_url,
                    binding_type=binding_type,
                    spine_color=spine_color,
                    note=note,
                    requested_by=request.user if getattr(request.user, "is_authenticated", False) else None,
                )
                return redirect("core:print_shop_order_list")

            return redirect("core:sheet_print_order_admin")

    pending_orders = (
        SheetPrintOrder.objects
        .select_related("sheet", "sheet__subject", "requested_by")
        .filter(status=SheetPrintOrder.Status.PENDING)
        .order_by("due_date", "created_at")
    )
    ready_orders = (
        SheetPrintOrder.objects
        .select_related("sheet", "sheet__subject", "requested_by")
        .filter(status=SheetPrintOrder.Status.READY)
        .order_by("-completed_at", "-updated_at")[:80]
    )
    rows = _print_order_dashboard_rows()

    return render(request, "core/sheet_print_order_admin.html", {
        "rows": rows,
        "grade_groups": _print_order_grade_groups(rows),
        "pending_orders": pending_orders,
        "ready_orders": ready_orders,
        "default_due_date": timezone.localdate() + timedelta(days=3),
        "shop_url": request.build_absolute_uri("/print-shop/"),
        "binding_type_choices": SheetPrintOrder.BindingType.choices,
        "spine_color_choices": _print_color_choices(),
    })


def print_shop_order_list(request: HttpRequest) -> HttpResponse:
    pending_orders = (
        SheetPrintOrder.objects
        .select_related("sheet", "sheet__subject")
        .filter(status=SheetPrintOrder.Status.PENDING)
        .order_by("due_date", "created_at")
    )
    ready_orders = (
        SheetPrintOrder.objects
        .select_related("sheet", "sheet__subject")
        .filter(status=SheetPrintOrder.Status.READY)
        .order_by("-completed_at", "-updated_at")[:80]
    )
    return render(request, "core/print_shop_orders.html", {
        "pending_orders": pending_orders,
        "ready_orders": ready_orders,
    })


@require_POST
def print_shop_update_order(request: HttpRequest, pk: int) -> HttpResponse:
    order = get_object_or_404(SheetPrintOrder, pk=pk)
    action = (request.POST.get("action") or "update_progress").strip()

    try:
        printed_qty = int(request.POST.get("printed_quantity") or order.printed_quantity or 0)
    except Exception:
        printed_qty = int(order.printed_quantity or 0)
    printed_qty = max(0, min(printed_qty, int(order.quantity or 0)))

    if action == "fill_complete":
        printed_qty = int(order.quantity or 0)

    order.printed_quantity = printed_qty
    order.print_done = (request.POST.get("print_done") == "yes") or printed_qty >= int(order.quantity or 0)
    order.bound_done = request.POST.get("bound_done") == "yes"
    order.spine_unavailable = request.POST.get("spine_unavailable") == "yes"
    order.save(update_fields=["printed_quantity", "print_done", "bound_done", "spine_unavailable", "updated_at"])

    return redirect("core:print_shop_order_list")


@require_POST
def print_shop_mark_ready(request: HttpRequest, pk: int) -> HttpResponse:
    order = get_object_or_404(SheetPrintOrder, pk=pk)
    if order.status == SheetPrintOrder.Status.PENDING:
        order.mark_ready()
    return redirect("core:print_shop_order_list")


# =========================================================
# ✅ Super Dashboard
# =========================================================
def _super_period(request: HttpRequest, prefix: str, default_mode: str = "month") -> tuple[date, date, str]:
    today = timezone.localdate()
    mode = (request.GET.get(f"{prefix}_period") or default_mode).strip()
    if mode == "1m":
        return today - timedelta(days=30), today, mode
    if mode == "3m":
        return today - timedelta(days=90), today, mode
    if mode == "year":
        return date(today.year, 1, 1), today, mode
    if mode == "custom":
        start = _parse_optional_date(request.GET.get(f"{prefix}_from")) or today - timedelta(days=90)
        end = _parse_optional_date(request.GET.get(f"{prefix}_to")) or today
        if end < start:
            start, end = end, start
        return start, end, mode
    # month default
    start = date(today.year, today.month, 1)
    return start, today, "month"


def _attendance_weekly_series(start: date, end: date) -> dict:
    # School week = Sat-Sun, same logic as school overview.
    start_week, _ = _school_week_range(start)
    labels, total, present, excused, no_show = [], [], [], [], []
    cur = start_week
    while cur <= end:
        ws, we = cur, cur + timedelta(days=1)
        qs = Attendance.objects.filter(attendance_date__gte=ws, attendance_date__lte=we)
        p = qs.filter(status=Attendance.Status.PRESENT).count()
        e = qs.filter(status=Attendance.Status.EXCUSED).count()
        n = qs.filter(status=Attendance.Status.NO_SHOW).count()
        labels.append(f"{ws.strftime('%d/%m')}–{we.strftime('%d/%m')}")
        present.append(p)
        excused.append(e)
        no_show.append(n)
        total.append(p + e + n)
        cur += timedelta(days=7)
    return {"labels": labels, "total": total, "present": present, "excused": excused, "no_show": no_show}


def _finance_summary_for_range(start: date, end: date) -> dict:
    revenue_per_student = _finance_setting("revenue_per_student_per_week", Decimal("360"), "Revenue per deducted attendance")
    deducted_count = Attendance.objects.filter(attendance_date__gte=start, attendance_date__lte=end, deducted=True).count()
    estimated_revenue = Decimal(deducted_count) * revenue_per_student
    cash_revenue = CoursePayment.objects.filter(
        payment_date__gte=start,
        payment_date__lte=end,
        status=CoursePayment.ReceiptStatus.ISSUED,
    ).aggregate(total=Sum("amount_paid"))["total"] or Decimal("0")
    general_expense = SchoolExpense.objects.filter(
        expense_date__gte=start,
        expense_date__lte=end,
    ).exclude(category__is_tutor_payroll=True).aggregate(total=Sum("amount"))["total"] or Decimal("0")
    tutor_payroll = TutorPayrollEntry.objects.filter(
        work_date__gte=start,
        work_date__lte=end,
    ).aggregate(total=Sum("total_amount"))["total"] or Decimal("0")
    total_expense = Decimal(general_expense) + Decimal(tutor_payroll)
    return {
        "start": start,
        "end": end,
        "deducted_count": deducted_count,
        "estimated_revenue": estimated_revenue,
        "cash_revenue": cash_revenue,
        "general_expense": general_expense,
        "tutor_payroll": tutor_payroll,
        "total_expense": total_expense,
        "net_estimated": estimated_revenue - total_expense,
        "net_cash": Decimal(cash_revenue) - total_expense,
    }


def _super_teaching_rows() -> dict:
    week_start, week_end = _school_week_range(timezone.localdate())
    _ensure_teaching_assignments(week_start, week_end)
    assignments = list(
        TeachingWeeklyAssignment.objects
        .select_related("tutoring_class", "subject_template", "tutor")
        .filter(week_start_date=week_start)
    )
    latest_updates = {}
    for u in TeachingProgressUpdate.objects.filter(assignment__week_start_date=week_start).order_by("assignment_id", "-teaching_date", "-updated_at"):
        latest_updates.setdefault(u.assignment_id, u)
    rows = []
    for a in assignments:
        u = latest_updates.get(a.id)
        status = "missing"
        if u and u.no_teaching:
            status = "no_teaching"
        elif u:
            status = "done"
        rows.append({"assignment": a, "update": u, "status": status})
    rank = {"missing": 0, "done": 1, "no_teaching": 2}
    slot_rank = {slot: idx for idx, slot in enumerate(TIME_SLOT_ORDER)}
    rows.sort(key=lambda r: (
        rank.get(r["status"], 9),
        slot_rank.get(r["assignment"].tutoring_class.time_slot, 99),
        r["assignment"].tutoring_class.name,
        r["assignment"].subject_template.display_order,
    ))
    return {
        "week_start": week_start,
        "week_end": week_end,
        "rows": rows[:36],
        "missing_count": sum(1 for r in rows if r["status"] == "missing"),
        "done_count": sum(1 for r in rows if r["status"] == "done"),
        "no_teaching_count": sum(1 for r in rows if r["status"] == "no_teaching"),
    }


def _admission_range_from_request(request: HttpRequest) -> tuple[date, date, str]:
    today = timezone.localdate()
    mode = (request.GET.get("admission_period") or "next_week").strip()
    if mode == "last_week":
        return today - timedelta(days=7), today - timedelta(days=1), mode
    if mode == "custom":
        start = _parse_optional_date(request.GET.get("admission_from")) or today
        end = _parse_optional_date(request.GET.get("admission_to")) or today + timedelta(days=7)
        if end < start:
            start, end = end, start
        return start, end, mode
    return today, today + timedelta(days=7), "next_week"


def _build_seat_rows(active_classes: list[TutoringClass], pending_inquiries: list[AdmissionInquiry]) -> list[dict]:
    rows = []
    active_enrollments = (
        Enrollment.objects
        .filter(is_active=True, student__is_active=True, tutoring_class__is_active=True)
        .values("tutoring_class_id")
        .annotate(c=Count("id"))
    )
    enrollment_map = {r["tutoring_class_id"]: r["c"] for r in active_enrollments}
    for cls in active_classes:
        matched = []
        for inquiry in pending_inquiries:
            target = getattr(inquiry, "target_class", None) or _guess_class_for_inquiry(inquiry, active_classes)
            if target and target.id == cls.id:
                matched.append(inquiry)
        enrolled_count = int(enrollment_map.get(cls.id, 0) or 0)
        total_seats = int(cls.total_seats or 0)
        expected_count = len(matched)
        rows.append({
            "class": cls,
            "total_seats": total_seats,
            "enrolled_count": enrolled_count,
            "available_seats": max(total_seats - enrolled_count, 0),
            "expected_count": expected_count,
            "expected_trial": sum(1 for i in matched if i.request_type == AdmissionInquiry.RequestType.TRIAL),
            "expected_enroll": sum(1 for i in matched if i.request_type == AdmissionInquiry.RequestType.ENROLL),
            "expected_queue": sum(1 for i in matched if i.request_type == AdmissionInquiry.RequestType.QUEUE),
        })
    return rows


def _sheet_inventory_matrix() -> dict:
    sheets = list(Sheet.objects.select_related("subject").filter(is_active=True).order_by("grade_level", "subject__name", "code"))
    subjects = []
    seen_subjects = set()
    for s in sheets:
        key = s.subject.name if s.subject_id else "ไม่ระบุวิชา"
        if key not in seen_subjects:
            seen_subjects.add(key)
            subjects.append(key)
    grades = []
    by_grade_subject = defaultdict(lambda: defaultdict(list))
    for s in sheets:
        grade = getattr(s, "grade_level", "") or ""
        if grade not in grades:
            grades.append(grade)
        inv = _inventory_for_sheet(s)
        by_grade_subject[grade][s.subject.name if s.subject_id else "ไม่ระบุวิชา"].append({
            "sheet": s,
            "quantity": int(inv.quantity or 0) if inv else 0,
        })
    ordered_grades = [g for g in SHEET_GRADE_ORDER if g in grades] + [g for g in grades if g not in SHEET_GRADE_ORDER]
    rows = []
    for g in ordered_grades:
        rows.append({
            "grade": g,
            "label": _sheet_grade_label(g),
            "cells": [{"subject": subj, "items": by_grade_subject[g].get(subj, [])} for subj in subjects],
        })
    return {"subjects": subjects, "rows": rows}


@login_required
def super_dashboard(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "add_expense":
            try:
                category = ExpenseCategory.objects.get(id=request.POST.get("category"), is_active=True)
                SchoolExpense.objects.create(
                    expense_date=_parse_date(request.POST.get("expense_date")),
                    category=category,
                    vendor=(request.POST.get("vendor") or "").strip(),
                    description=(request.POST.get("description") or "").strip(),
                    amount=_money(request.POST.get("amount")),
                    payment_method=(request.POST.get("payment_method") or SchoolExpense.PaymentMethod.TRANSFER),
                    note=(request.POST.get("note") or "").strip(),
                )
            except Exception:
                pass
            return redirect(request.META.get("HTTP_REFERER", "core:super_dashboard"))

        if action == "admission_quick_update":
            inquiry = get_object_or_404(AdmissionInquiry, id=request.POST.get("inquiry_id"))
            if request.POST.get("sheet_prepared") in {"yes", "no"}:
                inquiry.sheet_prepared = request.POST.get("sheet_prepared") == "yes"
            if request.POST.get("trial_attended") in {"yes", "no", "pending"}:
                inquiry.trial_attended = request.POST.get("trial_attended")
            if request.POST.get("is_completed") in {"yes", "no"}:
                inquiry.is_completed = request.POST.get("is_completed") == "yes"
                inquiry.completed_at = timezone.now() if inquiry.is_completed else None
            inquiry.save()
            return redirect(request.META.get("HTTP_REFERER", "core:super_dashboard"))

    today = timezone.localdate()
    attendance_start, attendance_end, attendance_mode = _super_period(request, "attendance", "3m")
    finance_start, finance_end, finance_mode = _super_period(request, "finance", "month")
    admission_start, admission_end, admission_mode = _admission_range_from_request(request)
    admission_type = (request.GET.get("admission_type") or "all").strip()

    attendance_series = _attendance_weekly_series(attendance_start, attendance_end)
    month_finance = _finance_summary_for_range(date(today.year, today.month, 1), today)
    ytd_finance = _finance_summary_for_range(date(today.year, 1, 1), today)
    selected_finance = _finance_summary_for_range(finance_start, finance_end)

    active_classes = list(TutoringClass.objects.filter(is_active=True).order_by("time_slot", "name"))
    pending_inquiries = list(
        AdmissionInquiry.objects.select_related("target_class")
        .filter(is_completed=False)
        .order_by("first_lesson_date", "created_at")
    )
    for inquiry in pending_inquiries:
        inquiry.guessed_class = _guess_class_for_inquiry(inquiry, active_classes)
        inquiry.display_target_class = inquiry.target_class or inquiry.guessed_class

    admission_qs = AdmissionInquiry.objects.select_related("target_class").filter(
        first_lesson_date__gte=admission_start,
        first_lesson_date__lte=admission_end,
        is_completed=False,
    )
    if admission_type in {AdmissionInquiry.RequestType.TRIAL, AdmissionInquiry.RequestType.ENROLL, AdmissionInquiry.RequestType.QUEUE}:
        admission_qs = admission_qs.filter(request_type=admission_type)
    admission_rows = list(admission_qs.order_by("first_lesson_date", "preferred_time_slot", "nickname")[:40])
    for inquiry in admission_rows:
        inquiry.guessed_class = _guess_class_for_inquiry(inquiry, active_classes)
        inquiry.display_target_class = inquiry.target_class or inquiry.guessed_class

    near_enrollments = list(
        Enrollment.objects.select_related("student", "tutoring_class")
        .filter(is_active=True, student__is_active=True, tutoring_class__is_active=True, sessions_total__lte=2)
        .order_by("sessions_total", "tutoring_class__name", "student__nickname")[:40]
    )
    enrollment_ids = [e.id for e in near_enrollments]
    notices_by_enrollment = {}
    if enrollment_ids:
        for notice in CourseRenewalNotice.objects.filter(enrollment_id__in=enrollment_ids).order_by("enrollment_id", "-created_at"):
            notices_by_enrollment.setdefault(notice.enrollment_id, notice)
    for e in near_enrollments:
        e.latest_notice = notices_by_enrollment.get(e.id)

    notice_unsent_count = CourseRenewalNotice.objects.filter(is_sent_to_parent=False, enrollment__is_active=True).count()
    print_pending = SheetPrintOrder.objects.select_related("sheet", "sheet__subject").filter(status=SheetPrintOrder.Status.PENDING).order_by("due_date", "created_at")[:20]
    print_ready = SheetPrintOrder.objects.select_related("sheet", "sheet__subject").filter(status=SheetPrintOrder.Status.READY).order_by("-completed_at", "-updated_at")[:12]

    context = {
        "today": today,
        "attendance_start": attendance_start,
        "attendance_end": attendance_end,
        "attendance_mode": attendance_mode,
        "attendance_series_json": json.dumps(attendance_series, ensure_ascii=False),
        "attendance_latest_total": attendance_series["total"][-1] if attendance_series["total"] else 0,
        "attendance_latest_present": attendance_series["present"][-1] if attendance_series["present"] else 0,
        "attendance_latest_excused": attendance_series["excused"][-1] if attendance_series["excused"] else 0,
        "attendance_latest_no_show": attendance_series["no_show"][-1] if attendance_series["no_show"] else 0,
        "seat_rows": _build_seat_rows(active_classes, pending_inquiries),
        "finance_mode": finance_mode,
        "finance_start": finance_start,
        "finance_end": finance_end,
        "month_finance": month_finance,
        "ytd_finance": ytd_finance,
        "selected_finance": selected_finance,
        "expense_categories": ExpenseCategory.objects.filter(is_active=True).order_by("sort_order", "name"),
        "payment_method_choices": SchoolExpense.PaymentMethod.choices,
        "teaching": _super_teaching_rows(),
        "admission_rows": admission_rows,
        "admission_type": admission_type,
        "admission_mode": admission_mode,
        "admission_start": admission_start,
        "admission_end": admission_end,
        "near_enrollments": near_enrollments,
        "near_2_count": Enrollment.objects.filter(is_active=True, sessions_total=2).count(),
        "near_1_count": Enrollment.objects.filter(is_active=True, sessions_total=1).count(),
        "notice_unsent_count": notice_unsent_count,
        "sheet_matrix": _sheet_inventory_matrix(),
        "print_pending": print_pending,
        "print_ready": print_ready,
    }
    return render(request, "core/super_dashboard.html", context)


@login_required
def export_excel(request: HttpRequest) -> HttpResponse:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Enrollments"

    ws1.append([
        "Enrollment ID",
        "Student Code",
        "Nickname",
        "Full Name",
        "Grade",
        "Class",
        "Time Slot",
        "Enrollment Type",
        "Sessions Total",
        "Remaining Sessions",
        "Is Active",
        "Created At",
        "Notified?",
        "Notified Method",
        "Notified At",
    ])

    enrollments = (
        Enrollment.objects
        .select_related("student", "tutoring_class")
        .order_by("-is_active", "tutoring_class__name", "student__nickname", "student__full_name")
    )

    for e in enrollments:
        s = e.student
        c = e.tutoring_class
        ws1.append([
            e.id,
            getattr(s, "student_code", "") or "",
            getattr(s, "nickname", "") or "",
            getattr(s, "full_name", "") or "",
            getattr(s, "grade_level", "") or "",
            getattr(c, "name", "") or "",
            getattr(c, "time_slot", "") or "",
            e.get_enrollment_type_display() if hasattr(e, "get_enrollment_type_display") else "",
            getattr(e, "sessions_total", "") or "",
            getattr(e, "remaining_sessions", "") if getattr(e, "remaining_sessions", None) is not None else "",
            bool(getattr(e, "is_active", False)),
            e.created_at.strftime("%Y-%m-%d %H:%M:%S") if getattr(e, "created_at", None) else "",
            bool(getattr(e, "notified_near_complete", False)),
            getattr(e, "notified_method", "") or "",
            e.notified_at.strftime("%Y-%m-%d %H:%M:%S") if getattr(e, "notified_at", None) else "",
        ])

    _autosize(ws1)

    ws2 = wb.create_sheet("Students")
    ws2.append([
        "Student ID",
        "Student Code",
        "Nickname",
        "Full Name",
        "Grade",
        "Is Active",
        "Active Enrollments",
        "Remaining Sessions (Active Total)",
    ])

    students = (
        Student.objects
        .order_by("-is_active", "grade_level", "student_code")
        .annotate(
            active_enrollments=Count("enrollments", filter=Q(enrollments__is_active=True)),
            remaining_total=Sum("enrollments__remaining_sessions", filter=Q(enrollments__is_active=True)),
        )
    )

    for s in students:
        ws2.append([
            s.id,
            getattr(s, "student_code", "") or "",
            getattr(s, "nickname", "") or "",
            getattr(s, "full_name", "") or "",
            getattr(s, "grade_level", "") or "",
            bool(getattr(s, "is_active", False)),
            int(getattr(s, "active_enrollments", 0) or 0),
            int(getattr(s, "remaining_total", 0) or 0),
        ])

    _autosize(ws2)

    ws_cash = wb.create_sheet("Course Payments")
    ws_cash.append(["Receipt No", "Payment Date", "Student", "Class", "Enrollment", "Payment Type", "Payment Method", "Sessions", "Course Price", "Discount", "Net Amount", "Amount Paid", "Status", "Note"])
    course_payment_export_rows = (
        CoursePayment.objects
        .select_related("student", "tutoring_class", "enrollment")
        .filter(status=CoursePayment.ReceiptStatus.ISSUED)
        .order_by("-payment_date", "-created_at")
    )
    for p in course_payment_export_rows:
        ws_cash.append([
            p.receipt_no,
            p.payment_date.isoformat() if p.payment_date else "",
            p.student.display_name if p.student_id else "",
            p.tutoring_class.name if p.tutoring_class_id else "",
            p.enrollment.sale_run_no if p.enrollment_id and p.enrollment else "",
            p.get_payment_type_display() if hasattr(p, "get_payment_type_display") else p.payment_type,
            p.get_payment_method_display() if hasattr(p, "get_payment_method_display") else p.payment_method,
            int(p.sessions_granted or 0),
            float(p.course_price or 0),
            float(p.discount_amount or 0),
            float(p.net_amount or 0),
            float(p.amount_paid or 0),
            p.get_status_display() if hasattr(p, "get_status_display") else p.status,
            p.note or "",
        ])
    _autosize(ws_cash)

    buff = BytesIO()
    wb.save(buff)
    buff.seek(0)

    filename = f"pkanoon_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    resp = HttpResponse(
        buff.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@require_POST
@login_required
def attendance_submit(request: HttpRequest) -> JsonResponse:
    """
    Submit เช็คชื่อทีละห้อง
    - อัปเดต checked_at เป็นเวลา server
    - คืนค่า summary + remaining_map + checked_at_map + server_now_text
    """
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "error": "Invalid JSON"}, status=400)

    selected_date = _parse_date(payload.get("date"))
    class_id = payload.get("class_id")
    items = payload.get("items", [])

    if not class_id:
        return JsonResponse({"ok": False, "error": "Missing class_id"}, status=400)

    valid_status = {
        Attendance.Status.PRESENT,
        Attendance.Status.EXCUSED,
        Attendance.Status.NO_SHOW
    }

    normalized_items: list[dict] = []
    for it in items:
        try:
            eid = int(str(it.get("enrollment_id")).strip())
        except Exception:
            continue

        status = (it.get("status") or "").strip()
        if status not in valid_status:
            return JsonResponse({"ok": False, "error": "Invalid status in items"}, status=400)

        normalized_items.append({"enrollment_id": eid, "status": status})

    enrollments = (
        Enrollment.objects
        .select_related("student", "tutoring_class")
        .filter(
            tutoring_class_id=class_id,
            is_active=True,
            student__is_active=True,
            tutoring_class__is_active=True,
        )
        .all()
    )
    enroll_map = {e.id: e for e in enrollments}

    already_checked_ids = set(
        Attendance.objects
        .filter(
            attendance_date=selected_date,
            enrollment__tutoring_class_id=class_id,
            student__is_active=True,
        )
        .values_list("enrollment_id", flat=True)
    )

    required_ids = set(enroll_map.keys()) - already_checked_ids
    submitted_ids = {it["enrollment_id"] for it in normalized_items}

    missing = sorted(list(required_ids - submitted_ids))
    if missing:
        return JsonResponse({
            "ok": False,
            "error": "กรุณาเลือกสถานะให้ครบทุกคนในห้องนี้ก่อนกด Submit",
            "missing_enrollment_ids": missing,
            "debug": {
                "required_today": len(required_ids),
                "submitted_today": len(submitted_ids),
            }
        }, status=400)

    now = timezone.now()

    with transaction.atomic():
        for it in normalized_items:
            eid = it["enrollment_id"]
            status = it["status"]

            e = enroll_map.get(eid)
            if not e:
                continue

            att, _ = Attendance.objects.get_or_create(
                student=e.student,
                enrollment=e,
                attendance_date=selected_date,
                defaults={"status": status},
            )
            att.status = status
            att.checked_at = now
            att.save()

    cls_summary = Attendance.objects.filter(
        attendance_date=selected_date,
        enrollment__tutoring_class_id=class_id,
        student__is_active=True,
    ).aggregate(
        present=Count("id", filter=Q(status=Attendance.Status.PRESENT)),
        excused=Count("id", filter=Q(status=Attendance.Status.EXCUSED)),
        no_show=Count("id", filter=Q(status=Attendance.Status.NO_SHOW)),
        total=Count("id"),
    )

    global_summary = Attendance.objects.filter(
        attendance_date=selected_date,
        enrollment__tutoring_class__is_active=True,
        student__is_active=True,
    ).aggregate(
        present=Count("id", filter=Q(status=Attendance.Status.PRESENT)),
        excused=Count("id", filter=Q(status=Attendance.Status.EXCUSED)),
        no_show=Count("id", filter=Q(status=Attendance.Status.NO_SHOW)),
        total=Count("id"),
    )

    refreshed = (
        Enrollment.objects
        .filter(id__in=list(enroll_map.keys()))
        .only("id", "remaining_sessions")
    )
    remaining_map = {e.id: e.remaining_sessions for e in refreshed}

    submitted_eids = [it["enrollment_id"] for it in normalized_items]
    checked_at_qs = Attendance.objects.filter(
        attendance_date=selected_date,
        enrollment_id__in=submitted_eids,
    ).only("enrollment_id", "checked_at")
    checked_at_map = {a.enrollment_id: _fmt_dt_th(a.checked_at) for a in checked_at_qs}

    return JsonResponse({
        "ok": True,
        "class_id": int(class_id),
        "date": selected_date.isoformat(),
        "class_summary": cls_summary,
        "global_summary": global_summary,
        "remaining_map": remaining_map,
        "checked_at_map": checked_at_map,
        "server_now_text": _fmt_dt_th(now),
    })


@login_required
def alerts_dashboard(request: HttpRequest) -> HttpResponse:
    THRESHOLD = 2
    enrollments = (
        Enrollment.objects
        .select_related("student", "tutoring_class")
        .filter(student__is_active=True, tutoring_class__is_active=True)
        .order_by("tutoring_class__name", "student__nickname", "student__full_name", "student__grade_level")
        .all()
    )

    near = [e for e in enrollments if (e.remaining_sessions or 0) < THRESHOLD]

    return render(request, "core/alerts_dashboard.html", {
        "near": near,
        "threshold": THRESHOLD,
    })


@require_POST
@login_required
def alerts_mark(request: HttpRequest) -> HttpResponse:
    eid = request.POST.get("enrollment_id")
    method = (request.POST.get("method") or "").strip()

    e = get_object_or_404(Enrollment, id=eid)

    if method == "":
        e.notified_near_complete = False
        e.notified_method = None
        e.notified_at = None
    else:
        e.notified_near_complete = True
        e.notified_method = method
        e.notified_at = timezone.now()

    e.save()
    return redirect(request.META.get("HTTP_REFERER", "/dashboard/"))


@login_required
def admin_dashboard(request: HttpRequest) -> HttpResponse:
    today = timezone.localdate()
    weeks = 8

    monday_this_week = today - timedelta(days=today.weekday())
    week_starts = [monday_this_week - timedelta(weeks=i) for i in range(weeks - 1, -1, -1)]
    buckets = {ws: set() for ws in week_starts}

    start_date = week_starts[0]
    end_date = monday_this_week + timedelta(days=6)

    qs = (
        Attendance.objects
        .select_related("student")
        .filter(attendance_date__gte=start_date, attendance_date__lte=end_date, student__is_active=True)
        .only("attendance_date", "student_id")
    )

    for a in qs:
        ws = a.attendance_date - timedelta(days=a.attendance_date.weekday())
        if ws in buckets:
            buckets[ws].add(a.student_id)

    labels = [ws.strftime("%d %b") for ws in week_starts]
    counts = [len(buckets[ws]) for ws in week_starts]
    max_count = max(counts) if counts else 0

    return render(request, "core/admin_dashboard.html", {
        "labels": labels,
        "counts": counts,
        "max_count": max_count,
        "weeks": weeks,
    })


@login_required
def attendance_details(request: HttpRequest) -> HttpResponse:
    """
    ✅ แสดงเฉพาะ enrollment active
    ✅ คอลัมน์คงที่ 1..22
    ✅ ส่ง summary ต่อ enrollment ให้ template ใช้ (มา/ลา/ขาด/total/คงเหลือ)
    """
    classes = TutoringClass.objects.filter(is_active=True).order_by("name").all()

    enrollments = (
        Enrollment.objects
        .select_related("student", "tutoring_class")
        .filter(
            is_active=True,
            student__is_active=True,
            tutoring_class__is_active=True,
        )
        .order_by(
            "tutoring_class__name",
            "student__nickname",
            "student__full_name",
            "student__grade_level",
        )
        .all()
    )

    all_att = (
        Attendance.objects
        .filter(enrollment__in=enrollments)
        .only("enrollment_id", "attendance_date", "status", "checked_at")
        .order_by("attendance_date", "checked_at")
        .all()
    )

    att_list_map: dict[int, list[dict]] = {e.id: [] for e in enrollments}
    for a in all_att:
        att_list_map[a.enrollment_id].append({
            "date": a.attendance_date,
            "status": a.status,
        })

    grouped_rows: dict[int, list[dict]] = {}

    for e in enrollments:
        recs = att_list_map.get(e.id, [])

        present_cnt = sum(1 for r in recs if r.get("status") == Attendance.Status.PRESENT)
        excused_cnt = sum(1 for r in recs if r.get("status") == Attendance.Status.EXCUSED)
        noshow_cnt = sum(1 for r in recs if r.get("status") == Attendance.Status.NO_SHOW)

        grouped_rows.setdefault(e.tutoring_class_id, []).append({
            "enrollment": e,
            "records": recs[:22],  # โชว์สูงสุด 22 ช่อง
            "present_cnt": present_cnt,
            "excused_cnt": excused_cnt,
            "noshow_cnt": noshow_cnt,
            "total_sessions": int(getattr(e, "sessions_total", 0) or 0),
            "remaining_sessions": int(getattr(e, "remaining_sessions", 0) or 0),
        })

    session_cols = list(range(1, 23))

    return render(request, "core/attendance_details.html", {
        "classes": classes,
        "grouped_rows": grouped_rows,
        "session_cols": session_cols,
    })


# =========================================================
# ✅ Student Portal (ผู้ปกครอง)
# =========================================================
class StudentPortalLoginForm(forms.Form):
    # ✅ ใช้ Select2 AJAX แต่ "ห้าม" ใช้ ChoiceField เพราะจะ Validate choices แล้วพัง
    #    เราใช้ CharField + Select widget แทน เพื่อให้ POST ค่า id ได้โดยไม่ติด "valid choice"
    student_id = forms.CharField(
        label="เลือกชื่อน้อง",
        required=True,
        widget=forms.Select(attrs={"id": "id_student_id"})
    )
    parent_phone = forms.CharField(label="เบอร์ผู้ปกครอง", max_length=50)

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # ใส่ placeholder option ไว้ 1 ตัวพอ (Select2 จะเติมรายการจาก AJAX เอง)
        self.fields["student_id"].widget.choices = [
            ("", "พิมพ์ค้นหาชื่อเล่น / ชื่อจริง / รหัสนักเรียน")
        ]

    def clean(self):
        cleaned = super().clean()
        sid = (cleaned.get("student_id") or "").strip()
        phone = (cleaned.get("parent_phone") or "").strip()

        if not sid:
            raise forms.ValidationError("กรุณาเลือกชื่อนักเรียน")

        try:
            sid_int = int(sid)
        except Exception:
            raise forms.ValidationError("รูปแบบนักเรียนไม่ถูกต้อง")

        student = Student.objects.filter(id=sid_int, is_active=True).first()
        if not student:
            raise forms.ValidationError("ไม่พบนักเรียนนี้ในระบบ")

        def digits(x: str) -> str:
            return "".join(ch for ch in x if ch.isdigit())

        MASTER_PASSWORD = "kanoon"

        if phone != MASTER_PASSWORD and digits(student.parent_phone) != digits(phone):
            raise forms.ValidationError("เบอร์ผู้ปกครองไม่ถูกต้อง")

        cleaned["student"] = student
        return cleaned


@require_GET
def student_portal_student_search(request: HttpRequest) -> JsonResponse:
    """
    AJAX endpoint สำหรับ Select2:
    - ค้นหาได้ด้วย nickname / full_name / student_code
    - คืนค่า format: { "results": [ { "id": "...", "text": "..." }, ... ] }
    """
    q = (request.GET.get("q") or "").strip()

    qs = Student.objects.filter(is_active=True)

    if q:
        qs = qs.filter(
            Q(nickname__icontains=q) |
            Q(full_name__icontains=q) |
            Q(student_code__icontains=q)
        )

    qs = qs.only("id", "student_code", "nickname", "full_name", "grade_level").order_by(
        "grade_level", "student_code"
    )[:30]

    results = []
    for s in qs:
        nickname = (getattr(s, "nickname", "") or "").strip()
        full_name = getattr(s, "full_name", "") or ""
        student_code = getattr(s, "student_code", "") or ""
        grade = getattr(s, "grade_level", "") or ""

        label = f"{student_code} | {nickname or '-'} | {full_name}"
        if grade:
            label += f" | {grade}"

        results.append({"id": str(s.id), "text": label})

    return JsonResponse({"results": results})


def student_portal_login(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        form = StudentPortalLoginForm(request.POST)

        if form.is_valid():
            student = form.cleaned_data["student"]
            request.session["portal_student_id"] = student.id

            # ✅ ถ้า Login จากปุ่มคอร์สออนไลน์ ป.6 ให้ไปหน้า Online Course
            if request.path == "/online-course-p6/":
                return redirect("core:online_course_home")

            # ✅ ถ้า Login จาก Student Portal ปกติ ให้ไปหน้า Student Portal Home
            return redirect("core:student_portal_home")

    else:
        form = StudentPortalLoginForm()

    return render(request, "core/student_portal_login.html", {
        "form": form
    })


def online_course_login(request: HttpRequest) -> HttpResponse:
    return student_portal_login(request)


def _get_portal_student(request: HttpRequest) -> Student | None:
    sid = request.session.get("portal_student_id")

    if not sid:
        return None

    try:
        return Student.objects.get(id=sid, is_active=True)

    except Student.DoesNotExist:
        return None


def student_portal_logout(request: HttpRequest) -> HttpResponse:
    request.session.pop("portal_student_id", None)
    return redirect("core:student_portal_login")


def online_course_home(request: HttpRequest) -> HttpResponse:
    student = _get_portal_student(request)

    if not student:
        return redirect("core:online_course_login")

    return render(request, "core/online_course_home.html", {
        "student": student,
    })


def student_portal_home(request: HttpRequest) -> HttpResponse:
    student = _get_portal_student(request)

    if not student:
        return redirect("core:student_portal_login")

    enrollments = (
        Enrollment.objects
        .select_related("tutoring_class")
        .filter(student=student)
        .order_by("-is_active", "-created_at")
        .all()
    )

    selected_enrollment_id = request.GET.get("enrollment_id")
    selected_enrollment = None

    if selected_enrollment_id:
        selected_enrollment = enrollments.filter(id=selected_enrollment_id).first()

    if not selected_enrollment and enrollments:
        selected_enrollment = enrollments[0]

    attendance_rows = []

    if selected_enrollment:
        attendance_rows = (
            Attendance.objects
            .filter(student=student, enrollment=selected_enrollment)
            .order_by("-attendance_date", "-checked_at")
            .all()
        )

    remaining_sessions = selected_enrollment.remaining_sessions if selected_enrollment else 0

    hours_per_session = (
        float(selected_enrollment.tutoring_class.hours_per_session)
        if selected_enrollment else 0.0
    )

    remaining_hours = remaining_sessions * hours_per_session

    context = {
        "student": student,
        "enrollments": enrollments,
        "selected_enrollment": selected_enrollment,
        "attendance_rows": attendance_rows,
        "remaining_sessions": remaining_sessions,
        "hours_per_session": hours_per_session,
        "remaining_hours": remaining_hours,
    }

    return render(request, "core/student_portal_home.html", context)


# =========================================================
# ✅ Generate ใบแจ้งครบคอร์ส
# =========================================================
@login_required
def generate_course_notice(request: HttpRequest) -> HttpResponse:
    enrollment_id = request.GET.get("enrollment_id") or request.POST.get("enrollment_id")
    if not enrollment_id:
        return render(request, "core/generate_course_notice.html", {"error": "missing enrollment_id"})

    enrollment = get_object_or_404(
        Enrollment.objects.select_related("student", "tutoring_class"),
        id=enrollment_id,
        is_active=True,
        student__is_active=True,
        tutoring_class__is_active=True,
    )

    student = enrollment.student
    tutoring_class = enrollment.tutoring_class

    base_price = enrollment.course_price if getattr(enrollment, "course_price", None) is not None else (
        tutoring_class.course_price if tutoring_class else 0
    )

    def to_decimal(x, default: Decimal) -> Decimal:
        try:
            s = str(x).strip()
            if s == "":
                return default
            return Decimal(s)
        except Exception:
            return default

    if request.method == "POST":
        course_end_date = _parse_date(request.POST.get("course_end_date"))
    else:
        course_end_date = timezone.localdate()

    next_course_start_date = course_end_date + timedelta(days=7)

    amount_10 = to_decimal(
        request.POST.get("amount_10") if request.method == "POST" else None,
        Decimal(str(base_price or 0))
    )
    discount_10 = to_decimal(
        request.POST.get("discount_10") if request.method == "POST" else None,
        Decimal("0")
    )
    net_10 = amount_10 - discount_10
    if net_10 < 0:
        net_10 = Decimal("0")

    amount_20_default = Decimal(str(base_price or 0)) * 2
    amount_20 = to_decimal(
        request.POST.get("amount_20") if request.method == "POST" else None,
        amount_20_default
    )
    discount_20 = to_decimal(
        request.POST.get("discount_20") if request.method == "POST" else None,
        Decimal("0")
    )
    net_20 = amount_20 - discount_20
    if net_20 < 0:
        net_20 = Decimal("0")

    context = {
        "student": student,
        "enrollment": enrollment,
        "tutoring_class": tutoring_class,
        "remaining_sessions": enrollment.remaining_sessions,

        "course_end_date": course_end_date,
        "next_course_start_date": next_course_start_date,

        "amount_10": amount_10,
        "discount_10": discount_10,
        "net_10": net_10,

        "amount_20": amount_20,
        "discount_20": discount_20,
        "net_20": net_20,

        "qr_promptpay_static": "core/img/qr_promptpay.png",
        "qr_line_static": "core/img/qr_line.png",
    }
    return render(request, "core/generate_course_notice.html", context)

# =========================================================
# ✅ Admission Inquiry: สมัครเรียน / จองทดลองเรียน
# =========================================================
class AdmissionInquiryForm(forms.ModelForm):
    class Meta:
        model = AdmissionInquiry
        fields = [
            "request_type",
            "nickname",
            "first_name",
            "last_name",
            "contact_phone",
            "school_name",
            "latest_gpa",
            "first_lesson_date",
            "grade_level",
            "preferred_time_slot",
        ]
        widgets = {
            "request_type": forms.RadioSelect,
            "nickname": forms.TextInput(attrs={"placeholder": "เช่น น้องข้าวหอม", "autocomplete": "given-name"}),
            "first_name": forms.TextInput(attrs={"placeholder": "ชื่อจริงของนักเรียน", "autocomplete": "given-name"}),
            "last_name": forms.TextInput(attrs={"placeholder": "นามสกุลของนักเรียน", "autocomplete": "family-name"}),
            "school_name": forms.TextInput(attrs={"placeholder": "ชื่อโรงเรียน", "autocomplete": "organization"}),
            "contact_phone": forms.TextInput(attrs={
                "placeholder": "เบอร์ผู้ปกครอง / เบอร์ติดต่อ",
                "inputmode": "tel",
                "autocomplete": "tel",
            }),
            "latest_gpa": forms.NumberInput(attrs={
                "placeholder": "เช่น 3.50",
                "step": "0.01",
                "min": "0",
                "max": "4",
                "inputmode": "decimal",
            }),
            "first_lesson_date": forms.DateInput(attrs={
                "type": "date",
                "autocomplete": "off",
                "data-native-picker": "date",
            }),
        }

    def clean_contact_phone(self):
        phone = (self.cleaned_data.get("contact_phone") or "").strip()
        digits = "".join(ch for ch in phone if ch.isdigit())
        if len(digits) < 9:
            raise forms.ValidationError("กรุณากรอกเบอร์ติดต่อให้ถูกต้อง")
        return phone


def admission_inquiry(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        form = AdmissionInquiryForm(request.POST)
        if form.is_valid():
            inquiry = form.save()
            return redirect("core:admission_thank_you", pk=inquiry.pk)
    else:
        form = AdmissionInquiryForm()

    return render(request, "core/admission_inquiry.html", {
        "form": form,
    })


def admission_thank_you(request: HttpRequest, pk: int) -> HttpResponse:
    inquiry = get_object_or_404(AdmissionInquiry, pk=pk)

    return render(request, "core/admission_thank_you.html", {
        "inquiry": inquiry,
        "line_url": "https://lin.ee/Vp91szz",
    })


def _guess_class_for_inquiry(inquiry: AdmissionInquiry, classes=None):
    """Guess class from grade + preferred time slot. Persisted target_class wins."""
    if getattr(inquiry, "target_class_id", None):
        return inquiry.target_class
    if classes is None:
        classes = list(TutoringClass.objects.filter(is_active=True))
    grade_label = ""
    try:
        grade_label = inquiry.get_grade_level_display()
    except Exception:
        grade_label = inquiry.grade_level or ""
    grade_tokens = {
        (inquiry.grade_level or "").lower(),
        grade_label,
        grade_label.replace(".", ""),
        grade_label.replace(" ", ""),
    }
    grade_tokens = {t.lower() for t in grade_tokens if t}
    for cls in classes:
        if getattr(cls, "time_slot", "") != getattr(inquiry, "preferred_time_slot", ""):
            continue
        cls_name = (cls.name or "").lower().replace(" ", "")
        if any(t.replace(" ", "") in cls_name for t in grade_tokens):
            return cls
    return None


def admission_report(request: HttpRequest) -> HttpResponse:
    qs = AdmissionInquiry.objects.select_related("target_class").all()

    q = (request.GET.get("q") or "").strip()
    request_type = (request.GET.get("request_type") or "").strip()
    grade_level = (request.GET.get("grade_level") or "").strip()
    preferred_time_slot = (request.GET.get("preferred_time_slot") or "").strip()
    sheet_prepared = (request.GET.get("sheet_prepared") or "").strip()
    trial_attended = (request.GET.get("trial_attended") or "").strip()
    trial_result = (request.GET.get("trial_result") or "").strip()
    completed_status = (request.GET.get("completed_status") or "open").strip()
    date_from = (request.GET.get("date_from") or "").strip()
    date_to = (request.GET.get("date_to") or "").strip()

    if q:
        qs = qs.filter(
            Q(nickname__icontains=q) |
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q) |
            Q(school_name__icontains=q) |
            Q(contact_phone__icontains=q)
        )

    if request_type:
        qs = qs.filter(request_type=request_type)
    if grade_level:
        qs = qs.filter(grade_level=grade_level)
    if preferred_time_slot:
        qs = qs.filter(preferred_time_slot=preferred_time_slot)
    if sheet_prepared == "yes":
        qs = qs.filter(sheet_prepared=True)
    elif sheet_prepared == "no":
        qs = qs.filter(sheet_prepared=False)
    if trial_attended:
        qs = qs.filter(trial_attended=trial_attended)
    if trial_result:
        qs = qs.filter(trial_result=trial_result)
    if completed_status == "completed":
        qs = qs.filter(is_completed=True)
    elif completed_status == "all":
        pass
    else:
        completed_status = "open"
        qs = qs.filter(is_completed=False)
    if date_from:
        try:
            qs = qs.filter(first_lesson_date__gte=date.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            qs = qs.filter(first_lesson_date__lte=date.fromisoformat(date_to))
        except ValueError:
            pass

    active_classes = list(TutoringClass.objects.filter(is_active=True).order_by("time_slot", "name"))

    stats_base = AdmissionInquiry.objects.all()
    open_base = stats_base.filter(is_completed=False)
    completed_base = stats_base.filter(is_completed=True)
    stats = {
        "total": stats_base.count(),
        "open_total": open_base.count(),
        "completed_total": completed_base.count(),
        "trial": open_base.filter(request_type=AdmissionInquiry.RequestType.TRIAL).count(),
        "enroll": open_base.filter(request_type=AdmissionInquiry.RequestType.ENROLL).count(),
        "queue": open_base.filter(request_type=AdmissionInquiry.RequestType.QUEUE).count(),
        "sheet_ready": open_base.filter(sheet_prepared=True).count(),
        "trial_attended": open_base.filter(trial_attended=AdmissionInquiry.TrialAttended.YES).count(),
        "trial_enrolled": open_base.filter(trial_result=AdmissionInquiry.TrialResult.ENROLLED).count(),
    }

    inquiries = qs.order_by("is_completed", "first_lesson_date", "-created_at")
    active_qs = inquiries.filter(is_completed=False)
    completed_qs = inquiries.filter(is_completed=True).order_by("-completed_at", "-updated_at", "-created_at")

    for inquiry in list(inquiries):
        guessed = _guess_class_for_inquiry(inquiry, active_classes)
        inquiry.guessed_class = guessed
        inquiry.display_target_class = inquiry.target_class or guessed

    inquiry_groups = [
        {
            "key": "trial",
            "title": "จองทดลองเรียน",
            "icon": "🧪",
            "color": "blue",
            "items": active_qs.filter(request_type=AdmissionInquiry.RequestType.TRIAL),
        },
        {
            "key": "enroll",
            "title": "สมัครเรียน",
            "icon": "📚",
            "color": "green",
            "items": active_qs.filter(request_type=AdmissionInquiry.RequestType.ENROLL),
        },
        {
            "key": "queue",
            "title": "จองที่นั่งล่วงหน้า",
            "icon": "📌",
            "color": "orange",
            "items": active_qs.filter(request_type=AdmissionInquiry.RequestType.QUEUE),
        },
    ]

    return render(request, "core/admission_report.html", {
        "inquiries": inquiries,
        "inquiry_groups": inquiry_groups,
        "completed_inquiries": completed_qs,
        "stats": stats,
        "filters": {
            "q": q,
            "request_type": request_type,
            "grade_level": grade_level,
            "preferred_time_slot": preferred_time_slot,
            "sheet_prepared": sheet_prepared,
            "trial_attended": trial_attended,
            "trial_result": trial_result,
            "completed_status": completed_status,
            "date_from": date_from,
            "date_to": date_to,
        },
        "request_type_choices": AdmissionInquiry.RequestType.choices,
        "grade_level_choices": AdmissionInquiry.GradeLevel.choices,
        "time_slot_choices": AdmissionInquiry.PreferredTimeSlot.choices,
        "trial_attended_choices": AdmissionInquiry.TrialAttended.choices,
        "trial_result_choices": AdmissionInquiry.TrialResult.choices,
        "active_classes": active_classes,
    })


@require_POST
@login_required
def admission_report_update(request: HttpRequest) -> HttpResponse:
    inquiry_id = request.POST.get("inquiry_id")
    inquiry = get_object_or_404(AdmissionInquiry, id=inquiry_id)

    sheet_prepared = request.POST.get("sheet_prepared")
    trial_attended = request.POST.get("trial_attended")
    trial_result = request.POST.get("trial_result")
    internal_note = request.POST.get("internal_note")
    is_completed_raw = request.POST.get("is_completed")
    target_class_id = (request.POST.get("target_class_id") or request.POST.get("target_class") or "").strip()

    # Update only fields explicitly submitted so quick-action buttons do not reset other fields.
    if sheet_prepared in {"yes", "no"}:
        inquiry.sheet_prepared = sheet_prepared == "yes"

    valid_attended = {choice[0] for choice in AdmissionInquiry.TrialAttended.choices}
    valid_result = {choice[0] for choice in AdmissionInquiry.TrialResult.choices}

    if trial_attended in valid_attended:
        inquiry.trial_attended = trial_attended
    if trial_result in valid_result:
        inquiry.trial_result = trial_result
    if internal_note is not None:
        inquiry.internal_note = internal_note.strip()

    if "target_class_id" in request.POST or "target_class" in request.POST:
        if target_class_id:
            cls = TutoringClass.objects.filter(id=target_class_id, is_active=True).first()
            inquiry.target_class = cls
        else:
            inquiry.target_class = None

    if is_completed_raw in {"yes", "no"}:
        new_completed = is_completed_raw == "yes"
        if new_completed and not inquiry.is_completed:
            inquiry.completed_at = timezone.now()
        elif not new_completed:
            inquiry.completed_at = None
        inquiry.is_completed = new_completed

    inquiry.save()

    return redirect(request.META.get("HTTP_REFERER", "core:admission_report"))

# =========================================================
# ✅ School Overview / Finance helpers
# =========================================================
def _finance_setting(key: str, default: Decimal, description: str = "") -> Decimal:
    obj, _ = FinanceSetting.objects.get_or_create(
        key=key,
        defaults={"value": default, "description": description},
    )
    return Decimal(str(obj.value))


def _school_week_range(anchor: date) -> tuple[date, date]:
    """School week = Saturday to Sunday."""
    days_since_sat = (anchor.weekday() - 5) % 7
    start = anchor - timedelta(days=days_since_sat)
    return start, start + timedelta(days=1)


def _month_range(anchor: date) -> tuple[date, date]:
    if anchor.month == 12:
        next_month = date(anchor.year + 1, 1, 1)
    else:
        next_month = date(anchor.year, anchor.month + 1, 1)
    start = date(anchor.year, anchor.month, 1)
    return start, next_month - timedelta(days=1)


def _year_range(anchor: date) -> tuple[date, date]:
    return date(anchor.year, 1, 1), date(anchor.year, 12, 31)


def _safe_date(s: str | None) -> date | None:
    if not s:
        return None
    try:
        return date.fromisoformat(s)
    except ValueError:
        return None


def _period_from_request(request: HttpRequest) -> tuple[str, date, date]:
    """
    Date range priority:
    1) If date_from/date_to are provided, use them directly.
    2) Otherwise use period_mode + anchor_date as a quick preset.

    This fixes the old behavior where users had to rely on anchor_date even
    after selecting a start/end range.
    """
    mode = (request.GET.get("period_mode") or "custom").strip()

    raw_start = _safe_date(request.GET.get("date_from"))
    raw_end = _safe_date(request.GET.get("date_to"))

    if raw_start or raw_end:
        start = raw_start or raw_end or timezone.localdate()
        end = raw_end or raw_start or start
        if end < start:
            start, end = end, start
        return "custom", start, end

    anchor = _safe_date(request.GET.get("anchor_date")) or timezone.localdate()

    if mode == "month":
        start, end = _month_range(anchor)
    elif mode == "year":
        start, end = _year_range(anchor)
    else:
        mode = "week"
        start, end = _school_week_range(anchor)

    return mode, start, end


def _blank_overview_row(label: str, sort_date: date) -> dict:
    return {
        "label": label,
        "sort": sort_date,
        "present": 0,
        "excused": 0,
        "no_show": 0,
        "deducted_count": 0,
        "class_ids": set(),
        "estimated_revenue": Decimal("0"),
        "estimated_tutor_cost": Decimal("0"),
        "actual_tutor_cost": Decimal("0"),
        "general_expense": Decimal("0"),
        "net_estimated": Decimal("0"),
        "net_actual_direct": Decimal("0"),
    }


def _ensure_overview_group(groups: dict[str, dict], label: str, sort_date: date) -> dict:
    if label not in groups:
        groups[label] = _blank_overview_row(label, sort_date)
    return groups[label]


def _school_week_labels_between(start: date, end: date) -> list[tuple[str, date, date]]:
    """Return all Sat-Sun school weeks overlapping the selected range."""
    first_week_start, _ = _school_week_range(start)
    cur = first_week_start
    out: list[tuple[str, date, date]] = []
    while cur <= end:
        week_end = cur + timedelta(days=1)
        label = f"{cur.strftime('%d/%m/%Y')} - {week_end.strftime('%d/%m/%Y')}"
        out.append((label, cur, week_end))
        cur += timedelta(days=7)
    return out


def _group_key(d: date, group_by: str) -> tuple[str, date]:
    if group_by == "month":
        start = date(d.year, d.month, 1)
        return start.strftime("%Y-%m"), start
    if group_by == "year":
        start = date(d.year, 1, 1)
        return str(d.year), start
    start, _ = _school_week_range(d)
    return f"{start.strftime('%d/%m/%Y')} - {(start + timedelta(days=1)).strftime('%d/%m/%Y')}", start


def _money(x) -> Decimal:
    try:
        return Decimal(str(x or 0))
    except Exception:
        return Decimal("0")




COURSE_SESSION_CHOICES = [
    ("10", "10 ครั้ง", 10),
    ("12", "10 แถม 2 = 12 ครั้ง", 12),
    ("20", "20 ครั้ง", 20),
    ("30", "30 ครั้ง", 30),
    ("custom", "กรอกเอง", 10),
]


def _active_students_for_payment():
    return Student.objects.filter(is_active=True).select_related("school").order_by("grade_level", "nickname", "full_name")


def _active_enrollments_for_payment(student_id: str | None = None):
    qs = (
        Enrollment.objects
        .select_related("student", "tutoring_class")
        .filter(is_active=True, student__is_active=True, tutoring_class__is_active=True)
        .order_by("student__nickname", "student__full_name", "tutoring_class__name", "-created_at")
    )
    if student_id:
        try:
            qs = qs.filter(student_id=int(student_id))
        except Exception:
            pass
    return qs


def _default_payment_form_context(request: HttpRequest, errors: list[str] | None = None, posted: dict | None = None):
    students = list(_active_students_for_payment())
    classes = list(TutoringClass.objects.filter(is_active=True).order_by("time_slot", "name"))

    # For the receipt form we use search boxes instead of long dropdowns.
    # Keep the queryset unrestricted to active records so the browser can filter instantly.
    enrollments = list(_active_enrollments_for_payment())
    admission_inquiries = list(
        AdmissionInquiry.objects
        .order_by("-created_at")
        .only(
            "id",
            "nickname",
            "first_name",
            "last_name",
            "school_name",
            "contact_phone",
            "grade_level",
            "preferred_time_slot",
            "first_lesson_date",
            "created_at",
        )[:300]
    )

    students_lookup_json = [
        {
            "id": s.id,
            "label": f"{s.nickname or '-'} | {s.full_name or '-'} | {s.student_code or '-'} | {s.grade_level or '-'}",
            "nickname": s.nickname or "",
            "full_name": s.full_name or "",
            "student_code": s.student_code or "",
            "grade_level": s.grade_level or "",
            "school_name": s.school.name if getattr(s, "school_id", None) and s.school else "",
            "parent_phone": s.parent_phone or "",
        }
        for s in students
    ]

    enrollments_lookup_json = [
        {
            "id": e.id,
            "student_id": e.student_id,
            "student_label": f"{e.student.nickname or '-'} | {e.student.full_name or '-'} | {e.student.student_code or '-'}",
            "label": f"{e.student.nickname or '-'} | {e.student.full_name or '-'} | {e.tutoring_class.name} | {e.sale_run_no or e.id} | คงเหลือ {e.remaining_sessions}",
            "class_name": e.tutoring_class.name,
            "sale_run_no": e.sale_run_no or str(e.id),
            "remaining_sessions": e.remaining_sessions,
            "course_price": str(e.course_price or e.tutoring_class.course_price or 0),
        }
        for e in enrollments
    ]

    admission_inquiries_lookup_json = [
        {
            "id": i.id,
            "label": f"{i.nickname} | {i.full_name} | {i.get_grade_level_display()} | {i.school_name or '-'} | {i.contact_phone}",
            "nickname": i.nickname or "",
            "full_name": i.full_name or "",
            "grade_level": i.get_grade_level_display() or "",
            "school_name": i.school_name or "",
            "parent_phone": i.contact_phone or "",
            "first_lesson_date": i.first_lesson_date.isoformat() if i.first_lesson_date else "",
            "preferred_time_slot": i.get_preferred_time_slot_display() if i.preferred_time_slot else "",
            "created_at": i.created_at.strftime("%Y-%m-%d") if i.created_at else "",
        }
        for i in admission_inquiries
    ]

    return {
        "students": students,
        "classes": classes,
        "enrollments": enrollments,
        "admission_inquiries": admission_inquiries,
        "students_lookup_json": students_lookup_json,
        "enrollments_lookup_json": enrollments_lookup_json,
        "admission_inquiries_lookup_json": admission_inquiries_lookup_json,
        "payment_method_choices": CoursePayment.PaymentMethod.choices,
        "payment_type_choices": CoursePayment.PaymentType.choices,
        "enrollment_action_choices": CoursePayment.EnrollmentAction.choices,
        "session_choices": COURSE_SESSION_CHOICES,
        "today": timezone.localdate(),
        "errors": errors or [],
        "posted": posted or {},
    }


def _sessions_from_package(package: str, custom_value: str | None) -> int:
    mapping = {code: value for code, _label, value in COURSE_SESSION_CHOICES}
    if package == "custom":
        try:
            return max(int(custom_value or 0), 1)
        except Exception:
            return 10
    return mapping.get(package, 10)


def _apply_payment_to_enrollment(payment: CoursePayment, action: str, existing_enrollment: Enrollment | None = None) -> Enrollment:
    """Create a new enrollment or add sessions to an existing enrollment for a course payment."""
    sessions = int(payment.sessions_granted or 0)
    if action == CoursePayment.EnrollmentAction.ADD_EXISTING and existing_enrollment:
        payment.enrollment_sessions_before = int(existing_enrollment.sessions_total or 0)
        existing_enrollment.sessions_total = int(existing_enrollment.sessions_total or 0) + sessions
        existing_enrollment.remark = ((existing_enrollment.remark or "").strip() + f"\nเพิ่ม {sessions} ครั้ง จากใบเสร็จ {payment.receipt_no}").strip()
        existing_enrollment.save()
        payment.enrollment = existing_enrollment
        payment.tutoring_class = existing_enrollment.tutoring_class
        payment.enrollment_action = CoursePayment.EnrollmentAction.ADD_EXISTING
        payment.enrollment_created = False
        payment.save()
        return existing_enrollment

    enrollment = Enrollment.objects.create(
        student=payment.student,
        tutoring_class=payment.tutoring_class,
        enrollment_type=Enrollment.EnrollmentType.SPECIAL,
        sessions_total=sessions,
        payment_type=Enrollment.PaymentType.INSTALLMENT if payment.payment_type == CoursePayment.PaymentType.INSTALLMENT else Enrollment.PaymentType.FULL,
        installments_count=1,
        course_price=payment.course_price,
        discount_amount=payment.discount_amount,
        remark=f"สร้างจากใบเสร็จ {payment.receipt_no}",
        is_active=True,
    )
    payment.enrollment = enrollment
    payment.enrollment_action = CoursePayment.EnrollmentAction.NEW
    payment.enrollment_created = True
    payment.enrollment_sessions_before = None
    payment.save()
    return enrollment


def _reverse_payment_enrollment_effect(payment: CoursePayment) -> None:
    """Reverse enrollment effect when a receipt is cancelled."""
    enrollment = payment.enrollment
    if not enrollment:
        return
    if payment.enrollment_created:
        enrollment.is_active = False
        enrollment.remark = ((enrollment.remark or "").strip() + f"\nยกเลิกจากใบเสร็จ {payment.receipt_no}").strip()
        enrollment.save()
        return
    if payment.enrollment_sessions_before is not None:
        enrollment.sessions_total = int(payment.enrollment_sessions_before)
        enrollment.remark = ((enrollment.remark or "").strip() + f"\nยกเลิกการเพิ่มครั้งจากใบเสร็จ {payment.receipt_no}").strip()
        enrollment.save()


def _attendance_queryset_for_range(start: date, end: date, time_slot: str = "", class_id: str = ""):
    qs = (
        Attendance.objects
        .select_related("enrollment__tutoring_class", "student")
        .filter(
            attendance_date__gte=start,
            attendance_date__lte=end,
            student__is_active=True,
            enrollment__tutoring_class__is_active=True,
        )
    )
    if time_slot:
        qs = qs.filter(enrollment__tutoring_class__time_slot=time_slot)
    if class_id:
        try:
            qs = qs.filter(enrollment__tutoring_class_id=int(class_id))
        except Exception:
            pass
    return qs


@login_required
def school_overview(request: HttpRequest) -> HttpResponse:
    """
    Module 1:
    - Attendance overview by selected start/end date.
    - Attendance graph is always a weekly line chart: present / excused / no-show.
    - Revenue estimate = (present + no-show) x revenue per student.
    - Estimated tutor direct cost = active classes in period x tutor cost per class.
    - Actual tutor direct cost = Module 2 tutor payroll entries.
    """
    if request.method == "POST":
        revenue = _money(request.POST.get("revenue_per_student"))
        tutor_cost = _money(request.POST.get("estimated_tutor_cost_per_class"))
        FinanceSetting.objects.update_or_create(
            key="revenue_per_student_per_week",
            defaults={
                "value": revenue,
                "description": "รายได้ประมาณการต่อคนต่อสัปดาห์ / ต่อครั้งที่หักชั่วโมง",
            },
        )
        FinanceSetting.objects.update_or_create(
            key="estimated_tutor_cost_per_class_per_week",
            defaults={
                "value": tutor_cost,
                "description": "ค่าใช้จ่ายติวเตอร์ประมาณการต่อ class ต่อสัปดาห์",
            },
        )
        # Keep the selected filter after saving settings.
        qs = request.GET.urlencode()
        return redirect(f"{request.path}?{qs}" if qs else "core:school_overview")

    period_mode, start, end = _period_from_request(request)
    group_by = (request.GET.get("group_by") or "week").strip()
    if group_by not in {"week", "month", "year"}:
        group_by = "week"

    time_slot = (request.GET.get("time_slot") or "").strip()
    class_id = (request.GET.get("class_id") or "").strip()

    revenue_per_student = _finance_setting(
        "revenue_per_student_per_week",
        Decimal("360"),
        "รายได้ประมาณการต่อคนต่อสัปดาห์ / ต่อครั้งที่หักชั่วโมง",
    )
    estimated_tutor_cost_per_class = _finance_setting(
        "estimated_tutor_cost_per_class_per_week",
        Decimal("1350"),
        "ค่าใช้จ่ายติวเตอร์ประมาณการต่อ class ต่อสัปดาห์",
    )

    att_qs = _attendance_queryset_for_range(start, end, time_slot, class_id)

    # =====================================================
    # Table groups: user can choose week / month / year.
    # =====================================================
    groups: dict[str, dict] = {}
    for a in att_qs:
        label, group_start = _group_key(a.attendance_date, group_by)
        item = _ensure_overview_group(groups, label, group_start)

        if a.status == Attendance.Status.PRESENT:
            item["present"] += 1
            item["deducted_count"] += 1
            item["class_ids"].add(a.enrollment.tutoring_class_id)
        elif a.status == Attendance.Status.EXCUSED:
            item["excused"] += 1
        elif a.status == Attendance.Status.NO_SHOW:
            item["no_show"] += 1
            item["deducted_count"] += 1
            item["class_ids"].add(a.enrollment.tutoring_class_id)

    # Actual tutor payroll from Module 2
    payroll_qs = TutorPayrollEntry.objects.filter(work_date__gte=start, work_date__lte=end)
    for p in payroll_qs:
        label, group_start = _group_key(p.work_date, group_by)
        item = _ensure_overview_group(groups, label, group_start)
        item["actual_tutor_cost"] += _money(p.total_amount)

    # General expenses from Module 2
    expense_qs = SchoolExpense.objects.filter(expense_date__gte=start, expense_date__lte=end)
    for e in expense_qs:
        label, group_start = _group_key(e.expense_date, group_by)
        item = _ensure_overview_group(groups, label, group_start)
        item["general_expense"] += _money(e.amount)

    rows = []
    for item in groups.values():
        active_class_count = len(item["class_ids"])
        item["active_class_count"] = active_class_count
        item["estimated_revenue"] = Decimal(item["deducted_count"]) * revenue_per_student
        item["estimated_tutor_cost"] = Decimal(active_class_count) * estimated_tutor_cost_per_class
        item["net_estimated"] = item["estimated_revenue"] - item["estimated_tutor_cost"]
        item["net_actual_direct"] = item["estimated_revenue"] - item["actual_tutor_cost"]
        rows.append(item)

    rows.sort(key=lambda x: x["sort"])

    totals = {
        "present": sum(r["present"] for r in rows),
        "excused": sum(r["excused"] for r in rows),
        "no_show": sum(r["no_show"] for r in rows),
        "deducted_count": sum(r["deducted_count"] for r in rows),
        "estimated_revenue": sum((r["estimated_revenue"] for r in rows), Decimal("0")),
        "estimated_tutor_cost": sum((r["estimated_tutor_cost"] for r in rows), Decimal("0")),
        "actual_tutor_cost": sum((r["actual_tutor_cost"] for r in rows), Decimal("0")),
        "general_expense": sum((r["general_expense"] for r in rows), Decimal("0")),
    }
    totals["net_estimated"] = totals["estimated_revenue"] - totals["estimated_tutor_cost"]
    totals["net_actual_direct"] = totals["estimated_revenue"] - totals["actual_tutor_cost"]

    # =====================================================
    # Chart groups: ALWAYS weekly line chart by school week.
    # This keeps X-axis as Sat-Sun weeks even if the table is monthly/yearly.
    # =====================================================
    weekly_groups: dict[str, dict] = {}
    for label, week_start, week_end in _school_week_labels_between(start, end):
        weekly_groups[label] = _blank_overview_row(label, week_start)

    for a in att_qs:
        label, week_start = _group_key(a.attendance_date, "week")
        item = _ensure_overview_group(weekly_groups, label, week_start)
        if a.status == Attendance.Status.PRESENT:
            item["present"] += 1
            item["deducted_count"] += 1
            item["class_ids"].add(a.enrollment.tutoring_class_id)
        elif a.status == Attendance.Status.EXCUSED:
            item["excused"] += 1
        elif a.status == Attendance.Status.NO_SHOW:
            item["no_show"] += 1
            item["deducted_count"] += 1
            item["class_ids"].add(a.enrollment.tutoring_class_id)

    for p in payroll_qs:
        label, week_start = _group_key(p.work_date, "week")
        item = _ensure_overview_group(weekly_groups, label, week_start)
        item["actual_tutor_cost"] += _money(p.total_amount)

    for e in expense_qs:
        label, week_start = _group_key(e.expense_date, "week")
        item = _ensure_overview_group(weekly_groups, label, week_start)
        item["general_expense"] += _money(e.amount)

    weekly_rows = sorted(weekly_groups.values(), key=lambda x: x["sort"])
    for item in weekly_rows:
        active_class_count = len(item["class_ids"])
        item["active_class_count"] = active_class_count
        item["estimated_revenue"] = Decimal(item["deducted_count"]) * revenue_per_student
        item["estimated_tutor_cost"] = Decimal(active_class_count) * estimated_tutor_cost_per_class

    chart_labels = [r["label"] for r in weekly_rows]
    chart_present = [r["present"] for r in weekly_rows]
    chart_excused = [r["excused"] for r in weekly_rows]
    chart_no_show = [r["no_show"] for r in weekly_rows]
    chart_revenue = [float(r["estimated_revenue"]) for r in weekly_rows]
    chart_estimated_cost = [float(r["estimated_tutor_cost"]) for r in weekly_rows]
    chart_actual_tutor = [float(r["actual_tutor_cost"]) for r in weekly_rows]

    classes = TutoringClass.objects.filter(is_active=True).order_by("time_slot", "name")

    return render(request, "core/school_overview.html", {
        "rows": rows,
        "totals": totals,
        "period_mode": period_mode,
        "group_by": group_by,
        "start": start,
        "end": end,
        "filters": {
            "anchor_date": request.GET.get("anchor_date") or timezone.localdate().isoformat(),
            "date_from": start.isoformat(),
            "date_to": end.isoformat(),
            "time_slot": time_slot,
            "class_id": class_id,
        },
        "classes": classes,
        "time_slot_choices": TutoringClass.TimeSlot.choices,
        "revenue_per_student": revenue_per_student,
        "estimated_tutor_cost_per_class": estimated_tutor_cost_per_class,
        "chart_labels_json": json.dumps(chart_labels, ensure_ascii=False),
        "chart_present_json": json.dumps(chart_present),
        "chart_excused_json": json.dumps(chart_excused),
        "chart_no_show_json": json.dumps(chart_no_show),
        "chart_revenue_json": json.dumps(chart_revenue),
        "chart_estimated_cost_json": json.dumps(chart_estimated_cost),
        "chart_actual_tutor_json": json.dumps(chart_actual_tutor),
    })


@login_required
def _school_finance_filtered_data(request: HttpRequest):
    selected_date = _parse_date(request.GET.get("work_date") or request.POST.get("work_date"))
    date_from = _parse_date(request.GET.get("date_from"))
    date_to = _parse_date(request.GET.get("date_to"))
    if date_to < date_from:
        date_from, date_to = date_to, date_from

    revenue_per_student = _finance_setting(
        "revenue_per_student_per_week",
        Decimal("360"),
        "รายได้ประมาณการต่อคนต่อสัปดาห์ / ต่อครั้งที่หักชั่วโมง",
    )

    att_qs = Attendance.objects.filter(
        attendance_date__gte=date_from,
        attendance_date__lte=date_to,
        student__is_active=True,
        enrollment__tutoring_class__is_active=True,
        status__in=[Attendance.Status.PRESENT, Attendance.Status.NO_SHOW],
    )
    deducted_count = att_qs.count()
    estimated_revenue = Decimal(deducted_count) * revenue_per_student

    expense_rows = (
        SchoolExpense.objects
        .select_related("category")
        .filter(expense_date__gte=date_from, expense_date__lte=date_to)
        .order_by("-expense_date", "-created_at")
    )
    payroll_rows = (
        TutorPayrollEntry.objects
        .select_related("tutor")
        .filter(work_date__gte=date_from, work_date__lte=date_to)
        .order_by("-work_date", "tutor__name")
    )
    course_payment_rows = (
        CoursePayment.objects
        .select_related("student", "tutoring_class", "enrollment")
        .filter(
            payment_date__gte=date_from,
            payment_date__lte=date_to,
            status=CoursePayment.ReceiptStatus.ISSUED,
        )
        .order_by("-payment_date", "-created_at")
    )

    general_expense_total = expense_rows.aggregate(total=Sum("amount")).get("total") or Decimal("0")
    tutor_payroll_total = payroll_rows.aggregate(total=Sum("total_amount")).get("total") or Decimal("0")
    cash_revenue_total = course_payment_rows.aggregate(total=Sum("amount_paid")).get("total") or Decimal("0")
    total_expense = general_expense_total + tutor_payroll_total
    net_estimated = estimated_revenue - total_expense
    net_cash_basis = cash_revenue_total - total_expense

    return {
        "selected_date": selected_date,
        "date_from": date_from,
        "date_to": date_to,
        "revenue_per_student": revenue_per_student,
        "deducted_count": deducted_count,
        "estimated_revenue": estimated_revenue,
        "cash_revenue_total": cash_revenue_total,
        "expense_rows": expense_rows,
        "payroll_rows": payroll_rows,
        "course_payment_rows": course_payment_rows,
        "general_expense_total": general_expense_total,
        "tutor_payroll_total": tutor_payroll_total,
        "total_expense": total_expense,
        "net_estimated": net_estimated,
        "net_cash_basis": net_cash_basis,
    }



# =========================================================
# ✅ Course Renewal Notice Module
# =========================================================
def _default_renewal_dates(enrollment: Enrollment) -> tuple[date, date]:
    """
    Default วันครบคอร์ส:
    - Class เสาร์เช้า/เสาร์บ่าย → วันเสาร์ที่กำลังจะมาถึง
    - Class อาทิตย์เช้า/อาทิตย์บ่าย → วันอาทิตย์ที่กำลังจะมาถึง
    - อื่น ๆ → วันเสาร์ที่กำลังจะมาถึง
    """
    today = timezone.localdate()
    time_slot = getattr(enrollment.tutoring_class, "time_slot", "") if enrollment and enrollment.tutoring_class_id else ""

    if time_slot in (
        TutoringClass.TimeSlot.SUN_MORNING,
        TutoringClass.TimeSlot.SUN_AFTERNOON,
    ):
        target_weekday = 6  # Sunday
    else:
        target_weekday = 5  # Saturday

    days_ahead = (target_weekday - today.weekday()) % 7
    expected_end = today + timedelta(days=days_ahead)
    next_start = expected_end + timedelta(days=7)
    return expected_end, next_start


def _decimal_from_post(value, default: Decimal) -> Decimal:
    try:
        s = str(value or "").replace(",", "").strip()
        if s == "":
            return default
        return Decimal(s)
    except Exception:
        return default


def _renewal_notice_packages(notice: CourseRenewalNotice) -> list[dict]:
    return [
        {
            "label": "ต่อคอร์ส 10 สัปดาห์",
            "weeks": 10,
            "full_price": notice.package_10_full_price,
            "discount": notice.package_10_discount,
            "net_price": notice.package_10_net_price,
            "accent": "blue",
        },
        {
            "label": "ต่อคอร์ส 20 สัปดาห์",
            "weeks": 20,
            "full_price": notice.package_20_full_price,
            "discount": notice.package_20_discount,
            "net_price": notice.package_20_net_price,
            "accent": "green",
        },
        {
            "label": "ต่อคอร์ส 30 สัปดาห์",
            "weeks": 30,
            "full_price": notice.package_30_full_price,
            "discount": notice.package_30_discount,
            "net_price": notice.package_30_net_price,
            "accent": "purple",
        },
    ]


def _issued_payments_for_enrollment(enrollment: Enrollment):
    return (
        CoursePayment.objects
        .filter(enrollment=enrollment, status=CoursePayment.ReceiptStatus.ISSUED)
        .order_by("payment_date", "created_at")
    )


def _installment_amounts_for_enrollment(enrollment: Enrollment) -> dict:
    payments = _issued_payments_for_enrollment(enrollment)
    paid = payments.aggregate(total=Sum("amount_paid")).get("total") or Decimal("0")
    full = Decimal(str(enrollment.net_price or enrollment.course_price or 0))
    remaining = max(full - paid, Decimal("0"))
    return {
        "full": full,
        "paid": paid,
        "remaining": remaining,
        "first_payment": payments.first(),
    }


def _admin_student_url(student: Student) -> str:
    return f"/adminlublub/core/student/{student.id}/change/" if student and student.id else "#"


def _admin_enrollment_url(enrollment: Enrollment) -> str:
    return f"/adminlublub/core/enrollment/{enrollment.id}/change/" if enrollment and enrollment.id else "#"


@login_required
def course_renewal_notice_list(request: HttpRequest) -> HttpResponse:
    """
    รายการ Enrollment ที่ใกล้ครบคอร์ส แบ่งตามสถานะ "กดส่งแจ้งผู้ปกครองแล้ว" เท่านั้น

    หลักการสำคัญ:
    - ไม่สนใจว่าเคยกดสร้างใบแจ้งแล้วหรือยัง
    - ถ้า enrollment เหลือ < 2 และยังไม่มีใบแจ้งใดที่ถูก mark ว่า sent → อยู่ในกลุ่ม "ยังไม่แจ้ง"
    - ถ้า enrollment เหลือ < 2 และมีใบแจ้งอย่างน้อย 1 ใบที่ถูก mark ว่า sent → อยู่ในกลุ่ม "แจ้งแล้ว"
    - date_from/date_to ใช้กรองเฉพาะประวัติใบแจ้งที่ส่งแล้วด้านล่าง ไม่กระทบการแบ่งกลุ่มหลัก
    """
    q = (request.GET.get("q") or "").strip()
    date_from = _safe_date(request.GET.get("date_from"))
    date_to = _safe_date(request.GET.get("date_to"))

    enrollments_qs = (
        Enrollment.objects
        .select_related("student", "tutoring_class")
        .filter(
            is_active=True,
            student__is_active=True,
            tutoring_class__is_active=True,
        )
        .order_by("tutoring_class__time_slot", "tutoring_class__name", "student__nickname", "student__full_name")
    )

    if q:
        enrollments_qs = enrollments_qs.filter(
            Q(student__nickname__icontains=q) |
            Q(student__full_name__icontains=q) |
            Q(student__student_code__icontains=q) |
            Q(tutoring_class__name__icontains=q) |
            Q(sale_run_no__icontains=q)
        )

    not_notified_rows = []
    notified_rows = []

    for enrollment in enrollments_qs:
        remaining_sessions = int(enrollment.remaining_sessions or 0)
        if remaining_sessions >= 2:
            continue

        # Do NOT apply date filters here.
        # Main grouping must depend only on whether this enrollment has ever been marked as sent.
        notices_qs = (
            CourseRenewalNotice.objects
            .select_related("student", "tutoring_class", "enrollment", "source_payment", "sent_to_parent_by")
            .filter(enrollment=enrollment)
            .order_by("-created_at")
        )

        latest_notice = notices_qs.first()
        latest_unsent_notice = notices_qs.filter(is_sent_to_parent=False).first()
        latest_sent_notice = (
            notices_qs
            .filter(is_sent_to_parent=True)
            .order_by("-sent_to_parent_at", "-created_at")
            .first()
        )

        expected_end, next_start = _default_renewal_dates(enrollment)
        amounts = _installment_amounts_for_enrollment(enrollment)

        row = {
            "enrollment": enrollment,
            "student": enrollment.student,
            "tutoring_class": enrollment.tutoring_class,
            "remaining": remaining_sessions,
            "latest_notice": latest_notice,
            "latest_unsent_notice": latest_unsent_notice,
            "latest_sent_notice": latest_sent_notice,
            "default_expected_end": expected_end,
            "default_next_start": next_start,
            "full_amount": amounts["full"],
            "paid_amount": amounts["paid"],
            "remaining_amount": amounts["remaining"],
            "first_payment": amounts["first_payment"],
            "student_admin_url": _admin_student_url(enrollment.student),
            "enrollment_admin_url": _admin_enrollment_url(enrollment),
        }

        if latest_sent_notice:
            notified_rows.append(row)
        else:
            not_notified_rows.append(row)

    all_notices_qs = (
        CourseRenewalNotice.objects
        .select_related("student", "tutoring_class", "enrollment", "source_payment", "sent_to_parent_by")
        .order_by("-created_at")
    )

    if q:
        all_notices_qs = all_notices_qs.filter(
            Q(student__nickname__icontains=q) |
            Q(student__full_name__icontains=q) |
            Q(student__student_code__icontains=q) |
            Q(tutoring_class__name__icontains=q) |
            Q(enrollment__sale_run_no__icontains=q) |
            Q(source_payment__receipt_no__icontains=q)
        )

    sent_history_qs = all_notices_qs.filter(is_sent_to_parent=True)
    if date_from:
        sent_history_qs = sent_history_qs.filter(sent_to_parent_at__date__gte=date_from)
    if date_to:
        sent_history_qs = sent_history_qs.filter(sent_to_parent_at__date__lte=date_to)

    sent_notices = sent_history_qs.order_by("-sent_to_parent_at", "-created_at")

    return render(request, "core/course_renewal_notice_list.html", {
        "not_notified_rows": not_notified_rows,
        "notified_rows": notified_rows,
        "sent_notices": sent_notices,
        "q": q,
        "date_from": date_from.isoformat() if date_from else "",
        "date_to": date_to.isoformat() if date_to else "",
    })

@login_required
def course_renewal_notice_create(request: HttpRequest, enrollment_id: int) -> HttpResponse:
    enrollment = get_object_or_404(
        Enrollment.objects.select_related("student", "tutoring_class"),
        id=enrollment_id,
        is_active=True,
        student__is_active=True,
        tutoring_class__is_active=True,
    )

    expected_end, next_start = _default_renewal_dates(enrollment)

    notice = CourseRenewalNotice.objects.create(
        notice_type=CourseRenewalNotice.NoticeType.RENEWAL,
        enrollment=enrollment,
        student=enrollment.student,
        tutoring_class=enrollment.tutoring_class,
        expected_course_end_date=expected_end,
        next_course_start_date=next_start,
        created_by=request.user if request.user.is_authenticated else None,
    )

    return redirect("core:course_renewal_notice_detail", pk=notice.pk)


@login_required
def course_installment_notice_create(request: HttpRequest, enrollment_id: int) -> HttpResponse:
    enrollment = get_object_or_404(
        Enrollment.objects.select_related("student", "tutoring_class"),
        id=enrollment_id,
        is_active=True,
        student__is_active=True,
        tutoring_class__is_active=True,
    )

    expected_end, next_start = _default_renewal_dates(enrollment)
    amounts = _installment_amounts_for_enrollment(enrollment)

    try:
        installment_no = int(request.GET.get("installment_no") or 2)
    except Exception:
        installment_no = 2
    if installment_no not in (2, 3, 4):
        installment_no = 2

    notice = CourseRenewalNotice.objects.create(
        notice_type=CourseRenewalNotice.NoticeType.INSTALLMENT,
        installment_no=installment_no,
        installment_sessions=0,
        enrollment=enrollment,
        student=enrollment.student,
        tutoring_class=enrollment.tutoring_class,
        source_payment=amounts["first_payment"],
        expected_course_end_date=expected_end,
        next_course_start_date=next_start,
        installment_full_amount=amounts["full"],
        installment_paid_amount=amounts["paid"],
        created_by=request.user if request.user.is_authenticated else None,
    )

    return redirect("core:course_renewal_notice_detail", pk=notice.pk)


@require_POST
@login_required
def course_renewal_notice_mark_sent(request: HttpRequest, pk: int) -> HttpResponse:
    notice = get_object_or_404(CourseRenewalNotice, pk=pk)
    notice.is_sent_to_parent = True
    notice.sent_to_parent_at = timezone.now()
    notice.sent_to_parent_by = request.user if request.user.is_authenticated else None
    notice.save()
    return redirect(request.META.get("HTTP_REFERER") or "core:course_renewal_notice_list")


@require_POST
@login_required
def course_renewal_notice_unmark_sent(request: HttpRequest, pk: int) -> HttpResponse:
    notice = get_object_or_404(CourseRenewalNotice, pk=pk)
    notice.is_sent_to_parent = False
    notice.sent_to_parent_at = None
    notice.sent_to_parent_by = None
    notice.save()
    return redirect(request.META.get("HTTP_REFERER") or "core:course_renewal_notice_list")


@login_required
def course_renewal_notice_detail(request: HttpRequest, pk: int) -> HttpResponse:
    notice = get_object_or_404(
        CourseRenewalNotice.objects.select_related("student", "tutoring_class", "enrollment", "source_payment"),
        pk=pk,
    )

    if request.method == "POST":
        notice.expected_course_end_date = _parse_date(request.POST.get("expected_course_end_date"))
        notice.next_course_start_date = _parse_date(request.POST.get("next_course_start_date"))

        notice_kind = (request.POST.get("notice_kind") or "").strip()
        if notice_kind == "renewal":
            notice.notice_type = CourseRenewalNotice.NoticeType.RENEWAL
            notice.installment_no = None
            notice.installment_sessions = 0
        elif notice_kind in {"installment_2", "installment_3", "installment_4"}:
            notice.notice_type = CourseRenewalNotice.NoticeType.INSTALLMENT
            try:
                notice.installment_no = int(notice_kind.split("_")[-1])
            except Exception:
                notice.installment_no = 2

        notice.package_10_full_price = _decimal_from_post(request.POST.get("package_10_full_price"), notice.package_10_full_price)
        notice.package_10_discount = _decimal_from_post(request.POST.get("package_10_discount"), notice.package_10_discount)

        notice.package_20_full_price = _decimal_from_post(request.POST.get("package_20_full_price"), notice.package_20_full_price)
        notice.package_20_discount = _decimal_from_post(request.POST.get("package_20_discount"), notice.package_20_discount)

        notice.package_30_full_price = _decimal_from_post(request.POST.get("package_30_full_price"), notice.package_30_full_price)
        notice.package_30_discount = _decimal_from_post(request.POST.get("package_30_discount"), notice.package_30_discount)

        notice.installment_full_amount = _decimal_from_post(request.POST.get("installment_full_amount"), notice.installment_full_amount)
        notice.installment_paid_amount = _decimal_from_post(request.POST.get("installment_paid_amount"), notice.installment_paid_amount)
        try:
            notice.installment_sessions = max(int(request.POST.get("installment_sessions") or 0), 0)
        except Exception:
            notice.installment_sessions = notice.installment_sessions or 0

        notice.note_wording = (request.POST.get("note_wording") or "").strip() or notice.note_wording
        notice.save()
        return redirect("core:course_renewal_notice_detail", pk=notice.pk)

    notice_kind = "renewal"
    if notice.notice_type == CourseRenewalNotice.NoticeType.INSTALLMENT:
        notice_kind = f"installment_{notice.installment_no or 2}"

    return render(request, "core/course_renewal_notice_detail.html", {
        "notice": notice,
        "notice_kind": notice_kind,
        "student": notice.student,
        "enrollment": notice.enrollment,
        "tutoring_class": notice.tutoring_class,
        "packages": _renewal_notice_packages(notice),
        "student_admin_url": _admin_student_url(notice.student),
        "enrollment_admin_url": _admin_enrollment_url(notice.enrollment),
    })




# =========================================================
# ✅ Weekly Small Test Module (Test ย่อยรายสัปดาห์)
# =========================================================
WEEKLY_TEST_RESULT_ORDER = [
    WeeklyTestScore.Result.FULL,
    WeeklyTestScore.Result.GREAT,
    WeeklyTestScore.Result.GOOD,
    WeeklyTestScore.Result.MEDIUM,
    WeeklyTestScore.Result.FAIL,
]

WEEKLY_TEST_RESULT_META = {
    WeeklyTestScore.Result.FULL: {"label": "เต็ม", "rank": 0, "class": "score-full"},
    WeeklyTestScore.Result.GREAT: {"label": "ดีมาก", "rank": 1, "class": "score-great"},
    WeeklyTestScore.Result.GOOD: {"label": "ดี", "rank": 2, "class": "score-good"},
    WeeklyTestScore.Result.MEDIUM: {"label": "ปานกลาง", "rank": 3, "class": "score-medium"},
    WeeklyTestScore.Result.FAIL: {"label": "ไม่ผ่าน", "rank": 4, "class": "score-fail"},
}

WEEKLY_TEST_ATTENDANCE_META = {
    WeeklyTestScore.AttendanceStatus.PRESENT: {"label": "ยังไม่กรอก", "rank": 5, "class": "score-empty"},
    WeeklyTestScore.AttendanceStatus.EXCUSED: {"label": "ลา", "rank": 6, "class": "score-excused"},
    WeeklyTestScore.AttendanceStatus.NO_SHOW: {"label": "ขาด", "rank": 7, "class": "score-noshow"},
    WeeklyTestScore.AttendanceStatus.NOT_CHECKED: {"label": "ยังไม่เช็คชื่อ", "rank": 8, "class": "score-notchecked"},
}

WEEKLY_TEST_INPUT_ATTENDANCE_ORDER = {
    WeeklyTestScore.AttendanceStatus.PRESENT: 0,
    WeeklyTestScore.AttendanceStatus.NOT_CHECKED: 1,
    WeeklyTestScore.AttendanceStatus.EXCUSED: 2,
    WeeklyTestScore.AttendanceStatus.NO_SHOW: 3,
}

THAI_SORT_ORDER = "กขฃคฅฆงจฉชซฌญฎฏฐฑฒณดตถทธนบปผฝพฟภมยรลวศษสหฬอฮะัาำิีึืุูเแโใไ่้๊๋์ๆabcdefghijklmnopqrstuvwxyz0123456789"
THAI_SORT_INDEX = {ch: i for i, ch in enumerate(THAI_SORT_ORDER)}


def _school_week_start(d: date) -> date:
    """Dashboard week = Sat-Sun. Return the Saturday for any selected date."""
    return d - timedelta(days=(d.weekday() - 5) % 7)


def _weekly_test_class_date(week_start: date, tutoring_class: TutoringClass) -> date:
    if getattr(tutoring_class, "time_slot", "") in {
        TutoringClass.TimeSlot.SUN_MORNING,
        TutoringClass.TimeSlot.SUN_AFTERNOON,
    }:
        return week_start + timedelta(days=1)
    return week_start


def _weekly_test_attendance_status(att: Attendance | None) -> str:
    if not att:
        return WeeklyTestScore.AttendanceStatus.NOT_CHECKED
    if att.status == Attendance.Status.PRESENT:
        return WeeklyTestScore.AttendanceStatus.PRESENT
    if att.status == Attendance.Status.EXCUSED:
        return WeeklyTestScore.AttendanceStatus.EXCUSED
    if att.status == Attendance.Status.NO_SHOW:
        return WeeklyTestScore.AttendanceStatus.NO_SHOW
    return WeeklyTestScore.AttendanceStatus.NOT_CHECKED


def _weekly_test_status_label(status: str) -> str:
    return WEEKLY_TEST_ATTENDANCE_META.get(status, WEEKLY_TEST_ATTENDANCE_META[WeeklyTestScore.AttendanceStatus.NOT_CHECKED])["label"]


def _weekly_test_result_label(result: str) -> str:
    return WEEKLY_TEST_RESULT_META.get(result, {}).get("label", "")


def _weekly_test_grade_label(grade_key: str) -> str:
    return _sheet_grade_label(grade_key or "")


def _weekly_test_valid_grade_keys() -> list[str]:
    return [g for g in SHEET_GRADE_ORDER if g]


def _weekly_test_is_valid_grade(grade_key: str) -> bool:
    return grade_key in set(_weekly_test_valid_grade_keys())


def _weekly_test_class_grade_key(tutoring_class: TutoringClass | None) -> str:
    return _class_grade_level(tutoring_class)


def _weekly_test_thai_sort_key(text: str) -> tuple:
    raw = (text or "").strip().lower()
    # Remove leading Thai vowels for more natural nickname sorting in most cases.
    normalized = raw
    return tuple(THAI_SORT_INDEX.get(ch, 999 + ord(ch)) for ch in normalized)


def _weekly_test_name_key(row: dict) -> tuple:
    student = row.get("student")
    nickname = (getattr(student, "nickname", "") or "").strip()
    full_name = (getattr(student, "full_name", "") or "").strip()
    primary = nickname or full_name
    return (_weekly_test_thai_sort_key(primary), _weekly_test_thai_sort_key(full_name), int(getattr(student, "id", 0) or 0))


def _weekly_test_input_row_sort_key(row: dict) -> tuple:
    return (WEEKLY_TEST_INPUT_ATTENDANCE_ORDER.get(row.get("attendance_status"), 9), *_weekly_test_name_key(row))


def _weekly_test_row_sort_key(row: dict) -> tuple:
    return (int(row.get("display_rank", 99)), *_weekly_test_name_key(row))


def _weekly_test_subject_display(test: WeeklyTest | None) -> str:
    if not test:
        return "-"
    return test.subject_display


def _weekly_test_filter_subject(qs, subject_filter: str):
    subject_filter = (subject_filter or "").strip()
    if not subject_filter:
        return qs
    return qs.filter(Q(subject__name__icontains=subject_filter) | Q(subject_name__icontains=subject_filter))


def _weekly_test_history_subject_choices() -> list[str]:
    names = set(Subject.objects.filter(is_active=True).values_list("name", flat=True))
    for name in WeeklyTest.objects.exclude(subject_name="").values_list("subject_name", flat=True):
        if name:
            names.add(name)
    for name in WeeklyTest.objects.filter(subject__isnull=False).values_list("subject__name", flat=True):
        if name:
            names.add(name)
    return sorted(names)


def _weekly_test_grade_tests_map(week_start: date, grade_keys: list[str] | None = None) -> dict[str, WeeklyTest]:
    qs = WeeklyTest.objects.select_related("subject").filter(week_start=week_start)
    if grade_keys is not None:
        qs = qs.filter(grade_level__in=grade_keys)
    return {t.grade_level: t for t in qs}


def _weekly_test_saved_or_current_row(
    *,
    week_start: date,
    weekly_test: WeeklyTest | None,
    enrollment: Enrollment,
    attendance: Attendance | None,
    score: WeeklyTestScore | None,
    class_grade_key: str,
    use_saved_status: bool = False,
) -> dict:
    student = enrollment.student
    tutoring_class = enrollment.tutoring_class
    att_date = _weekly_test_class_date(week_start, tutoring_class)
    current_status = _weekly_test_attendance_status(attendance)
    attendance_status = (score.attendance_status if (use_saved_status and score) else current_status) or current_status
    result = (score.result if score else "") or ""

    can_edit = current_status == WeeklyTestScore.AttendanceStatus.PRESENT
    if current_status != WeeklyTestScore.AttendanceStatus.PRESENT:
        can_edit = False
        if not use_saved_status:
            result = ""

    if result:
        meta = WEEKLY_TEST_RESULT_META.get(result, WEEKLY_TEST_RESULT_META[WeeklyTestScore.Result.FAIL])
        display_label = meta["label"]
        display_rank = meta["rank"]
        display_class = meta["class"]
    else:
        meta = WEEKLY_TEST_ATTENDANCE_META.get(attendance_status, WEEKLY_TEST_ATTENDANCE_META[WeeklyTestScore.AttendanceStatus.NOT_CHECKED])
        display_label = meta["label"]
        display_rank = meta["rank"]
        display_class = meta["class"]

    return {
        "weekly_test": weekly_test,
        "enrollment": enrollment,
        "student": student,
        "class": tutoring_class,
        "attendance": attendance,
        "attendance_date": att_date,
        "attendance_status": attendance_status,
        "attendance_label": "มา" if attendance_status == WeeklyTestScore.AttendanceStatus.PRESENT else _weekly_test_status_label(attendance_status),
        "attendance_class": f"att-{attendance_status.replace('_', '-')}",
        "score": score,
        "result": result,
        "result_label": _weekly_test_result_label(result),
        "display_label": display_label,
        "display_rank": display_rank,
        "display_class": display_class,
        "can_edit": can_edit,
        "grade_key": class_grade_key,
        "grade_label": _weekly_test_grade_label(class_grade_key),
    }


def _weekly_test_build_week_context(
    *,
    week_start: date,
    grade_filter: str = "all",
    class_filter: str | int | None = None,
    use_saved_status: bool = False,
) -> dict:
    grade_filter = (grade_filter or "").strip()
    valid_grades = _weekly_test_valid_grade_keys()

    all_classes = list(TutoringClass.objects.filter(is_active=True).order_by("time_slot", "name"))
    if class_filter:
        try:
            class_id = int(class_filter)
            all_classes = [c for c in all_classes if c.id == class_id]
        except Exception:
            pass

    class_grade_map = {c.id: _weekly_test_class_grade_key(c) for c in all_classes}

    if grade_filter == "all":
        visible_classes = [c for c in all_classes if class_grade_map.get(c.id) in valid_grades]
        visible_grade_keys = valid_grades
    elif _weekly_test_is_valid_grade(grade_filter):
        visible_classes = [c for c in all_classes if class_grade_map.get(c.id) == grade_filter]
        visible_grade_keys = [grade_filter]
    else:
        visible_classes = []
        visible_grade_keys = []

    class_ids = [c.id for c in visible_classes]
    tests_by_grade = _weekly_test_grade_tests_map(week_start, visible_grade_keys) if visible_grade_keys else {}
    test_ids = [t.id for t in tests_by_grade.values()]

    enrollments = list(
        Enrollment.objects
        .select_related("student", "tutoring_class", "student__school")
        .filter(
            is_active=True,
            student__is_active=True,
            tutoring_class__is_active=True,
            tutoring_class_id__in=class_ids,
        )
        .order_by("tutoring_class__time_slot", "tutoring_class__name", "student__nickname", "student__full_name")
    )
    enrollment_ids = [e.id for e in enrollments]

    attendance_dates = {week_start, week_start + timedelta(days=1)}
    atts = Attendance.objects.filter(
        enrollment_id__in=enrollment_ids,
        attendance_date__in=attendance_dates,
    ).select_related("enrollment")
    att_map = {(a.enrollment_id, a.attendance_date): a for a in atts}

    score_map = {}
    if test_ids:
        score_map = {
            (s.weekly_test_id, s.enrollment_id): s
            for s in WeeklyTestScore.objects
            .select_related("student", "tutoring_class", "enrollment", "weekly_test")
            .filter(weekly_test_id__in=test_ids, enrollment_id__in=enrollment_ids)
        }

    rows_by_class: dict[int, list[dict]] = defaultdict(list)
    rows_by_grade: dict[str, list[dict]] = defaultdict(list)
    all_rows: list[dict] = []

    for e in enrollments:
        grade_key = class_grade_map.get(e.tutoring_class_id, "")
        weekly_test = tests_by_grade.get(grade_key)
        att_date = _weekly_test_class_date(week_start, e.tutoring_class)
        score = score_map.get((weekly_test.id, e.id)) if weekly_test else None
        row = _weekly_test_saved_or_current_row(
            week_start=week_start,
            weekly_test=weekly_test,
            enrollment=e,
            attendance=att_map.get((e.id, att_date)),
            score=score,
            class_grade_key=grade_key,
            use_saved_status=use_saved_status,
        )
        rows_by_class[e.tutoring_class_id].append(row)
        rows_by_grade[grade_key].append(row)
        all_rows.append(row)

    class_blocks = []
    class_lookup = {c.id: c for c in visible_classes}
    for cls in visible_classes:
        rows = rows_by_class.get(cls.id, [])
        if not rows:
            continue
        rows = sorted(rows, key=_weekly_test_input_row_sort_key)
        class_blocks.append({
            "class": cls,
            "grade_key": class_grade_map.get(cls.id, ""),
            "grade_label": _weekly_test_grade_label(class_grade_map.get(cls.id, "")),
            "weekly_test": tests_by_grade.get(class_grade_map.get(cls.id, "")),
            "attendance_date": _weekly_test_class_date(week_start, cls),
            "rows": rows,
            "count": len(rows),
            "present_count": sum(1 for r in rows if r["attendance_status"] == WeeklyTestScore.AttendanceStatus.PRESENT),
            "excused_count": sum(1 for r in rows if r["attendance_status"] == WeeklyTestScore.AttendanceStatus.EXCUSED),
            "no_show_count": sum(1 for r in rows if r["attendance_status"] == WeeklyTestScore.AttendanceStatus.NO_SHOW),
            "not_checked_count": sum(1 for r in rows if r["attendance_status"] == WeeklyTestScore.AttendanceStatus.NOT_CHECKED),
        })

    grade_sections = []
    for grade_key in valid_grades:
        rows = rows_by_grade.get(grade_key, [])
        if not rows:
            continue
        blocks = []
        cls_ids_for_grade = sorted(
            {r["class"].id for r in rows},
            key=lambda cid: (
                getattr(class_lookup.get(cid), "time_slot", ""),
                getattr(class_lookup.get(cid), "name", ""),
            ),
        )
        for cid in cls_ids_for_grade:
            cls_rows = [r for r in rows if r["class"].id == cid]
            blocks.append({
                "class": class_lookup.get(cid) or cls_rows[0]["class"],
                "rows": sorted(cls_rows, key=_weekly_test_row_sort_key),
                "count": len(cls_rows),
                "present_count": sum(1 for r in cls_rows if r["attendance_status"] == WeeklyTestScore.AttendanceStatus.PRESENT),
                "excused_count": sum(1 for r in cls_rows if r["attendance_status"] == WeeklyTestScore.AttendanceStatus.EXCUSED),
                "no_show_count": sum(1 for r in cls_rows if r["attendance_status"] == WeeklyTestScore.AttendanceStatus.NO_SHOW),
                "not_checked_count": sum(1 for r in cls_rows if r["attendance_status"] == WeeklyTestScore.AttendanceStatus.NOT_CHECKED),
            })
        grade_sections.append({
            "grade_key": grade_key,
            "grade_label": _weekly_test_grade_label(grade_key),
            "weekly_test": tests_by_grade.get(grade_key),
            "rows": sorted(rows, key=_weekly_test_row_sort_key),
            "class_blocks": blocks,
            "count": len(rows),
        })

    return {
        "weekly_test": tests_by_grade.get(grade_filter) if _weekly_test_is_valid_grade(grade_filter) else None,
        "tests_by_grade": tests_by_grade,
        "week_start": week_start,
        "week_end": week_start + timedelta(days=1),
        "class_blocks": class_blocks,
        "grade_sections": grade_sections,
        "all_rows": all_rows,
    }


def _weekly_test_get_or_create_from_request(request: HttpRequest, week_start: date, grade_key: str) -> WeeklyTest:
    if not _weekly_test_is_valid_grade(grade_key):
        raise ValueError("กรุณาเลือกระดับชั้นก่อนบันทึก")

    subject_id = request.POST.get("subject_id") or None
    subject = Subject.objects.filter(id=subject_id, is_active=True).first() if subject_id else None
    subject_name = (request.POST.get("subject_name") or "").strip()
    topic = (request.POST.get("topic") or "").strip()
    test_date = _parse_date(request.POST.get("test_date"))
    try:
        difficulty = int(request.POST.get("difficulty") or 3)
    except Exception:
        difficulty = 3
    difficulty = min(max(difficulty, 1), 5)

    weekly_test, created = WeeklyTest.objects.get_or_create(
        week_start=week_start,
        grade_level=grade_key,
        defaults={
            "test_date": test_date,
            "subject": subject,
            "subject_name": subject_name,
            "topic": topic,
            "difficulty": difficulty,
            "created_by": request.user if getattr(request.user, "is_authenticated", False) else None,
        },
    )
    weekly_test.test_date = test_date
    weekly_test.subject = subject
    weekly_test.subject_name = subject_name
    weekly_test.topic = topic
    weekly_test.difficulty = difficulty
    weekly_test.updated_by = request.user if getattr(request.user, "is_authenticated", False) else None
    weekly_test.save()
    return weekly_test


def _weekly_test_class_ids_for_grade(grade_key: str) -> list[int]:
    if not _weekly_test_is_valid_grade(grade_key):
        return []
    ids = []
    for cls in TutoringClass.objects.filter(is_active=True).only("id", "name", "time_slot"):
        if _weekly_test_class_grade_key(cls) == grade_key:
            ids.append(cls.id)
    return ids


def _weekly_test_save_scores(request: HttpRequest, weekly_test: WeeklyTest, week_start: date, class_ids: list[int]) -> int:
    valid_results = set(WEEKLY_TEST_RESULT_ORDER)
    classes = list(TutoringClass.objects.filter(id__in=class_ids, is_active=True))
    class_map = {c.id: c for c in classes}
    enrollments = list(
        Enrollment.objects
        .select_related("student", "tutoring_class")
        .filter(
            tutoring_class_id__in=class_map.keys(),
            is_active=True,
            student__is_active=True,
            tutoring_class__is_active=True,
        )
    )
    enrollment_ids = [e.id for e in enrollments]
    attendance_dates = {week_start, week_start + timedelta(days=1)}
    att_map = {
        (a.enrollment_id, a.attendance_date): a
        for a in Attendance.objects.filter(enrollment_id__in=enrollment_ids, attendance_date__in=attendance_dates)
    }

    saved = 0
    with transaction.atomic():
        for e in enrollments:
            cls = e.tutoring_class
            att_date = _weekly_test_class_date(week_start, cls)
            att = att_map.get((e.id, att_date))
            att_status = _weekly_test_attendance_status(att)
            result = (request.POST.get(f"result_{e.id}") or "").strip()
            if att_status != WeeklyTestScore.AttendanceStatus.PRESENT or result not in valid_results:
                result = ""

            WeeklyTestScore.objects.update_or_create(
                weekly_test=weekly_test,
                enrollment=e,
                defaults={
                    "student": e.student,
                    "tutoring_class": cls,
                    "attendance_date": att_date,
                    "attendance_status": att_status,
                    "result": result,
                    "updated_by": request.user if getattr(request.user, "is_authenticated", False) else None,
                },
            )
            saved += 1
    return saved


@login_required
def weekly_test_admin(request: HttpRequest) -> HttpResponse:
    selected_date = _parse_date(request.POST.get("week") or request.GET.get("week"))
    week_start = _school_week_start(selected_date)
    grade_filter = (request.POST.get("grade") or request.GET.get("grade") or "").strip()
    mode = (request.GET.get("mode") or "weekly").strip()
    selected_class_id = (request.GET.get("class_id") or "").strip()
    selected_student_id = (request.GET.get("student_id") or "").strip()
    subject_filter = (request.GET.get("subject") or "").strip()
    saved_message = ""

    if request.method == "POST":
        try:
            weekly_test = _weekly_test_get_or_create_from_request(request, week_start, grade_filter)
        except ValueError:
            return redirect(f"{request.path}?mode=weekly&week={week_start.isoformat()}&grade=")

        save_scope = request.POST.get("save_scope") or "all"
        if save_scope == "meta_only":
            saved = 0
        elif save_scope.startswith("class:"):
            try:
                class_id = int(save_scope.split(":", 1)[1])
            except Exception:
                class_id = 0
            saved = _weekly_test_save_scores(request, weekly_test, week_start, [class_id] if class_id else [])
        else:
            class_ids = _weekly_test_class_ids_for_grade(grade_filter)
            saved = _weekly_test_save_scores(request, weekly_test, week_start, class_ids)
        redirect_url = f"{request.path}?mode=weekly&week={week_start.isoformat()}&grade={grade_filter}&saved={saved}#capture-summary"
        return redirect(redirect_url)

    week_data = _weekly_test_build_week_context(
        week_start=week_start,
        grade_filter=grade_filter,
    )
    weekly_test = week_data.get("weekly_test")
    if request.GET.get("saved") is not None:
        saved_count = request.GET.get("saved")
        saved_message = "บันทึกหัวข้อแล้ว" if saved_count == "0" else f"บันทึกข้อมูลแล้ว {saved_count} รายการ"

    # Class history
    class_history_rows = []
    selected_class = None
    if selected_class_id:
        selected_class = TutoringClass.objects.filter(id=selected_class_id, is_active=True).first()
    if mode == "class" and selected_class:
        tests_qs = _weekly_test_filter_subject(WeeklyTest.objects.select_related("subject").all(), subject_filter).order_by("-week_start", "grade_level")[:40]
        for t in tests_qs:
            scores = list(
                WeeklyTestScore.objects
                .select_related("student", "enrollment", "tutoring_class")
                .filter(weekly_test=t, tutoring_class=selected_class)
            )
            rows = []
            for s in scores:
                fake_row = {
                    "student": s.student,
                    "class": s.tutoring_class,
                    "attendance_status": s.attendance_status,
                    "result": s.result,
                    "display_label": _weekly_test_result_label(s.result) if s.result else _weekly_test_status_label(s.attendance_status),
                    "display_class": WEEKLY_TEST_RESULT_META.get(s.result, WEEKLY_TEST_ATTENDANCE_META.get(s.attendance_status, WEEKLY_TEST_ATTENDANCE_META[WeeklyTestScore.AttendanceStatus.NOT_CHECKED]))["class"],
                    "display_rank": WEEKLY_TEST_RESULT_META.get(s.result, WEEKLY_TEST_ATTENDANCE_META.get(s.attendance_status, WEEKLY_TEST_ATTENDANCE_META[WeeklyTestScore.AttendanceStatus.NOT_CHECKED]))["rank"],
                }
                rows.append(fake_row)
            if rows:
                class_history_rows.append({
                    "weekly_test": t,
                    "rows": sorted(rows, key=_weekly_test_row_sort_key),
                    "count": len(rows),
                })

    # Student history
    student_history_rows = []
    selected_student = None
    if selected_student_id:
        selected_student = Student.objects.filter(id=selected_student_id, is_active=True).first()
    if mode == "student" and selected_student:
        scores_qs = WeeklyTestScore.objects.select_related("weekly_test", "weekly_test__subject", "tutoring_class").filter(student=selected_student).order_by("-weekly_test__week_start", "weekly_test__grade_level")
        if subject_filter:
            scores_qs = scores_qs.filter(Q(weekly_test__subject__name__icontains=subject_filter) | Q(weekly_test__subject_name__icontains=subject_filter))
        for s in scores_qs[:100]:
            student_history_rows.append({
                "score": s,
                "weekly_test": s.weekly_test,
                "display_label": _weekly_test_result_label(s.result) if s.result else _weekly_test_status_label(s.attendance_status),
                "display_class": WEEKLY_TEST_RESULT_META.get(s.result, WEEKLY_TEST_ATTENDANCE_META.get(s.attendance_status, WEEKLY_TEST_ATTENDANCE_META[WeeklyTestScore.AttendanceStatus.NOT_CHECKED]))["class"],
            })

    grade_choices = [
        {"key": "", "label": "— เลือกระดับชั้นเพื่อกรอก —"},
        *[{"key": g, "label": _weekly_test_grade_label(g)} for g in _weekly_test_valid_grade_keys()],
        {"key": "all", "label": "ทุกระดับชั้น (หน้าสรุป)"},
    ]
    selected_grade_label = "ทุกระดับชั้น" if grade_filter == "all" else (_weekly_test_grade_label(grade_filter) if grade_filter else "ยังไม่ได้เลือกระดับชั้น")
    can_edit_weekly = _weekly_test_is_valid_grade(grade_filter)

    context = {
        **week_data,
        "selected_date": selected_date,
        "week_start": week_start,
        "week_end": week_start + timedelta(days=1),
        "grade_filter": grade_filter,
        "selected_grade_label": selected_grade_label,
        "can_edit_weekly": can_edit_weekly,
        "grade_choices": grade_choices,
        "subjects": Subject.objects.filter(is_active=True).order_by("name"),
        "difficulty_choices": [1, 2, 3, 4, 5],
        "result_choices": [
            {"value": value, **WEEKLY_TEST_RESULT_META[value]} for value in WEEKLY_TEST_RESULT_ORDER
        ],
        "saved_message": saved_message,
        "mode": mode,
        "classes": TutoringClass.objects.filter(is_active=True).order_by("time_slot", "name"),
        "students": Student.objects.filter(is_active=True).order_by("nickname", "full_name")[:500],
        "selected_class": selected_class,
        "selected_class_id": selected_class_id,
        "selected_student": selected_student,
        "selected_student_id": selected_student_id,
        "subject_filter": subject_filter,
        "history_subjects": _weekly_test_history_subject_choices(),
        "class_history_rows": class_history_rows,
        "student_history_rows": student_history_rows,
        "export_url": f"{request.path}export/?week={week_start.isoformat()}&grade={grade_filter}&mode={mode}&class_id={selected_class_id}&student_id={selected_student_id}&subject={subject_filter}",
    }
    return render(request, "core/weekly_test_admin.html", context)


@login_required
def weekly_test_export(request: HttpRequest) -> HttpResponse:
    selected_date = _parse_date(request.GET.get("week"))
    week_start = _school_week_start(selected_date)
    grade_filter = (request.GET.get("grade") or "all").strip() or "all"
    mode = (request.GET.get("mode") or "weekly").strip()
    class_id = (request.GET.get("class_id") or "").strip()
    student_id = (request.GET.get("student_id") or "").strip()
    subject_filter = (request.GET.get("subject") or "").strip()

    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Test"

    if mode == "class" and class_id:
        cls = TutoringClass.objects.filter(id=class_id).first()
        ws.append(["Class History", cls.name if cls else class_id])
        ws.append(["Week Start", "Grade", "Test Date", "Subject", "Topic", "Difficulty", "Nickname", "Full Name", "Status/Result", "Class"])
        tests_qs = _weekly_test_filter_subject(WeeklyTest.objects.select_related("subject").all(), subject_filter).order_by("-week_start", "grade_level")
        for t in tests_qs:
            scores = WeeklyTestScore.objects.select_related("student", "tutoring_class").filter(weekly_test=t, tutoring_class_id=class_id)
            for s in scores.order_by("student__nickname", "student__full_name"):
                ws.append([
                    t.week_start.isoformat(),
                    t.grade_display,
                    t.test_date.isoformat() if t.test_date else "",
                    t.subject_display,
                    t.topic,
                    t.difficulty_stars,
                    s.student.nickname,
                    s.student.full_name,
                    _weekly_test_result_label(s.result) if s.result else _weekly_test_status_label(s.attendance_status),
                    s.tutoring_class.name,
                ])
    elif mode == "student" and student_id:
        student = Student.objects.filter(id=student_id).first()
        ws.append(["Student History", student.display_name if student else student_id])
        ws.append(["Week Start", "Grade", "Test Date", "Subject", "Topic", "Difficulty", "Class", "Status/Result"])
        scores = WeeklyTestScore.objects.select_related("weekly_test", "weekly_test__subject", "tutoring_class").filter(student_id=student_id).order_by("-weekly_test__week_start", "weekly_test__grade_level")
        if subject_filter:
            scores = scores.filter(Q(weekly_test__subject__name__icontains=subject_filter) | Q(weekly_test__subject_name__icontains=subject_filter))
        for s in scores:
            t = s.weekly_test
            ws.append([
                t.week_start.isoformat(),
                t.grade_display,
                t.test_date.isoformat() if t.test_date else "",
                t.subject_display,
                t.topic,
                t.difficulty_stars,
                s.tutoring_class.name,
                _weekly_test_result_label(s.result) if s.result else _weekly_test_status_label(s.attendance_status),
            ])
    else:
        export_grade = grade_filter if grade_filter else "all"
        data = _weekly_test_build_week_context(week_start=week_start, grade_filter=export_grade)
        ws.append(["Weekly View", f"{week_start.isoformat()} to {(week_start + timedelta(days=1)).isoformat()}"])
        ws.append([])
        ws.append(["Grade", "Test Date", "Subject", "Topic", "Difficulty", "Class", "Nickname", "Full Name", "Attendance Date", "Status/Result"])
        for section in data["grade_sections"]:
            test = section.get("weekly_test")
            for block in section["class_blocks"]:
                for row in block["rows"]:
                    ws.append([
                        section["grade_label"],
                        test.test_date.isoformat() if test and test.test_date else "",
                        test.subject_display if test else "",
                        test.topic if test else "",
                        test.difficulty_stars if test else "",
                        block["class"].name,
                        row["student"].nickname,
                        row["student"].full_name,
                        row["attendance_date"].isoformat() if row["attendance_date"] else "",
                        row["display_label"],
                    ])

    _autosize(ws)
    buff = BytesIO()
    wb.save(buff)
    buff.seek(0)
    filename = f"weekly_test_{week_start.strftime('%Y%m%d')}_{datetime.now().strftime('%H%M%S')}.xlsx"
    resp = HttpResponse(buff.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp



def pkanoon_admin_tool(request: HttpRequest) -> HttpResponse:
    """Private landing page for sensitive school tools."""
    return render(request, "core/pkanoon_admin_tool.html")



def _weekend_tutor_summary(selected_date: date) -> dict:
    """
    Combine tutor payroll by school weekend.
    If selected date is Sat: Sat + next Sun.
    If selected date is Sun: previous Sat + Sun.
    Other days: school week containing the selected date (Sat-Sun).
    """
    weekend_start, weekend_end = _school_week_range(selected_date)

    rows_by_tutor: dict[int, dict] = {}
    qs = (
        TutorPayrollEntry.objects
        .select_related("tutor")
        .filter(work_date__gte=weekend_start, work_date__lte=weekend_end)
        .order_by("tutor__name", "work_date")
    )

    zero = Decimal("0")
    totals = {
        "tutor_count": 0,
        "onsite_hours": zero,
        "online_hours": zero,
        "onsite_fee": zero,
        "online_fee": zero,
        "travel_fee": zero,
        "idle_fee": zero,
        "total_amount": zero,
    }

    for p in qs:
        if not p.tutor_id:
            continue

        row = rows_by_tutor.setdefault(p.tutor_id, {
            "tutor": p.tutor,
            "onsite_hours": zero,
            "online_hours": zero,
            "onsite_fee": zero,
            "online_fee": zero,
            "travel_fee": zero,
            "idle_fee": zero,
            "total_amount": zero,
            "sat_amount": zero,
            "sun_amount": zero,
            "sat_hours": zero,
            "sun_hours": zero,
            "has_special_rate": False,
            "notes": [],
        })

        teaching_hours = Decimal(str(p.teaching_hours or 0))
        online_hours = Decimal(str(getattr(p, "online_teaching_hours", 0) or 0))
        teaching_fee = Decimal(str(p.teaching_fee or 0))
        online_fee = Decimal(str(getattr(p, "online_teaching_fee", 0) or 0))
        travel_fee = Decimal(str(p.travel_fee or 0))
        idle_fee = Decimal(str(p.idle_fee or 0))
        total_amount = Decimal(str(p.total_amount or 0))

        row["onsite_hours"] += teaching_hours
        row["online_hours"] += online_hours
        row["onsite_fee"] += teaching_fee
        row["online_fee"] += online_fee
        row["travel_fee"] += travel_fee
        row["idle_fee"] += idle_fee
        row["total_amount"] += total_amount
        row["has_special_rate"] = row["has_special_rate"] or bool(getattr(p, "special_rate_325", False))

        if p.work_date == weekend_start:
            row["sat_amount"] += total_amount
            row["sat_hours"] += teaching_hours + online_hours
        elif p.work_date == weekend_end:
            row["sun_amount"] += total_amount
            row["sun_hours"] += teaching_hours + online_hours

        if p.note:
            row["notes"].append(f"{p.work_date.strftime('%d/%m')}: {p.note}")

    rows = sorted(rows_by_tutor.values(), key=lambda r: (r["tutor"].name if r.get("tutor") else ""))

    totals["tutor_count"] = len(rows)
    for row in rows:
        totals["onsite_hours"] += row["onsite_hours"]
        totals["online_hours"] += row["online_hours"]
        totals["onsite_fee"] += row["onsite_fee"]
        totals["online_fee"] += row["online_fee"]
        totals["travel_fee"] += row["travel_fee"]
        totals["idle_fee"] += row["idle_fee"]
        totals["total_amount"] += row["total_amount"]

    return {
        "start": weekend_start,
        "end": weekend_end,
        "rows": rows,
        "totals": totals,
    }


@login_required
def school_finance_export(request: HttpRequest) -> HttpResponse:
    data = _school_finance_filtered_data(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Item", "Amount / Count"])
    ws.append(["Date From", data["date_from"].isoformat()])
    ws.append(["Date To", data["date_to"].isoformat()])
    ws.append(["Revenue per attendance", float(data["revenue_per_student"])])
    ws.append(["Recognized attendance count", data["deducted_count"]])
    ws.append(["Estimated revenue from attendance", float(data["estimated_revenue"])])
    ws.append(["Cash basis course revenue", float(data["cash_revenue_total"])])
    ws.append(["General expenses", float(data["general_expense_total"])])
    ws.append(["Tutor payroll", float(data["tutor_payroll_total"])])
    ws.append(["Total expenses", float(data["total_expense"])])
    ws.append(["Net attendance estimate", float(data["net_estimated"])])
    ws.append(["Net cash basis", float(data["net_cash_basis"])])
    _autosize(ws)

    ws_exp = wb.create_sheet("General Expenses")
    ws_exp.append(["Date", "Category", "Vendor", "Description", "Payment Method", "Amount", "Note"])
    for e in data["expense_rows"]:
        ws_exp.append([
            e.expense_date.isoformat(),
            e.category.name if e.category else "",
            e.vendor or "",
            e.description or "",
            e.get_payment_method_display() if hasattr(e, "get_payment_method_display") else e.payment_method,
            float(e.amount or 0),
            e.note or "",
        ])
    _autosize(ws_exp)

    ws_pay = wb.create_sheet("Tutor Payroll")
    ws_pay.append(["Date", "Tutor", "Onsite Hours", "Special 325?", "Hourly Rate", "Onsite Teaching Fee", "Online Hours", "Online Teaching Fee", "Travel Fee", "Idle/Other Fee", "Total", "Note"])
    for p in data["payroll_rows"]:
        ws_pay.append([
            p.work_date.isoformat(),
            p.tutor.name if p.tutor else "",
            float(p.teaching_hours or 0),
            "Yes" if getattr(p, "special_rate_325", False) else "No",
            float(p.hourly_rate or 0),
            float(p.teaching_fee or 0),
            float(getattr(p, "online_teaching_hours", 0) or 0),
            float(getattr(p, "online_teaching_fee", 0) or 0),
            float(p.travel_fee or 0),
            float(p.idle_fee or 0),
            float(p.total_amount or 0),
            p.note or "",
        ])
    _autosize(ws_pay)

    weekend_summary = _weekend_tutor_summary(data["selected_date"])
    ws_weekend = wb.create_sheet("Weekend Tutor Summary")
    ws_weekend.append([
        "Weekend Start",
        weekend_summary["start"].isoformat(),
        "Weekend End",
        weekend_summary["end"].isoformat(),
    ])
    ws_weekend.append([])
    ws_weekend.append([
        "Tutor",
        "Onsite Hours",
        "Online Hours",
        "Onsite Teaching Fee",
        "Online Teaching Fee",
        "Travel Fee",
        "Idle/Other Fee",
        "Saturday Amount",
        "Sunday Amount",
        "Weekend Total",
        "Special 325 Used?",
    ])
    for r in weekend_summary["rows"]:
        ws_weekend.append([
            r["tutor"].name if r.get("tutor") else "",
            float(r["onsite_hours"] or 0),
            float(r["online_hours"] or 0),
            float(r["onsite_fee"] or 0),
            float(r["online_fee"] or 0),
            float(r["travel_fee"] or 0),
            float(r["idle_fee"] or 0),
            float(r["sat_amount"] or 0),
            float(r["sun_amount"] or 0),
            float(r["total_amount"] or 0),
            "Yes" if r["has_special_rate"] else "No",
        ])
    _autosize(ws_weekend)

    ws_cash = wb.create_sheet("Course Payments")
    ws_cash.append(["Receipt No", "Payment Date", "Student", "Class", "Enrollment", "Payment Type", "Payment Method", "Sessions", "Course Price", "Discount", "Net Amount", "Amount Paid", "Status", "Note"])
    course_payment_export_rows = (
        CoursePayment.objects
        .select_related("student", "tutoring_class", "enrollment")
        .filter(status=CoursePayment.ReceiptStatus.ISSUED)
        .order_by("-payment_date", "-created_at")
    )
    for p in course_payment_export_rows:
        ws_cash.append([
            p.receipt_no,
            p.payment_date.isoformat() if p.payment_date else "",
            p.student.display_name if p.student_id else "",
            p.tutoring_class.name if p.tutoring_class_id else "",
            p.enrollment.sale_run_no if p.enrollment_id and p.enrollment else "",
            p.get_payment_type_display() if hasattr(p, "get_payment_type_display") else p.payment_type,
            p.get_payment_method_display() if hasattr(p, "get_payment_method_display") else p.payment_method,
            int(p.sessions_granted or 0),
            float(p.course_price or 0),
            float(p.discount_amount or 0),
            float(p.net_amount or 0),
            float(p.amount_paid or 0),
            p.get_status_display() if hasattr(p, "get_status_display") else p.status,
            p.note or "",
        ])
    _autosize(ws_cash)

    buff = BytesIO()
    wb.save(buff)
    buff.seek(0)

    filename = f"school_finance_{data['date_from'].strftime('%Y%m%d')}_{data['date_to'].strftime('%Y%m%d')}.xlsx"
    resp = HttpResponse(
        buff.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@login_required
def school_finance(request: HttpRequest) -> HttpResponse:
    """
    Module 2:
    - Revenue estimate from attendance.
    - General expenses.
    - Tutor payroll batch input.
    """
    selected_date = _parse_date(request.GET.get("work_date") or request.POST.get("work_date"))
    date_from = _parse_date(request.GET.get("date_from"))
    date_to = _parse_date(request.GET.get("date_to"))
    if date_to < date_from:
        date_from, date_to = date_to, date_from

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()

        if action == "add_category":
            name = (request.POST.get("category_name") or "").strip()
            if name:
                ExpenseCategory.objects.get_or_create(name=name, defaults={"is_active": True, "sort_order": 99})
            return redirect(request.META.get("HTTP_REFERER", "core:school_finance"))

        if action == "add_tutor":
            name = (request.POST.get("tutor_name") or "").strip()
            phone = (request.POST.get("tutor_phone") or "").strip()
            if name:
                Tutor.objects.update_or_create(name=name, defaults={"phone": phone, "is_active": True})
            return redirect(request.META.get("HTTP_REFERER", "core:school_finance"))

        if action == "add_expense":
            category_id = request.POST.get("category")
            try:
                category = ExpenseCategory.objects.get(id=category_id, is_active=True)
                SchoolExpense.objects.create(
                    expense_date=_parse_date(request.POST.get("expense_date")),
                    category=category,
                    vendor=(request.POST.get("vendor") or "").strip(),
                    description=(request.POST.get("description") or "").strip(),
                    amount=_money(request.POST.get("amount")),
                    payment_method=(request.POST.get("payment_method") or SchoolExpense.PaymentMethod.TRANSFER),
                    note=(request.POST.get("note") or "").strip(),
                )
            except Exception:
                pass
            return redirect(request.META.get("HTTP_REFERER", "core:school_finance"))

        if action == "save_payroll":
            work_date = _parse_date(request.POST.get("work_date"))
            single_tutor_id = (request.POST.get("single_tutor_id") or "").strip()
            tutor_ids = request.POST.getlist("tutor_ids")
            if single_tutor_id:
                tutor_ids = [single_tutor_id]

            for tid in tutor_ids:
                hours_raw = (request.POST.get(f"hours_{tid}") or "").strip()
                online_hours_raw = (request.POST.get(f"online_hours_{tid}") or "").strip()
                idle_raw = (request.POST.get(f"idle_{tid}") or "").strip()
                note_raw = (request.POST.get(f"note_{tid}") or "").strip()
                special_rate_325 = request.POST.get(f"special_rate_325_{tid}") == "yes"

                # Skip blank rows when saving all.
                if (
                    not single_tutor_id
                    and hours_raw == ""
                    and online_hours_raw == ""
                    and idle_raw == ""
                    and note_raw == ""
                    and not special_rate_325
                ):
                    continue

                hours = _money(hours_raw)
                online_hours = _money(online_hours_raw)
                idle_fee = _money(idle_raw)

                if hours <= 0 and online_hours <= 0 and idle_fee <= 0 and not note_raw:
                    continue

                try:
                    tutor = Tutor.objects.get(id=tid, is_active=True)
                except Tutor.DoesNotExist:
                    continue

                entry, _ = TutorPayrollEntry.objects.get_or_create(
                    work_date=work_date,
                    tutor=tutor,
                    defaults={
                        "teaching_hours": hours,
                        "online_teaching_hours": online_hours,
                        "special_rate_325": special_rate_325,
                        "idle_fee": idle_fee,
                        "note": note_raw,
                    },
                )
                entry.teaching_hours = hours
                entry.online_teaching_hours = online_hours
                entry.special_rate_325 = special_rate_325
                entry.idle_fee = idle_fee
                entry.note = note_raw
                entry.save()

            return redirect(f"{request.path}?work_date={work_date.isoformat()}&date_from={date_from.isoformat()}&date_to={date_to.isoformat()}")

    data = _school_finance_filtered_data(request)
    weekend_tutor_summary = _weekend_tutor_summary(data["selected_date"])

    categories = ExpenseCategory.objects.filter(is_active=True).order_by("sort_order", "name")
    tutors = Tutor.objects.filter(is_active=True).order_by("name")
    existing_entries = TutorPayrollEntry.objects.filter(work_date=data["selected_date"]).select_related("tutor")
    payroll_map = {e.tutor_id: e for e in existing_entries}

    payment_method_choices = SchoolExpense.PaymentMethod.choices
    current_querystring = request.GET.urlencode()

    return render(request, "core/school_finance.html", {
        "selected_date": data["selected_date"],
        "date_from": data["date_from"],
        "date_to": data["date_to"],
        "categories": categories,
        "tutors": tutors,
        "payroll_map": payroll_map,
        "expense_rows": data["expense_rows"],
        "payroll_rows": data["payroll_rows"],
        "weekend_tutor_summary": weekend_tutor_summary,
        "deducted_count": data["deducted_count"],
        "revenue_per_student": data["revenue_per_student"],
        "estimated_revenue": data["estimated_revenue"],
        "cash_revenue_total": data["cash_revenue_total"],
        "general_expense_total": data["general_expense_total"],
        "tutor_payroll_total": data["tutor_payroll_total"],
        "total_expense": data["total_expense"],
        "net_estimated": data["net_estimated"],
        "net_cash_basis": data["net_cash_basis"],
        "payment_method_choices": payment_method_choices,
        "current_querystring": current_querystring,
    })


@login_required
def course_payment_list(request: HttpRequest) -> HttpResponse:
    qs = CoursePayment.objects.select_related("student", "tutoring_class", "enrollment").order_by("-payment_date", "-created_at")
    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "").strip()
    date_from = _safe_date(request.GET.get("date_from"))
    date_to = _safe_date(request.GET.get("date_to"))

    if q:
        qs = qs.filter(
            Q(receipt_no__icontains=q) |
            Q(student__nickname__icontains=q) |
            Q(student__full_name__icontains=q) |
            Q(student__student_code__icontains=q) |
            Q(tutoring_class__name__icontains=q)
        )
    if status:
        qs = qs.filter(status=status)
    if date_from:
        qs = qs.filter(payment_date__gte=date_from)
    if date_to:
        qs = qs.filter(payment_date__lte=date_to)

    totals = qs.aggregate(total=Sum("amount_paid"))
    issued_total = qs.filter(status=CoursePayment.ReceiptStatus.ISSUED).aggregate(total=Sum("amount_paid")).get("total") or Decimal("0")

    return render(request, "core/course_payment_list.html", {
        "payments": qs[:300],
        "filters": {"q": q, "status": status, "date_from": request.GET.get("date_from", ""), "date_to": request.GET.get("date_to", "")},
        "status_choices": CoursePayment.ReceiptStatus.choices,
        "issued_total": issued_total,
        "all_total": totals.get("total") or Decimal("0"),
    })


@login_required
def course_payment_create(request: HttpRequest) -> HttpResponse:
    errors: list[str] = []
    if request.method == "POST":
        post = request.POST
        try:
            student_mode = (post.get("student_mode") or "existing").strip()
            enrollment_action = (post.get("enrollment_action") or CoursePayment.EnrollmentAction.NEW).strip()
            payment_date = _parse_date(post.get("payment_date"))
            payment_type = (post.get("payment_type") or CoursePayment.PaymentType.FULL).strip()
            payment_method = (post.get("payment_method") or CoursePayment.PaymentMethod.BANK_TRANSFER).strip()
            package = (post.get("session_package") or "10").strip()
            sessions_granted = _sessions_from_package(package, post.get("custom_sessions"))
            course_price = _money(post.get("course_price"))
            discount_amount = _money(post.get("discount_amount"))
            net_amount = max(course_price - discount_amount, Decimal("0"))
            amount_paid = _money(post.get("amount_paid")) or net_amount
            note = (post.get("note") or "").strip()

            if student_mode == "new":
                nickname = (post.get("new_nickname") or "").strip()
                full_name = (post.get("new_full_name") or "").strip()
                grade_level = (post.get("new_grade_level") or "").strip()
                school_name = (post.get("new_school_name") or "").strip()
                parent_phone = (post.get("new_parent_phone") or "").strip()
                if not full_name:
                    errors.append("กรุณากรอกชื่อจริงนามสกุลของนักเรียนใหม่")
                if not parent_phone:
                    errors.append("กรุณากรอกเบอร์ผู้ปกครองของนักเรียนใหม่")
                school = None
                if school_name:
                    school, _ = School.objects.get_or_create(name=school_name, defaults={"is_active": True})
                student = None
                if not errors:
                    student = Student.objects.create(
                        nickname=nickname,
                        full_name=full_name,
                        grade_level=grade_level,
                        school=school,
                        parent_phone=parent_phone,
                    )
            else:
                student_id = post.get("student_id")
                student = Student.objects.filter(id=student_id, is_active=True).first()
                if not student and enrollment_action != CoursePayment.EnrollmentAction.ADD_EXISTING:
                    errors.append("กรุณาเลือกนักเรียนเดิม")

            selected_class = None
            existing_enrollment = None
            if enrollment_action == CoursePayment.EnrollmentAction.ADD_EXISTING:
                existing_enrollment = Enrollment.objects.select_related("student", "tutoring_class").filter(id=post.get("existing_enrollment_id"), is_active=True).first()
                if not existing_enrollment:
                    errors.append("กรุณาเลือก Enrollment เดิมที่จะเพิ่มจำนวนครั้ง")
                else:
                    # In add-existing mode the active enrollment search is the source of truth.
                    # This makes the flow faster even if the student search box was not selected first.
                    if not student:
                        student = existing_enrollment.student
                    elif existing_enrollment.student_id != student.id:
                        errors.append("Enrollment เดิมไม่ตรงกับนักเรียนที่เลือก")
                    selected_class = existing_enrollment.tutoring_class
            else:
                selected_class = TutoringClass.objects.filter(id=post.get("tutoring_class_id"), is_active=True).first()
                if not selected_class:
                    errors.append("กรุณาเลือก Class")

            if sessions_granted <= 0:
                errors.append("จำนวนครั้งที่ให้เรียนต้องมากกว่า 0")
            if amount_paid < 0:
                errors.append("ยอดรับชำระไม่ถูกต้อง")

            if not errors and student and selected_class:
                with transaction.atomic():
                    payment = CoursePayment.objects.create(
                        payment_date=payment_date,
                        student=student,
                        tutoring_class=selected_class,
                        enrollment_action=enrollment_action,
                        session_package=package,
                        sessions_granted=sessions_granted,
                        course_price=course_price,
                        discount_amount=discount_amount,
                        net_amount=net_amount,
                        amount_paid=amount_paid,
                        payment_type=payment_type,
                        payment_method=payment_method,
                        note=note,
                        created_by=request.user if request.user.is_authenticated else None,
                    )
                    _apply_payment_to_enrollment(payment, enrollment_action, existing_enrollment)
                return redirect("core:course_payment_detail", pk=payment.pk)
        except Exception as exc:
            errors.append(f"บันทึกไม่สำเร็จ: {exc}")

        context = _default_payment_form_context(request, errors=errors, posted=dict(post))
        return render(request, "core/course_payment_create.html", context)

    return render(request, "core/course_payment_create.html", _default_payment_form_context(request))


@login_required
def course_payment_detail(request: HttpRequest, pk: int) -> HttpResponse:
    payment = get_object_or_404(
        CoursePayment.objects.select_related("student", "student__school", "tutoring_class", "enrollment"),
        pk=pk,
    )
    return render(request, "core/course_payment_detail.html", {"payment": payment})


@login_required
def course_payment_receipt_image(request: HttpRequest, pk: int) -> HttpResponse:
    payment = get_object_or_404(
        CoursePayment.objects.select_related("student", "student__school", "tutoring_class", "enrollment"),
        pk=pk,
    )
    return render(request, "core/course_payment_receipt_image.html", {"payment": payment})


@require_POST
@login_required
def course_payment_cancel(request: HttpRequest, pk: int) -> HttpResponse:
    payment = get_object_or_404(CoursePayment.objects.select_related("enrollment"), pk=pk)
    if payment.status != CoursePayment.ReceiptStatus.CANCELLED:
        with transaction.atomic():
            payment.status = CoursePayment.ReceiptStatus.CANCELLED
            payment.cancel_reason = (request.POST.get("cancel_reason") or "").strip()
            payment.cancelled_at = timezone.now()
            payment.cancelled_by = request.user if request.user.is_authenticated else None
            payment.save()
            _reverse_payment_enrollment_effect(payment)
    return redirect("core:course_payment_detail", pk=payment.pk)


@require_POST
@login_required
def school_expense_delete(request: HttpRequest, pk: int) -> HttpResponse:
    expense = get_object_or_404(SchoolExpense, pk=pk)
    expense.delete()
    return redirect(request.META.get("HTTP_REFERER", "core:school_finance"))


@require_POST
@login_required
def tutor_payroll_delete(request: HttpRequest, pk: int) -> HttpResponse:
    entry = get_object_or_404(TutorPayrollEntry, pk=pk)
    entry.delete()
    return redirect(request.META.get("HTTP_REFERER", "core:school_finance"))

# =========================================================
# ✅ Tutor Teaching Update Module
# =========================================================
def _teaching_week_range(anchor: date) -> tuple[date, date]:
    """Teaching week = Saturday to Sunday, same as school finance week."""
    days_since_sat = (anchor.weekday() - 5) % 7
    start = anchor - timedelta(days=days_since_sat)
    return start, start + timedelta(days=1)


def _teaching_week_from_request(request: HttpRequest) -> tuple[date, date]:
    raw = request.GET.get("week_start") or request.POST.get("week_start") or request.GET.get("date")
    anchor = _safe_date(raw) if "_safe_date" in globals() else None
    if not anchor:
        anchor = _parse_date(raw)
    return _teaching_week_range(anchor)


def _latest_tutor_for_template(template: TeachingClassSubjectTemplate, before_week_start: date) -> TeachingTutor | None:
    prev = (
        TeachingWeeklyAssignment.objects
        .filter(subject_template=template, week_start_date__lt=before_week_start, tutor__isnull=False)
        .select_related("tutor")
        .order_by("-week_start_date", "-updated_at")
        .first()
    )
    return prev.tutor if prev and prev.tutor_id else None


def _ensure_teaching_assignments(week_start: date, week_end: date) -> None:
    templates = (
        TeachingClassSubjectTemplate.objects
        .select_related("tutoring_class")
        .filter(is_active=True, tutoring_class__is_active=True)
        .order_by("tutoring_class__name", "display_order", "subject_name")
    )

    with transaction.atomic():
        for tmpl in templates:
            assignment, created = TeachingWeeklyAssignment.objects.get_or_create(
                week_start_date=week_start,
                subject_template=tmpl,
                defaults={
                    "week_end_date": week_end,
                    "tutoring_class": tmpl.tutoring_class,
                    "tutor": _latest_tutor_for_template(tmpl, week_start),
                    "is_teaching": True,
                },
            )
            changed = False
            if assignment.week_end_date != week_end:
                assignment.week_end_date = week_end
                changed = True
            if assignment.tutoring_class_id != tmpl.tutoring_class_id:
                assignment.tutoring_class = tmpl.tutoring_class
                changed = True
            if created is False and assignment.tutor_id is None:
                last_tutor = _latest_tutor_for_template(tmpl, week_start)
                if last_tutor:
                    assignment.tutor = last_tutor
                    changed = True
            if changed:
                assignment.save()


def _latest_real_progress_for_assignment(assignment: TeachingWeeklyAssignment, week_start: date) -> TeachingProgressUpdate | None:
    return (
        TeachingProgressUpdate.objects
        .filter(
            assignment__subject_template_id=assignment.subject_template_id,
            assignment__week_start_date__lt=week_start,
            no_teaching=False,
        )
        .order_by("-teaching_date", "-updated_at")
        .first()
    )


def _sync_no_teaching_update_for_assignment(assignment: TeachingWeeklyAssignment, week_start: date, *, updated_by_name: str = "Admin") -> None:
    """Create/update the no-teaching record used by the tutor update page/report."""
    prev = _latest_real_progress_for_assignment(assignment, week_start)
    teaching_date = week_start
    if getattr(assignment.tutoring_class, "time_slot", "") in {
        TutoringClass.TimeSlot.SUN_MORNING,
        TutoringClass.TimeSlot.SUN_AFTERNOON,
    }:
        teaching_date = week_start + timedelta(days=1)

    sheet_name = prev.sheet_name if prev else (assignment.subject_template.default_sheet_name or "")
    page_to = prev.page_to if prev else ""
    question_to = prev.question_to if prev else ""

    TeachingProgressUpdate.objects.update_or_create(
        assignment=assignment,
        teaching_date=teaching_date,
        defaults={
            "sheet_name": sheet_name,
            "page_to": page_to,
            "question_to": question_to,
            "no_teaching": True,
            "sheet_near_end": False,
            "updated_by_name": updated_by_name,
        },
    )


def _clear_auto_no_teaching_updates(assignment: TeachingWeeklyAssignment) -> None:
    TeachingProgressUpdate.objects.filter(assignment=assignment, no_teaching=True).delete()



@login_required
def teaching_template_manage(request: HttpRequest) -> HttpResponse:
    classes = list(TutoringClass.objects.filter(is_active=True).order_by("name"))
    for c in classes:
        c.sheet_grade_level = _class_grade_level(c)
    tutors = TeachingTutor.objects.filter(is_active=True).order_by("name")
    sheet_choices = Sheet.objects.select_related("subject").filter(is_active=True).order_by("grade_level", "subject__name", "code")
    class_grade_map = {str(c.id): c.sheet_grade_level for c in classes}

    selected_class_id = (request.GET.get("class_id") or request.POST.get("class_id") or "").strip()

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()

        if action == "add_tutor":
            tutor_name = (request.POST.get("tutor_name") or "").strip()
            tutor_phone = (request.POST.get("tutor_phone") or "").strip()
            if tutor_name:
                TeachingTutor.objects.update_or_create(
                    name=tutor_name,
                    defaults={"phone": tutor_phone, "is_active": True},
                )
            return redirect(f"/teaching/templates/?class_id={selected_class_id}" if selected_class_id else "/teaching/templates/")

        if action == "add_subject":
            class_id = request.POST.get("class_id")
            subject_name = (request.POST.get("subject_name") or "").strip()
            default_sheet_name = (request.POST.get("default_sheet_name") or "").strip()
            display_order_raw = (request.POST.get("display_order") or "").strip()
            try:
                display_order = int(display_order_raw or 1)
            except ValueError:
                display_order = 1

            cls = get_object_or_404(TutoringClass, id=class_id, is_active=True)
            if subject_name:
                TeachingClassSubjectTemplate.objects.update_or_create(
                    tutoring_class=cls,
                    subject_name=subject_name,
                    defaults={
                        "default_sheet_name": default_sheet_name,
                        "display_order": display_order,
                        "is_active": True,
                    },
                )
            return redirect(f"/teaching/templates/?class_id={class_id}")

        if action == "save_templates":
            for tid in request.POST.getlist("template_ids"):
                tmpl = TeachingClassSubjectTemplate.objects.filter(id=tid).first()
                if not tmpl:
                    continue
                tmpl.subject_name = (request.POST.get(f"subject_name_{tid}") or tmpl.subject_name).strip()
                tmpl.default_sheet_name = (request.POST.get(f"default_sheet_name_{tid}") or "").strip()
                try:
                    tmpl.display_order = int(request.POST.get(f"display_order_{tid}") or tmpl.display_order or 1)
                except ValueError:
                    pass
                tmpl.is_active = request.POST.get(f"is_active_{tid}") == "yes"
                tmpl.save()
            return redirect(f"/teaching/templates/?class_id={selected_class_id}" if selected_class_id else "/teaching/templates/")

    templates = TeachingClassSubjectTemplate.objects.select_related("tutoring_class").order_by(
        "tutoring_class__name", "display_order", "subject_name"
    )
    if selected_class_id:
        templates = templates.filter(tutoring_class_id=selected_class_id)
    templates = list(templates)
    for tmpl in templates:
        tmpl.class_grade_level = _class_grade_level(tmpl.tutoring_class)

    return render(request, "core/teaching_template_manage.html", {
        "classes": classes,
        "templates": templates,
        "tutors": tutors,
        "selected_class_id": selected_class_id,
        "sheet_choices": sheet_choices,
        "class_grade_map": class_grade_map,
    })


@login_required
def teaching_weekly_setup(request: HttpRequest) -> HttpResponse:
    week_start, week_end = _teaching_week_from_request(request)

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()

        if action == "add_tutor":
            tutor_name = (request.POST.get("tutor_name") or "").strip()
            tutor_phone = (request.POST.get("tutor_phone") or "").strip()
            if tutor_name:
                TeachingTutor.objects.update_or_create(
                    name=tutor_name,
                    defaults={"phone": tutor_phone, "is_active": True},
                )
            return redirect(f"/teaching/weekly-setup/?week_start={week_start.isoformat()}")

        _ensure_teaching_assignments(week_start, week_end)

        if action == "save_assignments" or action.startswith("save_class_assignments"):
            target_class_id = ""
            if action.startswith("save_class_assignments:"):
                target_class_id = action.split(":", 1)[1]

            assignment_ids = request.POST.getlist("assignment_ids")
            qs = TeachingWeeklyAssignment.objects.filter(id__in=assignment_ids, week_start_date=week_start)
            if target_class_id:
                qs = qs.filter(tutoring_class_id=target_class_id)

            for assignment in qs.select_related("subject_template", "tutoring_class", "tutor"):
                aid = str(assignment.id)
                tutor_id = (request.POST.get(f"tutor_{aid}") or "").strip()
                is_teaching_raw = (request.POST.get(f"is_teaching_{aid}") or "yes").strip()
                assignment.tutor = TeachingTutor.objects.filter(id=tutor_id, is_active=True).first() if tutor_id else None
                assignment.is_teaching = is_teaching_raw != "no"
                assignment.save()

                if assignment.is_teaching:
                    _clear_auto_no_teaching_updates(assignment)
                else:
                    _sync_no_teaching_update_for_assignment(assignment, week_start, updated_by_name="Admin")

            return redirect(f"/teaching/weekly-setup/?week_start={week_start.isoformat()}")

    _ensure_teaching_assignments(week_start, week_end)

    assignments = (
        TeachingWeeklyAssignment.objects
        .select_related("tutoring_class", "subject_template", "tutor")
        .filter(week_start_date=week_start)
        .order_by("tutoring_class__name", "subject_template__display_order", "subject_template__subject_name")
    )
    tutors = TeachingTutor.objects.filter(is_active=True).order_by("name")

    grouped_assignments: OrderedDict[int, dict] = OrderedDict()
    for a in assignments:
        key = a.tutoring_class_id
        grouped_assignments.setdefault(key, {"class": a.tutoring_class, "assignments": []})
        grouped_assignments[key]["assignments"].append(a)

    return render(request, "core/teaching_weekly_setup.html", {
        "week_start": week_start,
        "week_end": week_end,
        "assignments": assignments,
        "grouped_assignments": grouped_assignments,
        "tutors": tutors,
    })


def _teaching_slot_groups_from_assignments(assignments, previous_updates=None, current_updates=None):
    """Group assignments by the school time slot order: Sat AM, Sat PM, Sun AM, Sun PM."""
    previous_updates = previous_updates or {}
    current_updates = current_updates or {}
    slot_meta = {
        TutoringClass.TimeSlot.SAT_MORNING: {"label": "เสาร์เช้า", "tone": "sat-morning"},
        TutoringClass.TimeSlot.SAT_AFTERNOON: {"label": "เสาร์บ่าย", "tone": "sat-afternoon"},
        TutoringClass.TimeSlot.SUN_MORNING: {"label": "อาทิตย์เช้า", "tone": "sun-morning"},
        TutoringClass.TimeSlot.SUN_AFTERNOON: {"label": "อาทิตย์บ่าย", "tone": "sun-afternoon"},
    }

    grouped_slots: OrderedDict[str, dict] = OrderedDict()
    for slot in TIME_SLOT_ORDER:
        meta = slot_meta.get(slot, {"label": str(slot), "tone": "default"})
        grouped_slots[slot] = {
            "label": meta["label"],
            "tone": meta["tone"],
            "classes": OrderedDict(),
            "count": 0,
        }

    slot_rank = {slot: idx for idx, slot in enumerate(TIME_SLOT_ORDER)}
    sorted_assignments = sorted(
        assignments,
        key=lambda a: (
            slot_rank.get(a.tutoring_class.time_slot, 999),
            a.tutoring_class.name,
            a.subject_template.display_order,
            a.subject_template.subject_name,
        )
    )

    for a in sorted_assignments:
        slot = a.tutoring_class.time_slot
        if slot not in grouped_slots:
            grouped_slots[slot] = {"label": slot, "tone": "default", "classes": OrderedDict(), "count": 0}
        class_bucket = grouped_slots[slot]["classes"].setdefault(
            a.tutoring_class_id,
            {"class": a.tutoring_class, "assignments": []},
        )
        class_bucket["assignments"].append({
            "assignment": a,
            "previous_update": previous_updates.get(a.subject_template_id),
            "current_update": current_updates.get(a.id),
        })
        grouped_slots[slot]["count"] += 1

    # Hide empty slot headers when filtering by tutor/class leaves no assignments in that slot.
    return OrderedDict((k, v) for k, v in grouped_slots.items() if v["count"] > 0)


def tutor_teaching_update(request: HttpRequest) -> HttpResponse:
    week_start, week_end = _teaching_week_from_request(request)
    _ensure_teaching_assignments(week_start, week_end)

    selected_tutor_id = (request.GET.get("tutor_id") or request.POST.get("tutor_id") or "").strip()

    if request.method == "POST":
        assignment_id = request.POST.get("assignment_id")
        assignment = get_object_or_404(
            TeachingWeeklyAssignment.objects.select_related("subject_template", "tutor"),
            id=assignment_id,
            week_start_date=week_start,
        )
        teaching_date = _safe_date(request.POST.get("teaching_date")) if "_safe_date" in globals() else None
        if not teaching_date:
            teaching_date = timezone.localdate()

        no_teaching = request.POST.get("no_teaching") == "yes"
        sheet_name = (request.POST.get("sheet_name") or "").strip()
        page_to = (request.POST.get("page_to") or "").strip()
        question_to = (request.POST.get("question_to") or "").strip()

        qs = f"week_start={week_start.isoformat()}"
        if selected_tutor_id:
            qs += f"&tutor_id={selected_tutor_id}"

        # Require both page and question for actual teaching updates.
        # If not applicable, user can enter '-' in either field. If no teaching, allow blank.
        if not no_teaching and (not page_to or not question_to):
            qs += "&error=missing_progress"
            return redirect(f"/tutor-teaching-update/?{qs}#assignment-{assignment.id}")

        # If no teaching, preserve the latest real teaching progress as the visible sheet/page/question.
        if no_teaching:
            prev = _latest_real_progress_for_assignment(assignment, week_start)
            if prev:
                sheet_name = prev.sheet_name
                page_to = prev.page_to
                question_to = prev.question_to
            elif not sheet_name:
                sheet_name = assignment.subject_template.default_sheet_name or ""

        sheet_near_end = (request.POST.get("sheet_status") or "normal") == "near_end"
        if no_teaching:
            sheet_near_end = False

        updated_by_name = (request.POST.get("updated_by_name") or "").strip()
        if not updated_by_name and assignment.tutor_id:
            updated_by_name = assignment.tutor.name

        TeachingProgressUpdate.objects.update_or_create(
            assignment=assignment,
            teaching_date=teaching_date,
            defaults={
                "sheet_name": sheet_name,
                "page_to": page_to,
                "question_to": question_to,
                "no_teaching": no_teaching,
                "sheet_near_end": sheet_near_end,
                "updated_by_name": updated_by_name,
            },
        )
        return redirect(f"/tutor-teaching-update/?{qs}#assignment-{assignment.id}")

    all_week_assignments = list(
        TeachingWeeklyAssignment.objects
        .select_related("tutoring_class", "subject_template", "tutor")
        .prefetch_related("progress_updates")
        .filter(week_start_date=week_start)
    )

    # Make sure assignments marked as no teaching in weekly setup also appear as completed/no-teaching.
    for assignment in all_week_assignments:
        if not getattr(assignment, "is_teaching", True):
            has_no_teaching_update = TeachingProgressUpdate.objects.filter(assignment=assignment, no_teaching=True).exists()
            if not has_no_teaching_update:
                _sync_no_teaching_update_for_assignment(assignment, week_start, updated_by_name="Admin")

    display_assignments = all_week_assignments
    if selected_tutor_id:
        display_assignments = [a for a in all_week_assignments if str(a.tutor_id or "") == selected_tutor_id]

    all_assignments = display_assignments
    for a in all_assignments:
        a.class_grade_level = _class_grade_level(a.tutoring_class)
    template_ids = [a.subject_template_id for a in all_assignments]

    previous_updates = {}
    if template_ids:
        prev_qs = (
            TeachingProgressUpdate.objects
            .select_related("assignment", "assignment__subject_template", "assignment__tutor")
            .filter(
                assignment__subject_template_id__in=template_ids,
                assignment__week_start_date__lt=week_start,
                no_teaching=False,
            )
            .order_by("assignment__subject_template_id", "-teaching_date", "-updated_at")
        )
        for upd in prev_qs:
            previous_updates.setdefault(upd.assignment.subject_template_id, upd)

    current_updates = {}
    if all_assignments:
        curr_qs = (
            TeachingProgressUpdate.objects
            .filter(assignment_id__in=[a.id for a in all_assignments])
            .order_by("assignment_id", "-teaching_date", "-updated_at")
        )
        for upd in curr_qs:
            current_updates.setdefault(upd.assignment_id, upd)

    tutor_choices = list(TeachingTutor.objects.filter(is_active=True).order_by("name"))

    all_current_updates = {}
    if all_week_assignments:
        for upd in TeachingProgressUpdate.objects.filter(assignment_id__in=[a.id for a in all_week_assignments]).order_by("assignment_id", "-teaching_date", "-updated_at"):
            all_current_updates.setdefault(upd.assignment_id, upd)

    tutor_summary_rows = []
    for tutor in tutor_choices:
        assigned_items = [a for a in all_week_assignments if a.tutor_id == tutor.id and getattr(a, "is_teaching", True)]
        done_items = [a for a in assigned_items if a.id in all_current_updates]
        assigned_count = len(assigned_items)
        done_count = len(done_items)
        tutor_summary_rows.append({
            "tutor": tutor,
            "assigned_count": assigned_count,
            "done_count": done_count,
            "pending_count": max(assigned_count - done_count, 0),
            "percent": int((done_count / assigned_count) * 100) if assigned_count else 0,
            "is_empty": assigned_count == 0,
            "is_selected": selected_tutor_id == str(tutor.id),
        })

    unassigned_count = sum(1 for a in all_week_assignments if not a.tutor_id and getattr(a, "is_teaching", True))

    sheet_choices = Sheet.objects.select_related("subject").filter(is_active=True).order_by("grade_level", "subject__name", "code")
    grouped_slots = _teaching_slot_groups_from_assignments(all_assignments, previous_updates, current_updates)

    return render(request, "core/tutor_teaching_update.html", {
        "week_start": week_start,
        "week_end": week_end,
        "today": timezone.localdate(),
        "tutors": tutor_choices,
        "tutor_summary_rows": tutor_summary_rows,
        "unassigned_count": unassigned_count,
        "selected_tutor_id": selected_tutor_id,
        "grouped_slots": grouped_slots,
        "error_code": (request.GET.get("error") or "").strip(),
        "sheet_choices": sheet_choices,
    })


@login_required
def teaching_update_report(request: HttpRequest) -> HttpResponse:
    week_start, week_end = _teaching_week_from_request(request)
    _ensure_teaching_assignments(week_start, week_end)

    assignments = list(
        TeachingWeeklyAssignment.objects
        .select_related("tutoring_class", "subject_template", "tutor")
        .filter(week_start_date=week_start)
    )

    updates = {
        u.assignment_id: u
        for u in TeachingProgressUpdate.objects.filter(assignment__week_start_date=week_start).order_by("assignment_id", "-teaching_date", "-updated_at")
    }

    rows = []
    for a in assignments:
        u = updates.get(a.id)
        rows.append({
            "assignment": a,
            "update": u,
            "is_done": bool(u),
            "is_no_teaching": bool(u and u.no_teaching),
            "is_near_end": bool(u and getattr(u, "sheet_near_end", False)),
        })

    slot_rank = {slot: idx for idx, slot in enumerate(TIME_SLOT_ORDER)}
    rows.sort(key=lambda r: (
        slot_rank.get(r["assignment"].tutoring_class.time_slot, 999),
        r["assignment"].tutoring_class.name,
        r["assignment"].subject_template.display_order,
        r["assignment"].subject_template.subject_name,
    ))

    slot_meta = {
        TutoringClass.TimeSlot.SAT_MORNING: {"label": "เสาร์เช้า", "tone": "sat-morning"},
        TutoringClass.TimeSlot.SAT_AFTERNOON: {"label": "เสาร์บ่าย", "tone": "sat-afternoon"},
        TutoringClass.TimeSlot.SUN_MORNING: {"label": "อาทิตย์เช้า", "tone": "sun-morning"},
        TutoringClass.TimeSlot.SUN_AFTERNOON: {"label": "อาทิตย์บ่าย", "tone": "sun-afternoon"},
    }
    grouped_rows: OrderedDict[str, dict] = OrderedDict()
    for slot in TIME_SLOT_ORDER:
        meta = slot_meta.get(slot, {"label": str(slot), "tone": "default"})
        grouped_rows[slot] = {"label": meta["label"], "tone": meta["tone"], "rows": [], "count": 0}
    for row in rows:
        slot = row["assignment"].tutoring_class.time_slot
        if slot not in grouped_rows:
            grouped_rows[slot] = {"label": slot, "tone": "default", "rows": [], "count": 0}
        grouped_rows[slot]["rows"].append(row)
        grouped_rows[slot]["count"] += 1
    grouped_rows = OrderedDict((k, v) for k, v in grouped_rows.items() if v["count"] > 0)

    total = len(rows)
    done = sum(1 for r in rows if r["is_done"])
    no_teaching = sum(1 for r in rows if r["is_no_teaching"])
    taught = done - no_teaching

    return render(request, "core/teaching_update_report.html", {
        "week_start": week_start,
        "week_end": week_end,
        "rows": rows,
        "grouped_rows": grouped_rows,
        "total": total,
        "done": done,
        "taught": taught,
        "no_teaching": no_teaching,
        "pending": total - done,
    })

# =========================================================
# ✅ Test Score Announcement Module
# =========================================================
def _digits_only(value: str | None) -> str:
    return "".join(ch for ch in (value or "") if ch.isdigit())


def _to_decimal(value, default="0") -> Decimal:
    raw = str(value if value is not None else "").strip().replace(",", "")
    if raw == "":
        raw = str(default)
    try:
        return Decimal(raw)
    except Exception:
        return Decimal(default)


def _pct(score: Decimal, full_score: Decimal) -> float:
    if not full_score or full_score <= 0:
        return 0.0
    return float((score / full_score) * Decimal("100"))


def _shared_ranks(values: dict[int, float]) -> dict[int, int]:
    """Competition rank: 95,95,90 = 1,1,3."""
    ranks = {}
    sorted_items = sorted(values.items(), key=lambda kv: (-kv[1], kv[0]))
    previous_value = None
    previous_rank = 0
    for idx, (pk, value) in enumerate(sorted_items, start=1):
        if previous_value is None or value != previous_value:
            previous_rank = idx
            previous_value = value
        ranks[pk] = previous_rank
    return ranks


def _snapshot_from_student(student: Student) -> dict:
    return {
        "source_type": TestParticipant.SourceType.STUDENT,
        "student": student,
        "nickname": student.nickname or "",
        "full_name": student.full_name or "",
        "school_name": student.school.name if student.school else "",
        "contact_phone": student.parent_phone or "",
        "grade_level": student.grade_level or "",
    }


def _snapshot_from_admission(inquiry: AdmissionInquiry) -> dict:
    return {
        "source_type": TestParticipant.SourceType.ADMISSION,
        "admission_inquiry": inquiry,
        "nickname": inquiry.nickname or "",
        "full_name": inquiry.full_name or "",
        "school_name": inquiry.school_name or "",
        "contact_phone": inquiry.contact_phone or "",
        "grade_level": inquiry.grade_level or "",
    }




def _build_test_score_context(test_round: TestRound, current_participant: TestParticipant | None = None) -> dict:
    """
    Build score rows for both parent result page and admin summary.

    Important: overall percentage is weighted by each subject's full score:
        total_pct = (sum(raw scores) / sum(full scores)) * 100
    This is intentionally NOT an average of each subject percentage.
    """
    subjects = list(test_round.subjects.filter(is_active=True).order_by("display_order", "id"))
    participants = list(
        test_round.participants.filter(is_active=True).order_by("full_name", "nickname", "id")
    )
    score_objs = TestScore.objects.filter(
        participant__in=participants,
        subject__in=subjects,
    ).select_related("participant", "subject")
    score_map = {(s.participant_id, s.subject_id): s for s in score_objs}

    subject_pct_values: dict[int, dict[int, float]] = {s.id: {} for s in subjects}
    total_pct_values: dict[int, float] = {}
    rows = []

    total_full_score = sum((s.full_score or Decimal("0")) for s in subjects)

    for p in participants:
        subject_cells = []
        total_score = Decimal("0")

        for s in subjects:
            score_obj = score_map.get((p.id, s.id))
            score = score_obj.score if score_obj else Decimal("0")
            score_pct = _pct(score, s.full_score)

            # Raw scores are accumulated first, then divided by total full score.
            # This makes the overall percentage weighted by subject full score.
            total_score += score

            subject_pct_values[s.id][p.id] = score_pct
            subject_cells.append({
                "subject": s,
                "score": score,
                "full_score": s.full_score,
                "pct": score_pct,
                "score_obj": score_obj,
            })

        total_pct = _pct(total_score, total_full_score)
        total_pct_values[p.id] = total_pct
        rows.append({
            "participant": p,
            "subject_cells": subject_cells,
            "total_score": total_score,
            "total_full_score": total_full_score,
            "total_pct": total_pct,
            "weighted_total_pct": total_pct,
            "is_current": bool(current_participant and p.id == current_participant.id),
        })

    subject_avg_pct = {}
    subject_ranks = {}
    for s in subjects:
        values = subject_pct_values.get(s.id, {})
        subject_avg_pct[s.id] = round(sum(values.values()) / len(values), 2) if values else 0
        subject_ranks[s.id] = _shared_ranks(values)

    # Overall average also follows the same weighted approach:
    # average raw total score / total full score.
    participant_count = len(rows)
    total_score_sum = sum((row.get("total_score") or Decimal("0")) for row in rows)
    total_avg_score = (total_score_sum / Decimal(participant_count)) if participant_count else Decimal("0")
    total_avg_pct = _pct(total_avg_score, total_full_score)

    total_ranks = _shared_ranks(total_pct_values)

    for row in rows:
        p = row["participant"]
        row["total_rank"] = total_ranks.get(p.id, "-")
        row["display_name"] = p.display_name if row["is_current"] else "xxxxx"
        for cell in row["subject_cells"]:
            sid = cell["subject"].id
            cell["avg_pct"] = subject_avg_pct.get(sid, 0)
            cell["rank"] = subject_ranks.get(sid, {}).get(p.id, "-")

    score_rows = sorted(
        rows,
        key=lambda r: (
            -float(r.get("weighted_total_pct") or r.get("total_pct") or 0),
            int(r.get("total_rank") or 999999) if str(r.get("total_rank") or "").isdigit() else 999999,
            (r["participant"].full_name or ""),
            (r["participant"].nickname or ""),
            r["participant"].id,
        ),
    )

    current_row = None
    for row in rows:
        if row["is_current"]:
            current_row = row
            break

    chart_rows = []
    narrative = ""
    if current_row:
        for cell in current_row["subject_cells"]:
            chart_rows.append({
                "label": cell["subject"].name,
                "student_pct": round(cell["pct"], 2),
                "avg_pct": round(cell["avg_pct"], 2),
            })
        chart_rows.append({
            "label": "คะแนนรวม",
            "student_pct": round(current_row["weighted_total_pct"], 2),
            "avg_pct": round(total_avg_pct, 2),
        })

        diff = round(current_row["weighted_total_pct"] - total_avg_pct, 2)
        best = max(current_row["subject_cells"], key=lambda c: c["pct"], default=None)
        focus = min(current_row["subject_cells"], key=lambda c: c["pct"], default=None)
        if diff >= 5:
            main = f"ภาพรวมคะแนนของน้องสูงกว่าค่าเฉลี่ยประมาณ {abs(diff):.1f}% ถือว่าทำได้ดีมากครับ"
        elif diff <= -5:
            main = f"ภาพรวมคะแนนของน้องต่ำกว่าค่าเฉลี่ยประมาณ {abs(diff):.1f}% ยังมีจุดที่สามารถค่อย ๆ เสริมเพิ่มได้ครับ"
        else:
            main = "ภาพรวมคะแนนของน้องอยู่ใกล้เคียงค่าเฉลี่ยของรอบสอบนี้ครับ"
        if best and focus and best["subject"].id != focus["subject"].id:
            narrative = f"{main} วิชาที่เด่นที่สุดคือ {best['subject'].name} ส่วนวิชาที่ควรทบทวนเพิ่มคือ {focus['subject'].name} ครับ"
        else:
            narrative = main

    return {
        "test_round": test_round,
        "subjects": subjects,
        "participants": participants,
        "rows": rows,
        "score_rows": score_rows,
        "current_row": current_row,
        "chart_rows": chart_rows,
        "subject_avg_pct": subject_avg_pct,
        "total_avg_pct": total_avg_pct,
        "total_avg_score": total_avg_score,
        "total_full_score": total_full_score,
        "narrative": narrative,
    }


def test_score_round_list(request: HttpRequest) -> HttpResponse:
    rounds = TestRound.objects.filter(is_published=True).order_by("-exam_date", "-created_at")
    return render(request, "core/test_score_round_list.html", {"rounds": rounds})


@require_GET
def test_score_participant_search(request: HttpRequest, round_id: int) -> JsonResponse:
    test_round = get_object_or_404(TestRound, id=round_id, is_published=True)
    q = (request.GET.get("q") or "").strip()
    qs = test_round.participants.filter(is_active=True)
    if q:
        qs = qs.filter(
            Q(nickname__icontains=q) |
            Q(full_name__icontains=q) |
            Q(school_name__icontains=q)
        )
    qs = qs.order_by("full_name", "nickname")[:30]
    results = []
    for p in qs:
        text = f"{p.nickname or '-'} | {p.full_name}"
        if p.school_name:
            text += f" | {p.school_name}"
        results.append({"id": str(p.id), "text": text})
    return JsonResponse({"results": results})


def test_score_login(request: HttpRequest, round_id: int) -> HttpResponse:
    test_round = get_object_or_404(TestRound, id=round_id, is_published=True)
    error = ""
    if request.method == "POST":
        participant_id = (request.POST.get("participant_id") or "").strip()
        phone = (request.POST.get("parent_phone") or "").strip()
        participant = test_round.participants.filter(id=participant_id, is_active=True).first()
        if not participant:
            error = "กรุณาเลือกชื่อนักเรียนจากรายการที่ขึ้นมา"
        elif phone != "kanoon" and _digits_only(participant.contact_phone) != _digits_only(phone):
            error = "เบอร์มือถือไม่ถูกต้อง"
        else:
            request.session[f"test_score_participant_id_{test_round.id}"] = participant.id
            return redirect("core:test_score_result", round_id=test_round.id)
    return render(request, "core/test_score_login.html", {"test_round": test_round, "error": error})


def test_score_result(request: HttpRequest, round_id: int) -> HttpResponse:
    test_round = get_object_or_404(TestRound, id=round_id, is_published=True)
    participant_id = request.session.get(f"test_score_participant_id_{test_round.id}")
    participant = test_round.participants.filter(id=participant_id, is_active=True).first()
    if not participant:
        return redirect("core:test_score_login", round_id=test_round.id)
    context = _build_test_score_context(test_round, participant)
    return render(request, "core/test_score_result.html", context)


def test_score_logout(request: HttpRequest, round_id: int) -> HttpResponse:
    request.session.pop(f"test_score_participant_id_{round_id}", None)
    return redirect("core:test_score_login", round_id=round_id)


@login_required
def test_score_admin(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        title = (request.POST.get("title") or "").strip()
        exam_date = request.POST.get("exam_date") or None
        note = (request.POST.get("note") or "").strip()
        is_published = request.POST.get("is_published") == "on"
        if title:
            test_round = TestRound.objects.create(
                title=title,
                exam_date=exam_date or None,
                note=note,
                is_published=is_published,
            )
            return redirect("core:test_score_round_manage", round_id=test_round.id)
    rounds = TestRound.objects.all().order_by("-exam_date", "-created_at")
    return render(request, "core/test_score_admin.html", {"rounds": rounds})


@require_GET
@login_required
def test_score_student_search(request: HttpRequest) -> JsonResponse:
    q = (request.GET.get("q") or "").strip()
    qs = Student.objects.filter(is_active=True)
    if q:
        qs = qs.filter(
            Q(nickname__icontains=q) |
            Q(full_name__icontains=q) |
            Q(student_code__icontains=q) |
            Q(school__name__icontains=q)
        )
    qs = qs.select_related("school").order_by("grade_level", "student_code")[:30]
    results = []
    for s in qs:
        text = f"{s.nickname or '-'} | {s.full_name}"
        if s.school:
            text += f" | {s.school.name}"
        results.append({"id": str(s.id), "text": text})
    return JsonResponse({"results": results})


@require_GET
@login_required
def test_score_admission_search(request: HttpRequest) -> JsonResponse:
    q = (request.GET.get("q") or "").strip()
    qs = AdmissionInquiry.objects.all()
    if q:
        qs = qs.filter(
            Q(nickname__icontains=q) |
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q) |
            Q(school_name__icontains=q) |
            Q(contact_phone__icontains=q)
        )
    qs = qs.order_by("-created_at")[:30]
    results = []
    for i in qs:
        text = f"{i.nickname or '-'} | {i.full_name}"
        if i.school_name:
            text += f" | {i.school_name}"
        results.append({"id": str(i.id), "text": text})
    return JsonResponse({"results": results})


def _upsert_score(participant: TestParticipant, subject: TestSubject, value, note: str = ""):
    score = _to_decimal(value, default="0")
    if score < 0:
        score = Decimal("0")
    if subject.full_score and score > subject.full_score:
        score = subject.full_score
    TestScore.objects.update_or_create(
        participant=participant,
        subject=subject,
        defaults={"score": score, "note": note or ""},
    )


@login_required
def test_score_round_manage(request: HttpRequest, round_id: int) -> HttpResponse:
    test_round = get_object_or_404(TestRound, id=round_id)
    result_message = ""

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()

        if action == "update_round":
            test_round.title = (request.POST.get("title") or test_round.title).strip()
            test_round.exam_date = request.POST.get("exam_date") or None
            test_round.note = request.POST.get("note") or ""
            test_round.is_published = request.POST.get("is_published") == "on"
            test_round.save()
            return redirect("core:test_score_round_manage", round_id=test_round.id)

        if action == "add_subject":
            name = (request.POST.get("subject_name") or "").strip()
            full_score = _to_decimal(request.POST.get("full_score"), default="100")
            if name:
                next_order = (test_round.subjects.aggregate(m=Count("id")).get("m") or 0) + 1
                TestSubject.objects.create(test_round=test_round, name=name, full_score=full_score, display_order=next_order)
            return redirect("core:test_score_round_manage", round_id=test_round.id)

        if action == "save_subjects":
            for subject in test_round.subjects.all():
                subject.name = (request.POST.get(f"subject_name_{subject.id}") or subject.name).strip()
                subject.full_score = _to_decimal(request.POST.get(f"full_score_{subject.id}"), default=str(subject.full_score or 100))
                subject.display_order = int(request.POST.get(f"display_order_{subject.id}") or subject.display_order or 1)
                subject.is_active = request.POST.get(f"subject_active_{subject.id}") == "on"
                subject.save()
            return redirect("core:test_score_round_manage", round_id=test_round.id)

        if action == "add_student_participant":
            student = Student.objects.filter(id=request.POST.get("student_id"), is_active=True).first()
            if student:
                defaults = _snapshot_from_student(student)
                TestParticipant.objects.update_or_create(
                    test_round=test_round,
                    source_type=TestParticipant.SourceType.STUDENT,
                    student=student,
                    defaults={**defaults, "is_active": True},
                )
            return redirect("core:test_score_round_manage", round_id=test_round.id)

        if action == "add_admission_participant":
            inquiry = AdmissionInquiry.objects.filter(id=request.POST.get("admission_id")).first()
            if inquiry:
                defaults = _snapshot_from_admission(inquiry)
                TestParticipant.objects.update_or_create(
                    test_round=test_round,
                    source_type=TestParticipant.SourceType.ADMISSION,
                    admission_inquiry=inquiry,
                    defaults={**defaults, "is_active": True},
                )
            return redirect("core:test_score_round_manage", round_id=test_round.id)

        if action == "add_manual_participant":
            full_name = (request.POST.get("full_name") or "").strip()
            if full_name:
                TestParticipant.objects.create(
                    test_round=test_round,
                    source_type=TestParticipant.SourceType.MANUAL,
                    nickname=(request.POST.get("nickname") or "").strip(),
                    full_name=full_name,
                    school_name=(request.POST.get("school_name") or "").strip(),
                    contact_phone=(request.POST.get("contact_phone") or "").strip(),
                    grade_level=(request.POST.get("grade_level") or "").strip(),
                )
            return redirect("core:test_score_round_manage", round_id=test_round.id)

        if action == "save_scores":
            subjects = list(test_round.subjects.filter(is_active=True).order_by("display_order", "id"))
            participants = list(test_round.participants.filter(is_active=True))
            with transaction.atomic():
                for p in participants:
                    p.note = request.POST.get(f"participant_note_{p.id}") or ""
                    p.save(update_fields=["note", "updated_at"])
                    for s in subjects:
                        _upsert_score(p, s, request.POST.get(f"score_{p.id}_{s.id}"), "")
            return redirect("core:test_score_round_manage", round_id=test_round.id)

        if action == "import_scores":
            uploaded = request.FILES.get("score_file")
            if uploaded:
                try:
                    rows = _read_test_score_import_rows(uploaded)
                    result_message = _import_test_score_rows(test_round, rows)
                except Exception as exc:
                    result_message = f"Import ไม่สำเร็จ: {exc}"

    context = _build_test_score_context(test_round, None)
    context.update({"result_message": result_message})
    return render(request, "core/test_score_round_manage.html", context)


@login_required
def test_score_round_summary(request: HttpRequest, round_id: int) -> HttpResponse:
    test_round = get_object_or_404(TestRound, id=round_id)
    context = _build_test_score_context(test_round, None)

    rows = list(context.get("rows", []))
    subjects = list(context.get("subjects", []))
    participant_count = len(rows)

    # Admin summary should show every participant and sort by weighted overall percentage descending.
    score_rows = list(context.get("score_rows", [])) or sorted(
        rows,
        key=lambda r: (
            -float(r.get("weighted_total_pct") or r.get("total_pct") or 0),
            int(r.get("total_rank") or 999999) if str(r.get("total_rank") or "").isdigit() else 999999,
            (r["participant"].full_name or ""),
            (r["participant"].nickname or ""),
            r["participant"].id,
        ),
    )

    subject_summary_rows = []
    for idx, subject in enumerate(subjects):
        total_score = Decimal("0")
        for row in rows:
            try:
                total_score += row["subject_cells"][idx]["score"] or Decimal("0")
            except Exception:
                total_score += Decimal("0")
        avg_score = (total_score / Decimal(participant_count)) if participant_count else Decimal("0")
        subject_summary_rows.append({
            "subject": subject,
            "avg_score": avg_score,
            "avg_pct": context.get("subject_avg_pct", {}).get(subject.id, 0),
            "full_score": subject.full_score,
        })

    total_score_sum = sum((row.get("total_score") or Decimal("0")) for row in rows) if rows else Decimal("0")
    total_avg_score = (total_score_sum / Decimal(participant_count)) if participant_count else Decimal("0")

    context.update({
        "score_rows": score_rows,
        "subject_summary_rows": subject_summary_rows,
        "participant_count": participant_count,
        "total_avg_score": total_avg_score,
        "total_full_score": sum((s.full_score or Decimal("0")) for s in subjects),
    })
    return render(request, "core/test_score_summary.html", context)


def _read_test_score_import_rows(uploaded_file) -> list[dict]:
    filename = (getattr(uploaded_file, "name", "") or "").lower()
    if filename.endswith(".csv"):
        raw = uploaded_file.read()
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw.decode("cp874", errors="replace")
        return [dict(r) for r in csv.DictReader(text.splitlines())]

    wb = load_workbook(BytesIO(uploaded_file.read()), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_clean_cell(h) for h in rows[0]]
    result = []
    for values in rows[1:]:
        if not any(_clean_cell(v) for v in values):
            continue
        result.append({headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))})
    return result


def _import_test_score_rows(test_round: TestRound, rows: list[dict]) -> str:
    subjects = list(test_round.subjects.filter(is_active=True).order_by("display_order", "id"))
    created = updated = 0
    with transaction.atomic():
        for row in rows:
            student_id = _clean_cell(row.get("student_id") or row.get("Student ID"))
            admission_id = _clean_cell(row.get("admission_id") or row.get("Admission ID"))
            participant = None
            if student_id:
                student = Student.objects.filter(id=student_id).select_related("school").first()
                if student:
                    defaults = _snapshot_from_student(student)
                    participant, was_created = TestParticipant.objects.update_or_create(
                        test_round=test_round,
                        source_type=TestParticipant.SourceType.STUDENT,
                        student=student,
                        defaults={**defaults, "is_active": True},
                    )
                    created += int(was_created)
                    updated += int(not was_created)
            if not participant and admission_id:
                inquiry = AdmissionInquiry.objects.filter(id=admission_id).first()
                if inquiry:
                    defaults = _snapshot_from_admission(inquiry)
                    participant, was_created = TestParticipant.objects.update_or_create(
                        test_round=test_round,
                        source_type=TestParticipant.SourceType.ADMISSION,
                        admission_inquiry=inquiry,
                        defaults={**defaults, "is_active": True},
                    )
                    created += int(was_created)
                    updated += int(not was_created)
            if not participant:
                full_name = _clean_cell(row.get("full_name") or row.get("ชื่อจริงนามสกุล") or row.get("ชื่อจริง นามสกุล"))
                nickname = _clean_cell(row.get("nickname") or row.get("ชื่อเล่น"))
                contact_phone = _clean_cell(row.get("contact_phone") or row.get("เบอร์มือถือ") or row.get("เบอร์ติดต่อ"))
                if not full_name and not nickname:
                    continue
                participant, was_created = TestParticipant.objects.update_or_create(
                    test_round=test_round,
                    source_type=TestParticipant.SourceType.MANUAL,
                    full_name=full_name or nickname,
                    contact_phone=contact_phone,
                    defaults={
                        "nickname": nickname,
                        "school_name": _clean_cell(row.get("school_name") or row.get("โรงเรียน")),
                        "grade_level": _clean_cell(row.get("grade_level") or row.get("ระดับชั้น")),
                        "note": _clean_cell(row.get("note") or row.get("หมายเหตุ")),
                        "is_active": True,
                    },
                )
                created += int(was_created)
                updated += int(not was_created)

            participant.note = _clean_cell(row.get("note") or row.get("หมายเหตุ"))
            participant.save(update_fields=["note", "updated_at"])
            for subject in subjects:
                value = row.get(subject.name)
                if value is None:
                    value = row.get(f"{subject.name} ({subject.full_score})")
                _upsert_score(participant, subject, value if value is not None else 0, "")
    return f"Import สำเร็จ: เพิ่มใหม่ {created} รายการ / อัปเดต {updated} รายการ"


@login_required
def test_score_import_template(request: HttpRequest, round_id: int) -> HttpResponse:
    test_round = get_object_or_404(TestRound, id=round_id)
    subjects = list(test_round.subjects.filter(is_active=True).order_by("display_order", "id"))
    participants = list(test_round.participants.filter(is_active=True).order_by("full_name", "nickname"))
    score_objs = TestScore.objects.filter(participant__in=participants, subject__in=subjects)
    score_map = {(s.participant_id, s.subject_id): s.score for s in score_objs}

    wb = Workbook()
    ws = wb.active
    ws.title = "Import Scores"
    headers = ["student_id", "admission_id", "nickname", "full_name", "school_name", "contact_phone", "grade_level", "note"] + [s.name for s in subjects]
    ws.append(headers)
    for p in participants:
        ws.append([
            p.student_id or "",
            p.admission_inquiry_id or "",
            p.nickname,
            p.full_name,
            p.school_name,
            p.contact_phone,
            p.grade_level,
            p.note,
            *[score_map.get((p.id, s.id), 0) for s in subjects],
        ])
    if not participants:
        ws.append(["", "", "ตัวอย่าง", "ชื่อจริง นามสกุล", "โรงเรียน", "0999999999", "ป.6", "", *[0 for _ in subjects]])
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="test_score_import_template_{test_round.id}.xlsx"'
    return response

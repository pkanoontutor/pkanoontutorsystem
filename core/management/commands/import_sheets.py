from pathlib import Path
import csv
from openpyxl import load_workbook

from django.core.management.base import BaseCommand
from django.db import transaction

from core.models import Sheet, Subject


class Command(BaseCommand):
    help = "Import Sheet data from CSV/XLSX (no pandas)"

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            type=str,
            required=True,
            help="Path to .csv or .xlsx file",
        )

    def _resolve_path(self, raw_path: str) -> Path:
        """
        Resolve path relative to current working dir (where manage.py is run),
        and expand user/home if needed.
        """
        p = Path(raw_path).expanduser()
        if p.is_absolute():
            return p
        return (Path.cwd() / p).resolve()

    def _read_rows_from_csv(self, file_path: Path) -> list[dict]:
        rows = []
        with file_path.open(newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            required = {"code", "title", "subject"}
            headers = set(reader.fieldnames or [])
            if not required.issubset(headers):
                raise ValueError(f"CSV ต้องมี header: {', '.join(sorted(required))} (เจอ: {reader.fieldnames})")
            for r in reader:
                rows.append(r)
        return rows

    def _read_rows_from_xlsx(self, file_path: Path) -> list[dict]:
        wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)
        ws = wb.active  # ใช้ชีทแรก

        # อ่าน header แถวแรก
        header = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
            header.append((str(cell).strip() if cell is not None else ""))

        required = {"code", "title", "subject"}
        header_set = set(header)
        if not required.issubset(header_set):
            raise ValueError(f"XLSX ต้องมี header: {', '.join(sorted(required))} (เจอ: {header})")

        idx = {name: header.index(name) for name in required}

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            def get_col(col_name: str) -> str:
                v = row[idx[col_name]] if idx[col_name] < len(row) else None
                return (str(v).strip() if v is not None else "")

            rows.append({
                "code": get_col("code"),
                "title": get_col("title"),
                "subject": get_col("subject"),
            })
        return rows

    def handle(self, *args, **options):
        file_path = self._resolve_path(options["file"])

        if not file_path.exists():
            self.stderr.write(f"❌ ไม่พบไฟล์: {file_path}")
            self.stderr.write("   👉 เช็คด้วย: ls -la และ ls -la data/")
            return

        suffix = file_path.suffix.lower()
        if suffix not in {".csv", ".xlsx"}:
            self.stderr.write("❌ รองรับเฉพาะไฟล์ .csv หรือ .xlsx")
            return

        try:
            if suffix == ".csv":
                rows = self._read_rows_from_csv(file_path)
            else:
                rows = self._read_rows_from_xlsx(file_path)
        except Exception as e:
            self.stderr.write(f"❌ อ่านไฟล์ไม่สำเร็จ: {e}")
            return

        created = 0
        updated = 0
        skipped = 0
        missing_subject = 0

        with transaction.atomic():
            for r in rows:
                code = (r.get("code") or "").strip()
                title = (r.get("title") or "").strip()
                subject_name = (r.get("subject") or "").strip()

                if not code or not title or not subject_name:
                    skipped += 1
                    continue

                subject = Subject.objects.filter(name=subject_name, is_active=True).first()
                if not subject:
                    self.stderr.write(f"⚠️ ไม่พบ subject '{subject_name}' (code={code})")
                    missing_subject += 1
                    continue

                _, is_created = Sheet.objects.update_or_create(
                    code=code,
                    defaults={
                        "title": title,
                        "subject": subject,
                        "is_active": True,
                    },
                )

                if is_created:
                    created += 1
                else:
                    updated += 1

        self.stdout.write("✅ Import completed")
        self.stdout.write(f"  ➕ created: {created}")
        self.stdout.write(f"  ♻️ updated: {updated}")
        self.stdout.write(f"  ⏭ skipped (blank row): {skipped}")
        self.stdout.write(f"  ⚠️ missing subject: {missing_subject}")
